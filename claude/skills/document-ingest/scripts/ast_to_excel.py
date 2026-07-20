#!/usr/bin/env python3
"""Apply a modification plan to an Excel file, writing back only changed cells.

Usage:
    python ast_to_excel.py <input.xlsx> --plan <modifications.json>
                           [-o output.xlsx] [--dry-run] [--backup] [--force]

The modification plan JSON:
{
  "source_file": "original.xlsx",
  "changes": [
    {"sheet": "Sheet1", "cell": "E12", "new_value": 45.0, "old_value": 42.5},
    {"sheet": "Sheet1", "cell": "F12", "new_value": "=E12*1.15", "old_value": null}
  ]
}

Each change entry:
  - sheet (required): target sheet name
  - cell  (required): target cell coordinate, e.g. "E12"
  - new_value (required): new value or formula string
  - old_value (optional): expected current value, for validation
  - reason   (optional): human-readable reason for the change
"""

import json
import sys
import argparse
import shutil
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.utils import coordinate_to_tuple, get_column_letter
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")


def _col_idx(coord):
    """'E12' -> column index (5)."""
    return coordinate_to_tuple(coord)[1]


def _row_idx(coord):
    """'E12' -> row (12)."""
    return coordinate_to_tuple(coord)[0]


class PlanValidator:
    """Validate a modification plan before applying."""

    def __init__(self, wb, plan):
        self.wb = wb
        self.plan = plan
        self.errors = []
        self.warnings = []

    def validate(self):
        changes = self.plan.get("changes", [])
        if not changes:
            self.errors.append("Modification plan contains no changes")
            return self.errors, self.warnings

        sheet_names = set(self.wb.sheetnames)

        for i, ch in enumerate(changes):
            self._validate_change(i, ch, sheet_names)

        self._check_duplicate_cells(changes)
        self._check_merged_cell_safety(changes)

        return self.errors, self.warnings

    def _validate_change(self, idx, ch, sheet_names):
        sheet = ch.get("sheet")
        cell = ch.get("cell")
        new_val = ch.get("new_value")

        if not sheet:
            self.errors.append(f"Change [{idx}]: missing 'sheet' field")
            return
        if sheet not in sheet_names:
            self.errors.append(
                f"Change [{idx}]: sheet '{sheet}' not found. "
                f"Available: {sorted(sheet_names)}")
            return
        if not cell:
            self.errors.append(
                f"Change [{idx}]: missing 'cell' coordinate")
            return
        if new_val is None and "new_value" not in ch:
            self.errors.append(
                f"Change [{idx}]: missing 'new_value'")

        ws = self.wb[sheet]
        try:
            existing = ws[cell]
        except Exception:
            self.errors.append(f"Change [{idx}]: invalid cell '{cell}'")
            return

        existing_val = existing.value
        if isinstance(existing_val, str) and existing_val.startswith("="):
            if not (isinstance(new_val, str) and new_val.startswith("=")):
                self.warnings.append(
                    f"Change [{idx}] {sheet}!{cell}: overwriting FORMULA "
                    f"'{existing_val}' with non-formula value '{new_val}'. "
                    f"Use --force to proceed.")

        old_val = ch.get("old_value")
        if old_val is not None and old_val != existing_val:
            self.warnings.append(
                f"Change [{idx}] {sheet}!{cell}: old_value mismatch. "
                f"Expected '{old_val}', actual '{existing_val}'")

    def _check_duplicate_cells(self, changes):
        seen = defaultdict(list)
        for i, ch in enumerate(changes):
            key = (ch.get("sheet"), ch.get("cell"))
            seen[key].append(i)
        for key, indices in seen.items():
            if len(indices) > 1:
                self.warnings.append(
                    f"Cell {key[0]}!{key[1]} appears in {len(indices)} "
                    f"changes (indices {indices}) — last write wins")

    def _check_merged_cell_safety(self, changes):
        for i, ch in enumerate(changes):
            sheet = ch.get("sheet")
            cell = ch.get("cell")
            if not sheet or not cell:
                continue
            ws = self.wb[sheet]
            for mr in ws.merged_cells.ranges:
                row, col = _row_idx(cell), _col_idx(cell)
                if (mr.min_row <= row <= mr.max_row and
                        mr.min_col <= col <= mr.max_col):
                    top_left = f"{get_column_letter(mr.min_col)}{mr.min_row}"
                    if cell != top_left:
                        self.warnings.append(
                            f"Change [{i}] {sheet}!{cell}: cell is in merged "
                            f"region {mr}. Writing to top-left {top_left} "
                            f"instead.")


class DryRunReporter:
    """Preview changes without writing."""

    @staticmethod
    def report(wb, plan, output_path):
        changes = plan.get("changes", [])
        print(f"DRY RUN — {len(changes)} change(s) would be applied")
        print(f"Source: {plan.get('source_file', 'N/A')}")
        print(f"Output: {output_path}")
        print("-" * 60)
        for i, ch in enumerate(changes):
            sheet = ch.get("sheet", "?")
            cell = ch.get("cell", "?")
            old = ch.get("old_value", "?")
            new = ch.get("new_value", "?")
            reason = ch.get("reason", "")
            existing = None
            try:
                existing = wb[sheet][cell].value
            except Exception:
                pass
            print(f"[{i}] {sheet}!{cell}")
            print(f"    Current: {existing}")
            print(f"    New:     {new}")
            if reason:
                print(f"    Reason:  {reason}")
        print("-" * 60)
        print("Use without --dry-run to apply changes.")


class SafetyGuard:
    """Post-write safety checks."""

    def __init__(self, wb, plan, orig_formula_count, orig_merge_count):
        self.wb = wb
        self.plan = plan
        self.orig_formula_count = orig_formula_count
        self.orig_merge_count = orig_merge_count

    def check(self):
        issues = []
        changes = self.plan.get("changes", [])

        # Formula loss check
        current_formulas = 0
        for sn in self.wb.sheetnames:
            ws = self.wb[sn]
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if isinstance(v, str) and v.startswith("="):
                        current_formulas += 1
        if current_formulas < self.orig_formula_count:
            issues.append(
                f"Formula count decreased: {self.orig_formula_count} -> "
                f"{current_formulas}")

        # Merged cell check
        current_merges = sum(
            len(self.wb[sn].merged_cells.ranges)
            for sn in self.wb.sheetnames)
        if current_merges < self.orig_merge_count:
            issues.append(
                f"Merged cell count decreased: {self.orig_merge_count} -> "
                f"{current_merges}")

        # Change ratio check
        total_cells = 0
        for sn in self.wb.sheetnames:
            ws = self.wb[sn]
            total_cells += (ws.max_row or 0) * (ws.max_column or 0)
        if total_cells > 0 and len(changes) / total_cells > 0.05 and len(changes) > 10:
            issues.append(
                f"Large change ratio: {len(changes)}/{total_cells} "
                f"({len(changes)/total_cells*100:.1f}%) cells modified. "
                f"Use --force to proceed.")

        return issues


class SurgicalWriter:
    """Apply changes to specific cells, preserving everything else."""

    def __init__(self, wb, plan):
        self.wb = wb
        self.plan = plan
        self.stats = {"cells_written": 0, "cells_skipped": 0}

    def apply(self):
        changes = self.plan.get("changes", [])
        for ch in changes:
            sheet = ch.get("sheet")
            cell = ch.get("cell")
            new_val = ch.get("new_value")
            if not sheet or not cell:
                self.stats["cells_skipped"] += 1
                continue

            ws = self.wb[sheet]
            target_cell = cell

            for mr in ws.merged_cells.ranges:
                row, col = _row_idx(cell), _col_idx(cell)
                if (mr.min_row <= row <= mr.max_row and
                        mr.min_col <= col <= mr.max_col):
                    target_cell = (f"{get_column_letter(mr.min_col)}"
                                   f"{mr.min_row}")
                    break

            ws[target_cell] = new_val
            self.stats["cells_written"] += 1

        return self.stats


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main():
    p = argparse.ArgumentParser(
        description="Apply modification plan to Excel workbook")
    p.add_argument("input", help="Path to source .xlsx file")
    p.add_argument("--plan", required=True,
                   help="Path to modification plan JSON")
    p.add_argument("-o", "--output", help="Output .xlsx path (required)")
    p.add_argument("--dry-run", action="store_true",
                   help="Preview changes without writing")
    p.add_argument("--backup", action="store_true",
                   help="Create backup of original before writing")
    p.add_argument("--force", action="store_true",
                   help="Proceed despite warnings")
    args = p.parse_args()

    src = Path(args.input)
    if not src.exists():
        sys.exit(f"Source file not found: {args.input}")

    plan_path = Path(args.plan)
    if not plan_path.exists():
        sys.exit(f"Plan file not found: {args.plan}")

    try:
        plan = json.loads(plan_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as e:
        sys.exit(f"Invalid plan JSON: {e}")

    if "changes" not in plan:
        sys.exit("Plan must contain a 'changes' array")

    output = args.output
    if not output:
        stem = src.stem
        output = str(src.parent / f"{stem}_modified{src.suffix}")
        print(f"[ast_to_excel] output not specified, using: {output}",
              file=sys.stderr)

    wb = openpyxl.load_workbook(src, keep_vba=True)

    # Snapshot pre-modification metrics
    orig_formulas = 0
    orig_merges = 0
    for sn in wb.sheetnames:
        ws = wb[sn]
        orig_merges += len(ws.merged_cells.ranges)
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    orig_formulas += 1

    # Validate
    validator = PlanValidator(wb, plan)
    errors, warnings = validator.validate()

    if errors:
        print("VALIDATION ERRORS:", file=sys.stderr)
        for e in errors:
            print(f"  ERROR: {e}", file=sys.stderr)
        sys.exit(1)

    if warnings:
        print("WARNINGS:", file=sys.stderr)
        for w in warnings:
            print(f"  WARN: {w}", file=sys.stderr)
        if not args.force:
            sys.exit(
                "Use --force to proceed despite warnings, or fix the plan.")

    # Dry run
    if args.dry_run:
        DryRunReporter.report(wb, plan, output)
        wb.close()
        return

    # Backup
    if args.backup:
        backup_path = str(src.parent / f"{src.stem}_backup{src.suffix}")
        shutil.copy2(src, backup_path)
        print(f"[ast_to_excel] backup: {backup_path}", file=sys.stderr)

    # Apply changes
    writer = SurgicalWriter(wb, plan)
    stats = writer.apply()
    print(f"[ast_to_excel] written {stats['cells_written']} cells, "
          f"skipped {stats['cells_skipped']}", file=sys.stderr)

    # Safety check
    guard = SafetyGuard(wb, plan, orig_formulas, orig_merges)
    issues = guard.check()
    if issues:
        print("SAFETY WARNINGS:", file=sys.stderr)
        for iss in issues:
            print(f"  SAFETY: {iss}", file=sys.stderr)
        if not args.force:
            wb.close()
            sys.exit("Safety check failed. Use --force to override.")

    # Save
    wb.save(output)
    wb.close()
    print(f"[ast_to_excel] saved to {output}", file=sys.stderr)


if __name__ == "__main__":
    main()
