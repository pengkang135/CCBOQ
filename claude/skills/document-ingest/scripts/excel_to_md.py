#!/usr/bin/env python3
"""Convert a simple Excel workbook to Markdown tables.

Only intended for "flat" workbooks — no merged cells, no formulas, small size.
For complex workbooks use excel_to_ast.py instead.

Usage:
    python excel_to_md.py <input.xlsx> [-o output.md] [--sheet NAME]
                          [--max-rows N] [--include-hidden]
"""

import argparse
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")


def _fmt_cell(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v).replace("|", "\\|").replace("\n", " ")
    return s.strip()


def _sheet_to_md(ws, max_rows=None):
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if max_rows and i >= max_rows:
            break
        rows.append([_fmt_cell(c) for c in row])

    if not rows:
        return "_(empty sheet)_\n"

    max_col = max(len(r) for r in rows)
    for r in rows:
        r.extend([""] * (max_col - len(r)))

    header = rows[0]
    if all(not c for c in header):
        header = [get_column_letter(i + 1) for i in range(max_col)]
        body = rows
    else:
        for i, c in enumerate(header):
            if not c:
                header[i] = get_column_letter(i + 1)
        body = rows[1:]

    lines = ["| " + " | ".join(header) + " |",
             "|" + "|".join(["---"] * max_col) + "|"]
    for r in body:
        lines.append("| " + " | ".join(r) + " |")
    return "\n".join(lines) + "\n"


def convert(path, target_sheet=None, max_rows=None, include_hidden=False):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = [f"# {Path(path).name}\n"]

    sheet_names = [target_sheet] if target_sheet else wb.sheetnames
    for name in sheet_names:
        if name not in wb.sheetnames:
            raise SystemExit(f"Sheet not found: {name}")
        ws = wb[name]
        if not include_hidden and ws.sheet_state != "visible":
            continue
        out.append(f"## {name}\n")
        out.append(_sheet_to_md(ws, max_rows=max_rows))
        out.append("")

    wb.close()
    return "\n".join(out)


def main():
    p = argparse.ArgumentParser(description="Excel -> Markdown (flat tables)")
    p.add_argument("input", help="Path to .xlsx file")
    p.add_argument("-o", "--output", help="Output .md path (default stdout)")
    p.add_argument("--sheet", help="Convert only this sheet (default: all visible)")
    p.add_argument("--max-rows", type=int, help="Cap rows per sheet")
    p.add_argument("--include-hidden", action="store_true",
                   help="Include hidden sheets")
    args = p.parse_args()

    path = Path(args.input)
    if not path.exists():
        sys.exit(f"File not found: {args.input}")

    md = convert(path, target_sheet=args.sheet, max_rows=args.max_rows,
                 include_hidden=args.include_hidden)

    if args.output:
        Path(args.output).write_text(md, encoding="utf-8")
        print(f"[excel_to_md] written to {args.output}", file=sys.stderr)
    else:
        sys.stdout.reconfigure(encoding="utf-8")
        print(md)


if __name__ == "__main__":
    main()
