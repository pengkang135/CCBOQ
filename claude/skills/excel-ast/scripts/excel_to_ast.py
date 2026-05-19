#!/usr/bin/env python3
"""Convert Excel workbook to JSON AST representation.

Usage:
    python excel_to_ast.py <input.xlsx> [--mode MODE] [-o output.json]
                           [--sheet NAME] [--range A1:Z100]
                           [--max-rows N] [--indent N]

Modes:
    workbook_summary    Sheet names, ranges, metadata (default)
    sheet_ast           Full cell-level AST with coordinates
    semantic_analysis   AST + detected regions (headers, tables, summaries)
"""

import json
import sys
import argparse
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.utils import range_boundaries, coordinate_to_tuple
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

# e.g. E12 -> (12, 5)
def _coord_to_rc(coord):
    from openpyxl.utils import coordinate_to_tuple
    return coordinate_to_tuple(coord)

# e.g. 5 -> "E"
def _col_letter(idx):
    from openpyxl.utils import get_column_letter
    return get_column_letter(idx)

def _parse_range(rng_str):
    """Parse 'A1:Z100' -> (min_col, min_row, max_col, max_row)."""
    parts = rng_str.replace("$", "").split(":")
    if len(parts) != 2:
        raise ValueError(f"Invalid range: {rng_str}")
    r1, c1 = _coord_to_rc(parts[0])
    r2, c2 = _coord_to_rc(parts[1])
    return (r1, c1, r2, c2)


class WorkbookSummarizer:
    """Extract workbook-level metadata."""

    def __init__(self, path):
        self.path = Path(path)

    def summarize(self):
        wb = openpyxl.load_workbook(self.path, read_only=True)
        result = {
            "name": self.path.name,
            "path": str(self.path.resolve()),
            "sheet_count": len(wb.sheetnames),
            "sheets": {}
        }
        for name in wb.sheetnames:
            ws = wb[name]
            dim = ws.calculate_dimension()
            result["sheets"][name] = {
                "dimensions": dim,
                "state": ws.sheet_state,
            }
            if dim and dim != "A1:A1":
                rows = ws.max_row or 0
                cols = ws.max_column or 0
                result["sheets"][name]["max_row"] = rows if rows else None
                result["sheets"][name]["max_column"] = cols if cols else None
        wb.close()
        return result


class SheetASTExtractor:
    """Extract cell-level AST for a single sheet."""

    def __init__(self, path, sheet_name, cell_range=None, max_rows=None,
                 include_empty=False):
        self.path = path
        self.sheet_name = sheet_name
        self.cell_range = cell_range
        self.max_rows = max_rows
        self.include_empty = include_empty

    def extract(self):
        wb = openpyxl.load_workbook(self.path, read_only=False)
        ws = wb[self.sheet_name]

        min_r, max_r, min_c, max_c = self._bounds(ws)
        merged_block = self._collect_merged(ws)

        cells = []
        row_count = 0

        for row_idx in range(min_r, max_r + 1):
            if self.max_rows and row_count >= self.max_rows:
                break
            row_has_data = False
            for col_idx in range(min_c, max_c + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                entry = self._cell_entry(cell, merged_block)
                if entry or self.include_empty:
                    cells.append(entry)
                    row_has_data = True
            if row_has_data:
                row_count += 1

        result = {
            "sheet": self.sheet_name,
            "range": self._make_range(min_r, min_c, max_r, max_c),
            "freeze_pane": str(ws.freeze_panes) if ws.freeze_panes else None,
            "merged_cells": [str(m) for m in ws.merged_cells.ranges],
            "row_count": row_count,
            "column_count": max_c - min_c + 1,
            "cells": cells
        }
        wb.close()
        return result

    def _collect_merged(self, ws):
        """Build dict mapping every cell in a merged range to its parent."""
        merged = {}
        for rng in ws.merged_cells.ranges:
            parent = None
            for row in range(rng.min_row, rng.max_row + 1):
                for col in range(rng.min_col, rng.max_col + 1):
                    coord = f"{_col_letter(col)}{row}"
                    if parent is None:
                        parent = coord
                    merged[coord] = parent
        return merged

    def _bounds(self, ws):
        if self.cell_range:
            r1, c1, r2, c2 = _parse_range(self.cell_range)
            return r1, r2, c1, c2
        dim = ws.calculate_dimension()
        if dim and dim != "A1:A1":
            r1, c1, r2, c2 = _parse_range(dim)
            return r1, r2, c1, c2
        return 1, 1, 1, 1

    def _cell_entry(self, cell, merged_block):
        v = cell.value
        if v is None and not self.include_empty:
            return None
        coord = f"{_col_letter(cell.column)}{cell.row}"

        formula = None
        number_format = cell.number_format if cell.number_format != "General" else None
        style_role = self._guess_style_role(cell)

        if isinstance(v, str) and v.startswith("="):
            formula = v

        return {
            "cell": coord,
            "row": cell.row,
            "column": _col_letter(cell.column),
            "value": v,
            "formula": formula,
            "number_format": number_format,
            "is_merged": coord in merged_block,
            "merged_parent": merged_block.get(coord),
            "style_role": style_role
        }

    def _guess_style_role(self, cell):
        """Assign style_role by formatting cues."""
        font = cell.font
        fill = cell.fill
        if font and font.bold:
            if fill and fill.fgColor and fill.fgColor.rgb and fill.fgColor.rgb != "00000000":
                return "header"
            return "title"
        if fill and fill.fgColor and fill.fgColor.rgb and fill.fgColor.rgb != "00000000":
            return "note"
        return "data"

    @staticmethod
    def _make_range(r1, c1, r2, c2):
        return f"{_col_letter(c1)}{r1}:{_col_letter(c2)}{r2}"


class SemanticAnalyzer:
    """Detect semantic regions from sheet AST."""

    def analyze(self, sheet_data):
        cells = sheet_data.get("cells", [])
        if not cells:
            return {"regions": [], "header_tree": None}

        by_row = defaultdict(list)
        for c in cells:
            by_row[c["row"]].append(c)

        max_col = max((c["row"] for c in cells), default=0)
        header_rows = self._find_header_rows(by_row, max_col)
        data_regions = self._find_data_regions(cells, by_row)
        summary_rows = self._find_summary_rows(by_row)
        formula_cols = self._find_formula_columns(cells)

        return {
            "regions": {
                "header_rows": header_rows,
                "data_tables": data_regions,
                "summary_rows": summary_rows,
                "formula_columns": formula_cols
            },
            "header_tree": self._build_header_tree(header_rows, cells)
        }

    def _find_header_rows(self, by_row, max_row):
        """Rows where most cells have style_role=header or bold."""
        headers = []
        for row_idx in sorted(by_row.keys()):
            row_cells = by_row[row_idx]
            header_count = sum(1 for c in row_cells if c.get("style_role") in ("header", "title"))
            if header_count >= max(1, len(row_cells) * 0.5):
                headers.append(row_idx)
        return headers

    def _find_data_regions(self, cells, by_row):
        """Find contiguous rectangular data regions below headers."""
        regions = []
        all_rows = sorted(by_row.keys())
        if not all_rows:
            return regions

        header_set = set()
        for row_idx in all_rows:
            row_cells = by_row[row_idx]
            if sum(1 for c in row_cells if c.get("style_role") == "header") >= max(1, len(row_cells) * 0.3):
                header_set.add(row_idx)

        visited = set()
        for row_idx in all_rows:
            if row_idx in visited or row_idx in header_set:
                continue
            data_cols = set()
            for c in by_row[row_idx]:
                if c.get("style_role") == "data":
                    data_cols.add(c["column"])
            if not data_cols:
                continue
            col_indices = sorted(_col_letter_to_idx(l) for l in data_cols)
            min_c = col_indices[0]
            max_c = col_indices[-1]
            end_row = row_idx
            for r in range(row_idx, all_rows[-1] + 1):
                if r in header_set:
                    break
                has_data = any(
                    _col_letter_to_idx(c["column"]) in col_indices
                    for c in by_row.get(r, [])
                )
                if has_data:
                    end_row = r
                    visited.add(r)
                else:
                    break
            if end_row > row_idx:
                regions.append({
                    "range": f"{_col_letter(min_c)}{row_idx}:{_col_letter(max_c)}{end_row}",
                    "header_row": row_idx - 1 if row_idx > 1 else None,
                    "data_start_row": row_idx,
                    "data_end_row": end_row,
                    "column_count": max_c - min_c + 1,
                    "row_count": end_row - row_idx + 1
                })
        return regions

    def _find_summary_rows(self, by_row):
        """Rows containing SUM/SUBTOTAL/AGGREGATE formulas."""
        summary = []
        for row_idx in sorted(by_row.keys()):
            for c in by_row[row_idx]:
                f = c.get("formula") or ""
                if any(kw in f.upper() for kw in
                       ("SUM(", "SUBTOTAL(", "AGGREGATE(", "SUMPRODUCT(")):
                    summary.append(row_idx)
                    break
        return summary

    def _find_formula_columns(self, cells):
        """Columns where >40% of data cells contain formulas."""
        col_formulas = defaultdict(lambda: [0, 0])
        for c in cells:
            idx = _col_letter_to_idx(c["column"])
            col_formulas[idx][1] += 1
            if c.get("formula"):
                col_formulas[idx][0] += 1
        return [
            {"column": _col_letter(idx), "formula_ratio": round(f / t, 2)}
            for idx, (f, t) in col_formulas.items()
            if t > 3 and f / t > 0.4
        ]

    def _build_header_tree(self, header_rows, cells):
        """Build multi-level header tree from header rows."""
        if len(header_rows) < 2:
            return None
        tree = {"levels": len(header_rows), "rows": header_rows}
        by_row = defaultdict(list)
        for c in cells:
            if c.get("style_role") == "header" and c["row"] in header_rows:
                by_row[c["row"]].append(c)
        nodes = []
        for row_idx in header_rows:
            for c in sorted(by_row[row_idx], key=lambda x: _col_letter_to_idx(x["column"])):
                nodes.append({
                    "level": header_rows.index(row_idx),
                    "row": row_idx,
                    "column": c["column"],
                    "text": str(c.get("display_text", c.get("value", "")))
                })
        tree["nodes"] = nodes
        return tree


def _col_letter_to_idx(letter):
    result = 0
    for ch in letter.upper():
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main():
    p = argparse.ArgumentParser(description="Excel -> JSON AST converter")
    p.add_argument("input", help="Path to .xlsx file")
    p.add_argument("--mode", choices=["workbook_summary", "sheet_ast",
                                      "semantic_analysis"],
                   default="workbook_summary")
    p.add_argument("-o", "--output", help="Output JSON path (default stdout)")
    p.add_argument("--sheet", help="Target sheet name")
    p.add_argument("--range", help="Cell range filter, e.g. A1:Z500")
    p.add_argument("--max-rows", type=int, help="Max data rows to return")
    p.add_argument("--indent", type=int, default=2, help="JSON indent")
    args = p.parse_args()

    path = Path(args.input)
    if not path.exists():
        sys.exit(f"File not found: {args.input}")

    if args.mode == "workbook_summary":
        result = {
            "mode": "workbook_summary",
            "source": str(path.resolve()),
            "workbook": WorkbookSummarizer(path).summarize()
        }
    elif args.mode == "sheet_ast":
        sheet = args.sheet
        if not sheet:
            # use first visible sheet
            wb = openpyxl.load_workbook(path, read_only=True)
            sheet = wb.sheetnames[0]
            wb.close()
        extractor = SheetASTExtractor(path, sheet, args.range, args.max_rows)
        result = {
            "mode": "sheet_ast",
            "source": str(path.resolve()),
            "sheets": [extractor.extract()]
        }
    elif args.mode == "semantic_analysis":
        sheet = args.sheet
        if not sheet:
            wb = openpyxl.load_workbook(path, read_only=True)
            sheet = wb.sheetnames[0]
            wb.close()
        extractor = SheetASTExtractor(path, sheet, args.range, args.max_rows)
        sheet_data = extractor.extract()
        analyzer = SemanticAnalyzer()
        semantic = analyzer.analyze(sheet_data)
        result = {
            "mode": "semantic_analysis",
            "source": str(path.resolve()),
            "sheets": [{**sheet_data, "semantic": semantic}]
        }

    json_text = json.dumps(result, ensure_ascii=False, indent=args.indent,
                           default=str)
    if args.output:
        Path(args.output).write_text(json_text, encoding="utf-8")
        print(f"[excel_to_ast] written to {args.output}", file=sys.stderr)
    else:
        sys.stdout.reconfigure(encoding="utf-8")
        print(json_text)


if __name__ == "__main__":
    main()
