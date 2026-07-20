"""Batch apply LLM-reviewed classification corrections to BOQ xlsx.

Usage:
  python batch_correct.py <input.xlsx> [--corrections corrections.json] [--output out.xlsx]

Corrections JSON format (array of objects):
  [
    {
      "desc_regex": "pattern to match in description",
      "current_discipline": "expected current value",
      "current_category": "expected current value",
      "current_subcategory": "expected current value",
      "new_discipline": "...",
      "new_category": "...",
      "new_subcategory": "..."
    }
  ]

Or pass inline via --inline:
  python batch_correct.py input.xlsx --inline '[{"desc_regex":"B1 Half-brick.*",...}]'
"""
import openpyxl, sys, json, argparse, re
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')


def load_corrections_from_json(path):
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def load_corrections_from_inline(json_str):
    return json.loads(json_str)


def apply_corrections(ws, corrections, desc_col, disc_col, cat_col, subcat_col, dry_run=False):
    applied = 0
    checked = 0
    per_rule = defaultdict(int)
    unmatched = []

    for idx, corr in enumerate(corrections):
        pattern = corr['desc_regex']
        exp_d = corr.get('current_discipline', '')
        exp_c = corr.get('current_category', '')
        exp_s = corr.get('current_subcategory', '')
        new_d = corr['new_discipline']
        new_c = corr['new_category']
        new_s = corr['new_subcategory']

        rule_matched = 0
        for row in range(1, ws.max_row + 1):
            desc = ws.cell(row=row, column=desc_col).value
            if not desc or not isinstance(desc, str):
                continue

            if not re.search(pattern, desc, re.IGNORECASE):
                continue

            cur_d = str(ws.cell(row=row, column=disc_col).value or '')
            cur_c = str(ws.cell(row=row, column=cat_col).value or '')
            cur_s = str(ws.cell(row=row, column=subcat_col).value or '')

            if cur_d != exp_d or cur_c != exp_c or cur_s != exp_s:
                continue

            rule_matched += 1
            if not dry_run:
                ws.cell(row=row, column=disc_col).value = new_d
                ws.cell(row=row, column=cat_col).value = new_c
                ws.cell(row=row, column=subcat_col).value = new_s

        if rule_matched:
            rule_key = f'{exp_d}/{exp_c}/{exp_s} → {new_d}/{new_c}/{new_s}'
            per_rule[rule_key] += rule_matched
            applied += rule_matched
        else:
            unmatched.append(f'  Rule #{idx}: {pattern} ({exp_d}/{exp_c}/{exp_s})')

        checked += 1

    return applied, per_rule, unmatched


def main():
    parser = argparse.ArgumentParser(description='Batch apply classification corrections to BOQ xlsx')
    parser.add_argument('xlsx', help='Path to classified BOQ xlsx')
    parser.add_argument('--corrections', '-c', help='Path to corrections JSON file')
    parser.add_argument('--inline', help='Inline corrections JSON string')
    parser.add_argument('--output', '-o', help='Output path (default: input_corrected.xlsx)')
    parser.add_argument('--sheet', default='ZOO BQ', help='Sheet name (default: ZOO BQ)')
    parser.add_argument('--desc-col', type=int, default=2, help='Description column 1-based (default: 2 = col B)')
    parser.add_argument('--disc-col', type=int, default=13, help='Discipline column 1-based (default: 13 = col M)')
    parser.add_argument('--cat-col', type=int, default=14, help='Category column 1-based (default: 14 = col N)')
    parser.add_argument('--subcat-col', type=int, default=15, help='Subcategory column 1-based (default: 15 = col O)')
    parser.add_argument('--dry-run', action='store_true', help='Preview only, do not modify file')
    args = parser.parse_args()

    if args.corrections:
        corrections = load_corrections_from_json(args.corrections)
    elif args.inline:
        corrections = load_corrections_from_inline(args.inline)
    else:
        print("Error: --corrections or --inline required")
        sys.exit(1)

    print(f"Loading workbook (this may take a minute)...")
    wb = openpyxl.load_workbook(args.xlsx)
    ws = wb[args.sheet]

    applied, per_rule, unmatched = apply_corrections(
        ws, corrections, args.desc_col, args.disc_col, args.cat_col, args.subcat_col,
        dry_run=args.dry_run
    )

    if args.dry_run:
        print(f"\n[DRY RUN] Would apply {applied} corrections in {len(per_rule)} rules:\n")
    else:
        print(f"\nApplied {applied} corrections across {len(per_rule)} rules:\n")

    for rule, count in sorted(per_rule.items(), key=lambda x: -x[1]):
        print(f"  [{count:3d}] {rule}")

    if unmatched:
        print(f"\n{len(unmatched)} rules had NO matches (check current classification):")
        for u in unmatched:
            print(u)

    if not args.dry_run:
        out = args.output or args.xlsx.replace('.xlsx', '_corrected.xlsx')
        wb.save(out)
        print(f"\nSaved to: {out}")
    else:
        print("\nNo changes written (dry run).")


if __name__ == '__main__':
    main()
