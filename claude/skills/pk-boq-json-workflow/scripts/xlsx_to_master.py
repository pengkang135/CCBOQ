"""
xlsx_to_master.py — Import a BOQ Excel sheet into JSONL master state.

Reads a BOQ Excel sheet and writes:
  - master.jsonl (one JSON object per Excel row, incl. project/chapter/subheading context)
  - schema.json  (column mapping + import metadata)
  - xlsx_backup.xlsx (copy of the original for later merge)
  - change_log.jsonl (empty file, seeded with an "import" entry)

Structure conventions supported (BOQ hierarchy markers):
  【Project Block】   top-level
  《Chapter Name》    chapter (may contain code like "01-1 ST")
  {Sub-heading}      sub-topic
  - item / plain     leaf item (may have Unit + Qty)

Usage:
  python xlsx_to_master.py --xlsx path/to/BQ.xlsx --sheet "ZOO BQ" \
      --header-row 1 --output-dir temp/boq_workflow

Column detection:
  --auto (default)  Auto-detect standard columns (No, Description, Unit, Quantity, Rate, ...).
  --columns col=idx Manual override (e.g. desc=2 unit=3 qty=4 mat_rate=5).

Existing classification columns (Discipline / Category / Material / Spec / MatUnit / MatQty /
Subcategory / Confidence / Note) are auto-detected via header names and included as
`current_*` fields per row.
"""
import argparse, json, shutil, sys, pathlib, datetime
sys.stdout.reconfigure(encoding='utf-8')

def _clean(x):
    if x is None:
        return None
    s = str(x).strip()
    if not s or s.lower() == 'nan':
        return None
    return s

def detect_columns(header_cells):
    """Return dict of standard-name -> column-index (1-based).

    Handles headers that repeat (e.g. 'Material' appearing twice — once as
    the source BOQ MatRate column, then again as a classification Material column).
    Rule: the LAST occurrence wins for material/spec/mat_unit/mat_qty (classification
    columns are usually appended to the right of the sheet). For no/desc/unit/qty
    (source columns), FIRST occurrence wins."""
    m = {}
    NORM_FIRST = {
        'no': 'no', 'no.': 'no', '序号': 'no',
        'description': 'desc', 'desc': 'desc', '描述': 'desc', '名称': 'desc',
        'unit': 'unit', '单位': 'unit',
        'quantity': 'qty', 'qty': 'qty', '数量': 'qty', 'quan.': 'qty',
        'rate': 'rate', '单价': 'rate',
        'amount': 'amount', 'total': 'total', '合价': 'amount',
        'labor': 'labor', 'labor rate': 'labor_rate', 'remark': 'remark',
    }
    NORM_LAST = {  # classification columns — last one wins
        'material': 'material',
        'discipline': 'disc',
        'category': 'cat',
        'subcategory': 'subcat',
        'spec': 'spec',
        'materialunit': 'mat_unit', 'material unit': 'mat_unit',
        'materialqty': 'mat_qty', 'material qty': 'mat_qty',
        'confidence': 'conf', 'note': 'note',
    }
    for i, h in enumerate(header_cells, 1):
        if h is None: continue
        k = str(h).strip().lower().strip('【】《》')
        if k in NORM_FIRST:
            m.setdefault(NORM_FIRST[k], i)
        elif k in NORM_LAST:
            m[NORM_LAST[k]] = i  # overwrite (last wins)
    return m


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument('--xlsx', required=True)
    ap.add_argument('--sheet', required=True)
    ap.add_argument('--header-row', type=int, default=1,
                    help='Row (1-based) containing column headers. Default 1.')
    ap.add_argument('--data-start-row', type=int, default=None,
                    help='First row of data. Default = header_row + 1 (or +2 if row 2 is subheader).')
    ap.add_argument('--output-dir', required=True)
    ap.add_argument('--columns', nargs='*', default=[],
                    help='Manual column overrides, e.g. desc=2 unit=3 qty=4')
    ap.add_argument('--force', action='store_true', help='Overwrite existing output dir')
    args = ap.parse_args()

    import openpyxl

    xlsx = pathlib.Path(args.xlsx)
    outdir = pathlib.Path(args.output_dir)
    if outdir.exists() and any(outdir.iterdir()) and not args.force:
        print(f'ERROR: output dir {outdir} not empty. Use --force to overwrite.', file=sys.stderr)
        sys.exit(2)
    outdir.mkdir(parents=True, exist_ok=True)

    print(f'Loading {xlsx} sheet={args.sheet!r}...')
    wb = openpyxl.load_workbook(xlsx, data_only=False)
    if args.sheet not in wb.sheetnames:
        print(f'ERROR: sheet {args.sheet!r} not in workbook. Available: {wb.sheetnames}', file=sys.stderr)
        sys.exit(2)
    ws = wb[args.sheet]
    print(f'  rows={ws.max_row} cols={ws.max_column}')

    header = [ws.cell(row=args.header_row, column=c).value for c in range(1, ws.max_column + 1)]
    col_map = detect_columns(header)
    for kv in args.columns:
        k, v = kv.split('=')
        col_map[k.strip()] = int(v)
    print(f'Column map: {col_map}')

    data_start = args.data_start_row or (args.header_row + 1)
    # Detect subheader row: if data_start row column A is empty AND row is header-like
    if data_start == args.header_row + 1:
        r_next = ws.cell(row=data_start, column=1).value
        if r_next is None and any(ws.cell(row=data_start, column=c).value for c in range(1, ws.max_column + 1)):
            data_start += 1
            print(f'Detected subheader on row {data_start-1}; data starts at {data_start}')

    # Walk rows and track context
    current_project = None
    current_chapter = None
    current_chapter_code = None  # extracted short code like "01-1 ST"
    current_subheading = None

    n_written = 0
    with open(outdir / 'master.jsonl', 'w', encoding='utf-8') as f:
        for r in range(data_start, ws.max_row + 1):
            row_vals = {c: ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)}
            desc = _clean(row_vals.get(col_map.get('desc', 2)))

            # Detect heading markers
            heading_type = None
            if desc:
                if desc.startswith('【') and '】' in desc:
                    current_project = desc
                    current_chapter = None
                    current_chapter_code = None
                    current_subheading = None
                    heading_type = 'project'
                elif desc.startswith('《') and desc.endswith('》'):
                    current_chapter = desc
                    current_subheading = None
                    inner = desc.strip('《》').strip()
                    current_chapter_code = inner
                    heading_type = 'chapter'
                elif desc.startswith('{') and desc.endswith('}'):
                    current_subheading = desc
                    heading_type = 'subheading'

            rec = {
                'excel_row': r,
                'heading_type': heading_type,   # None for leaf items, else project/chapter/subheading
                'no': _clean(row_vals.get(col_map.get('no', 1))),
                'desc': desc,
                'unit': _clean(row_vals.get(col_map.get('unit', 3))),
                'qty': _clean(row_vals.get(col_map.get('qty', 4))),
                'project': current_project,
                'chapter': current_chapter,
                'chapter_code': current_chapter_code,
                'subheading': current_subheading,
            }

            for name in ('disc', 'cat', 'subcat', 'material', 'spec', 'mat_unit', 'mat_qty', 'conf', 'note', 'rate', 'amount', 'labor_rate', 'remark'):
                if name in col_map:
                    rec[f'current_{name}' if name in ('disc','cat','subcat','material','spec','mat_unit','mat_qty','conf','note') else name] = _clean(row_vals.get(col_map[name]))

            f.write(json.dumps(rec, ensure_ascii=False) + '\n')
            n_written += 1

    print(f'Wrote {n_written} rows -> {outdir/"master.jsonl"}')

    # Copy xlsx backup
    shutil.copy2(xlsx, outdir / 'xlsx_backup.xlsx')
    print(f'Backup -> {outdir/"xlsx_backup.xlsx"}')

    # schema
    schema = {
        'source_xlsx': str(xlsx),
        'sheet': args.sheet,
        'header_row': args.header_row,
        'data_start_row': data_start,
        'max_row': ws.max_row,
        'max_col': ws.max_column,
        'column_map': col_map,
        'imported_at': datetime.datetime.now().isoformat(),
        'row_count': n_written,
    }
    with open(outdir / 'schema.json', 'w', encoding='utf-8') as f:
        json.dump(schema, f, ensure_ascii=False, indent=2)
    print(f'Schema -> {outdir/"schema.json"}')

    # Seed change_log
    with open(outdir / 'change_log.jsonl', 'w', encoding='utf-8') as f:
        f.write(json.dumps({
            'ts': datetime.datetime.now().isoformat(),
            'op': 'import',
            'source': str(xlsx),
            'rows': n_written,
        }, ensure_ascii=False) + '\n')
    print(f'Change log seeded -> {outdir/"change_log.jsonl"}')

    print(f'\nReady. Use scripts/query.py to explore, scripts/shard.py to prep batches.')


if __name__ == '__main__':
    main()
