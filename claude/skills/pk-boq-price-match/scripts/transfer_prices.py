#!/usr/bin/env python
"""
BOQ 单价套价：从源清单提取单价，匹配到目标清单并写入公式。

用法:
  python transfer_prices.py \
    --source source.xlsx --source-sheet "Sheet1" \
    --target target.xlsm --target-sheet "BOQ" \
    --source-col-no A --source-col-name B --source-col-unit D --source-col-price F --source-col-qty E \
    --target-col-no D --target-col-name E --target-col-unit F --target-col-qty I \
    --output output.xlsm \
    --schemes "原案" "备选案"
"""
import argparse
import json
import re
import sys
from pathlib import Path

import pandas as pd
from rapidfuzz import fuzz


def clean_text(s):
    if pd.isna(s):
        return ''
    return re.sub(r'\s+', ' ', str(s).strip())


def col_letter_to_idx(letter):
    """Convert Excel column letter(s) to 0-based index. A=0, Z=25, AA=26."""
    result = 0
    for char in letter.upper():
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result - 1


def extract_leaf_items(filepath, sheet, col_no, col_name, col_unit, col_price, col_qty=None, header_rows=4):
    """Extract leaf items (with unit + unit_price) from source sheet."""
    df = pd.read_excel(filepath, sheet_name=sheet, header=None)
    items = []
    for i in range(header_rows, len(df)):
        no = clean_text(df.iloc[i, col_no])
        name = clean_text(df.iloc[i, col_name])
        unit = clean_text(df.iloc[i, col_unit])
        price = df.iloc[i, col_price]
        qty = df.iloc[i, col_qty] if col_qty is not None else None

        if unit and pd.notna(price):
            items.append({
                'row': i,
                'no': no,
                'name': name,
                'unit': unit,
                'qty': qty if (qty is not None and pd.notna(qty)) else 0,
                'unit_price': price,
            })
    return items


def build_target_items(filepath, sheet, col_no, col_name, col_unit, col_qty, header_rows=4):
    """Build target item list with lookup index."""
    df = pd.read_excel(filepath, sheet_name=sheet, header=None)
    items = []
    no_to_rows = {}

    for i in range(header_rows, len(df)):
        no = clean_text(df.iloc[i, col_no])
        name = clean_text(df.iloc[i, col_name])
        unit = clean_text(df.iloc[i, col_unit])
        qty = df.iloc[i, col_qty]

        has_unit = unit != ''
        has_qty = pd.notna(qty)

        items.append({
            'row': i,
            'no': no,
            'name': name,
            'unit': unit,
            'qty': qty if has_qty else None,
            'has_data': has_unit and has_qty,
        })

        if no:
            no_to_rows.setdefault(no, []).append(i)

    return items, no_to_rows


def match_item(source_item, target_items, no_to_rows, min_name_score=50, min_fuzzy_score=85):
    """
    Two-level matching:
    1. Exact match on item number (with name similarity check)
    2. Fuzzy name matching (token_sort_ratio, unit must match)
    Returns target row index (0-based in DataFrame) or None.
    """
    no = source_item['no']
    name = source_item['name']
    unit = source_item['unit']

    # Level 1: item number match
    no_match_row = None
    if no and no in no_to_rows:
        candidates = no_to_rows[no]
        best_score = 0
        best_row = None
        for row in candidates:
            target_it = target_items[row - target_items[0]['row']]
            score = fuzz.token_sort_ratio(name.lower(), target_it['name'].lower())
            if score > best_score:
                best_score = score
                best_row = row
        if best_score >= min_name_score:
            no_match_row = best_row

    if no_match_row is not None:
        return no_match_row

    # Level 2: fuzzy name match
    if name:
        best_score = 0
        best_row = None
        for target_it in target_items:
            if not target_it['has_data']:
                continue
            if unit and target_it['unit'] and unit.lower() != target_it['unit'].lower():
                continue
            score = fuzz.token_sort_ratio(name.lower(), target_it['name'].lower())
            if score > best_score:
                best_score = score
                best_row = target_it['row']
        if best_score >= min_fuzzy_score:
            return best_row

    return None


def verify_match(source_items, target_items, no_to_rows):
    """Verify completeness (qty sum) and reasonableness (total amount deviation)."""
    all_boq_qty = sum(it['qty'] for it in source_items)

    matched_qty = 0
    target_qty_sum = 0
    source_total = 0
    target_total = 0
    matched_count = 0
    unmatched = []

    for item in source_items:
        target_row = match_item(item, target_items, no_to_rows)
        if target_row is not None:
            matched_count += 1
            matched_qty += item['qty']
            target_it = target_items[target_row - target_items[0]['row']]
            if target_it['qty']:
                target_qty_sum += target_it['qty']
                target_total += item['unit_price'] * target_it['qty']
            source_total += item['unit_price'] * item['qty']
        else:
            unmatched.append(f"{item['no']} {item['name'][:60]}")

    qty_diff = abs(all_boq_qty - matched_qty)
    qty_ok = qty_diff < 1

    if source_total > 0:
        deviation = (target_total - source_total) / source_total * 100
    else:
        deviation = 0
    amount_ok = abs(deviation) <= 50

    return {
        'matched_count': matched_count,
        'total_items': len(source_items),
        'all_boq_qty': all_boq_qty,
        'matched_qty': matched_qty,
        'target_qty_sum': target_qty_sum,
        'qty_diff': qty_diff,
        'qty_ok': qty_ok,
        'source_total': source_total,
        'target_total': target_total,
        'deviation': deviation,
        'amount_ok': amount_ok,
        'unmatched': unmatched,
    }


def fix_qty_formulas(ws, qty_col, header_rows, qty_values):
    """Replace Qty formulas with hardcoded values (openpyxl .xlsm safety)."""
    fixed = 0
    for row_idx, qty_val in qty_values.items():
        excel_row = row_idx + 1  # 0-based to 1-based
        cell = ws.cell(row=excel_row, column=qty_col)
        current = cell.value
        if current is not None and str(current).startswith('='):
            cell.value = qty_val
            fixed += 1
        elif current is None:
            cell.value = qty_val
            fixed += 1
    return fixed


def compute_qty_values(filepath, sheet, true_qty_col, factor_col, header_rows):
    """Compute Qty = TrueQty * Factor from target file columns."""
    df = pd.read_excel(filepath, sheet_name=sheet, header=None)
    values = {}
    for i in range(header_rows, len(df)):
        true_qty = df.iloc[i, true_qty_col]
        factor = df.iloc[i, factor_col]
        if pd.notna(true_qty):
            if pd.notna(factor):
                values[i] = true_qty * factor
            else:
                values[i] = true_qty
    return values


def main():
    parser = argparse.ArgumentParser(description='BOQ 单价套价：从源清单匹配单价到目标清单')
    parser.add_argument('--source', required=True, help='源清单文件（提供单价）')
    parser.add_argument('--source-sheet', required=True, help='源清单 sheet 名')
    parser.add_argument('--source-col-no', required=True, help='源清单项目编号列 (如 A)')
    parser.add_argument('--source-col-name', required=True, help='源清单项目名称列 (如 B)')
    parser.add_argument('--source-col-unit', required=True, help='源清单单位列 (如 D)')
    parser.add_argument('--source-col-price', required=True, help='源清单单价列 (如 F)')
    parser.add_argument('--source-col-qty', help='源清单工程量列 (可选, 如 E)')

    parser.add_argument('--target', required=True, help='目标清单文件（接收单价）')
    parser.add_argument('--target-sheet', required=True, help='目标清单 sheet 名')
    parser.add_argument('--target-col-no', required=True, help='目标清单项目编号列 (如 D)')
    parser.add_argument('--target-col-name', required=True, help='目标清单项目名称列 (如 E)')
    parser.add_argument('--target-col-unit', required=True, help='目标清单单位列 (如 F)')
    parser.add_argument('--target-col-qty', required=True, help='目标清单工程量列 (如 I)')
    parser.add_argument('--target-true-qty-col', help='目标清单实际工程量列 (用于修复 Qty 公式)')
    parser.add_argument('--target-factor-col', help='目标清单放大系数列 (用于修复 Qty 公式)')

    parser.add_argument('--header-rows', type=int, default=4, help='表头行数 (默认 4)')
    parser.add_argument('--schemes', nargs='+', required=True, help='方案名称列表 (对应源清单中的 sheet 或标识)')
    parser.add_argument('--output', help='输出文件路径 (默认覆盖 --target)')
    parser.add_argument('--dry-run', action='store_true', help='仅验证匹配，不写入文件')
    parser.add_argument('--json', action='store_true', help='输出 JSON 格式结果')

    args = parser.parse_args()

    # Convert column letters to indices
    src_no = col_letter_to_idx(args.source_col_no)
    src_name = col_letter_to_idx(args.source_col_name)
    src_unit = col_letter_to_idx(args.source_col_unit)
    src_price = col_letter_to_idx(args.source_col_price)
    src_qty = col_letter_to_idx(args.source_col_qty) if args.source_col_qty else None

    tgt_no = col_letter_to_idx(args.target_col_no)
    tgt_name = col_letter_to_idx(args.target_col_name)
    tgt_unit = col_letter_to_idx(args.target_col_unit)
    tgt_qty = col_letter_to_idx(args.target_col_qty)

    # Extract target data
    target_items, no_to_rows = build_target_items(
        args.target, args.target_sheet, tgt_no, tgt_name, tgt_unit, tgt_qty, args.header_rows
    )

    # Process each scheme
    results = {}
    all_matches = {}

    for scheme_name in args.schemes:
        # Try: scheme_name as sheet name, or find matching sheet
        source_items = extract_leaf_items(
            args.source, args.source_sheet if len(args.schemes) == 1 else scheme_name,
            src_no, src_name, src_unit, src_price, src_qty, args.header_rows
        )

        scheme_matches = {}
        for item in source_items:
            target_row = match_item(item, target_items, no_to_rows)
            if target_row is not None:
                scheme_matches[target_row] = item['unit_price']

        verification = verify_match(source_items, target_items, no_to_rows)
        results[scheme_name] = {
            'matches': len(scheme_matches),
            'total': len(source_items),
            **verification,
        }
        all_matches[scheme_name] = scheme_matches

    if args.json:
        output = {
            name: {k: v for k, v in r.items() if k != 'unmatched'}
            for name, r in results.items()
        }
        print(json.dumps(output, indent=2, default=str))
    else:
        for name, r in results.items():
            print(f"\n=== {name} ===")
            print(f"  Matched: {r['matched_count']}/{r['total_items']}")
            print(f"  Qty check: {'PASS' if r['qty_ok'] else 'FAIL'} (diff={r['qty_diff']:,.0f})")
            print(f"  Amount deviation: {r['deviation']:.1f}% {'PASS' if r['amount_ok'] else 'WARNING'}")

    if args.dry_run:
        print("\nDry run — no file written.")
        return

    # Write to target file
    output_path = args.output or args.target
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    wb = load_workbook(args.target, keep_vba=args.target.lower().endswith('.xlsm'))
    ws = wb[args.target_sheet]

    # Fix Qty formulas if columns provided
    if args.target_true_qty_col and args.target_factor_col:
        true_qty_col = col_letter_to_idx(args.target_true_qty_col)
        factor_col = col_letter_to_idx(args.target_factor_col)
        qty_col_1based = tgt_qty + 1  # openpyxl is 1-based
        qty_values = compute_qty_values(args.target, args.target_sheet, true_qty_col, factor_col, args.header_rows)
        fixed = fix_qty_formulas(ws, qty_col_1based, args.header_rows, qty_values)
        print(f"Fixed {fixed} Qty formula cells")

    # Find rightmost column for new columns
    max_col = ws.max_column
    start_col = max_col + 1  # 1-based

    scheme_names = args.schemes
    for idx, scheme_name in enumerate(scheme_names):
        price_col = start_col + idx * 2
        total_col = start_col + idx * 2 + 1
        price_letter = get_column_letter(price_col)
        total_letter = get_column_letter(total_col)
        qty_letter = get_column_letter(tgt_qty + 1)

        # Headers
        ws.cell(row=1, column=price_col, value=f'{scheme_name} Unit Price')
        ws.cell(row=2, column=price_col, value=f'{scheme_name} 单价')
        ws.cell(row=1, column=total_col, value=f'{scheme_name} Amount')
        ws.cell(row=2, column=total_col, value=f'{scheme_name} 合价')

        # Data
        scheme_matches = all_matches[scheme_name]
        for target_row, unit_price in scheme_matches.items():
            excel_row = target_row + 1
            ws.cell(row=excel_row, column=price_col, value=unit_price)
            ws.cell(row=excel_row, column=total_col,
                    value=f'={price_letter}{excel_row}*{qty_letter}{excel_row}')

    wb.save(output_path)
    wb.close()
    print(f"\nFile saved: {output_path}")
    print(f"New columns: {get_column_letter(start_col)} - {get_column_letter(start_col + len(scheme_names) * 2 - 1)}")


if __name__ == '__main__':
    main()
