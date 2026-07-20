#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Build unified Excel material price table from quotation data.json.

Accepts the same data.json format as upload-rates.js, producing a standardized
16-column 人材机价格表 Excel file suitable for sharing and review.

Usage:
  python build_batch_excel.py <data.json> [-o output.xlsx] [--title ...] [--subtitle ...]
"""

import argparse, json, sys, os
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill


def load_data(path: str) -> dict:
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def build_xlsx(data: dict, output_path: str, title: str, subtitle: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quotation"

    project = data.get("project", {})
    suppliers = data.get("suppliers", [])

    # Styles
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    header_font = Font(name="Microsoft YaHei", size=11, bold=True)
    title_font = Font(name="Microsoft YaHei", size=14, bold=True)
    data_font = Font(name="Microsoft YaHei", size=10)
    group_font = Font(name="Microsoft YaHei", size=10, bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    group_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Column widths (A-P, 16 columns)
    col_widths = {
        "A": 8, "B": 14, "C": 40, "D": 60, "E": 8, "F": 14,
        "G": 12, "H": 14, "I": 14, "J": 8, "K": 28, "L": 10,
        "M": 14, "N": 30, "O": 30, "P": 40,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # Row 1: Title
    ws.merge_cells("A1:P1")
    c = ws["A1"]
    c.value = title
    c.font = title_font
    c.alignment = center_align

    # Row 2: Headers
    headers = [
        "编号", "专业", "名称", "项目特征", "单位", "除税单价", "税金", "含税单价",
        "日期", "币种", "供应商", "联系人", "电话", "地址", "备注", "来源",
    ]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align

    # Row 3: Subtitle
    ws.merge_cells("A3:P3")
    c = ws["A3"]
    c.value = subtitle
    c.font = Font(name="Microsoft YaHei", size=9, bold=True)
    c.alignment = left_align

    # Data rows (start from row 4)
    row = 4
    seq = 1
    all_prices = []

    for supplier in suppliers:
        supplier_name = supplier.get("name_cn") or supplier.get("name", "")
        source_file = supplier.get("sourceFile", "")

        # Supplier group header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=16)
        cell = ws.cell(row=row, column=1, value=f"《{supplier_name}》")
        cell.font = group_font
        cell.fill = group_fill
        cell.alignment = left_align
        for col_idx in range(1, 17):
            ws.cell(row=row, column=col_idx).border = thin_border
            ws.cell(row=row, column=col_idx).fill = group_fill
        row += 1

        for item in supplier.get("items", []):
            price_excl = float(item.get("price_excl_tax", 0))
            price_incl = float(item.get("price_incl_tax", 0))
            tax = price_incl - price_excl
            all_prices.append(price_excl)

            vals = [
                seq,
                project.get("specialty", ""),
                item.get("name_cn") or item.get("name", ""),
                item.get("features_cn") or item.get("features", ""),
                item.get("unit", ""),
                price_excl,
                round(tax, 2),
                price_incl,
                item.get("date", ""),
                item.get("currency", "THB"),
                supplier_name,
                supplier.get("contact", ""),
                supplier.get("phone", ""),
                supplier.get("address_cn") or supplier.get("address", ""),
                supplier.get("remarks", ""),
                source_file,
            ]

            for col_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = data_font
                cell.border = thin_border
                if col_idx in (1, 5, 6, 7, 8, 9, 10, 12, 13, 14):
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align

            seq += 1
            row += 1

    # Summary rows
    row += 1
    total = seq - 1
    sources = sorted(set(s.get("sourceFile", "") for s in suppliers))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=16)
    cell = ws.cell(row=row, column=1)
    cell.value = f"数据来源：{', '.join(sources)}"
    cell.font = Font(name="Microsoft YaHei", size=9)
    cell.alignment = left_align
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=16)
    cell = ws.cell(row=row, column=1)
    if all_prices:
        lo, hi = min(all_prices), max(all_prices)
        avg = sum(all_prices) / len(all_prices)
        cell.value = f"价格统计：共{total}项，最低{lo:,.2f}，最高{hi:,.2f}，平均{avg:,.2f}"
    else:
        cell.value = f"价格统计：共{total}项"
    cell.font = Font(name="Microsoft YaHei", size=9)
    cell.alignment = left_align

    wb.save(output_path)
    return total, (min(all_prices) if all_prices else 0), (max(all_prices) if all_prices else 0)


def main():
    parser = argparse.ArgumentParser(description="Build unified quotation Excel from data.json")
    parser.add_argument("data", help="JSON data file (same format as upload-rates.js input)")
    parser.add_argument("-o", "--output", default=None, help="Output xlsx path")
    parser.add_argument("--title", default="人材机价格表", help="Main title")
    parser.add_argument("--subtitle", default="", help="Subtitle in row 3")
    args = parser.parse_args()

    data = load_data(args.data)
    output = args.output or Path(args.data).with_suffix(".xlsx").name

    total, lo, hi = build_xlsx(data, output, args.title, args.subtitle)
    print(f"Generated: {output}")
    print(f"  Entries: {total}, Price range: {lo:,.2f} - {hi:,.2f}")


if __name__ == "__main__":
    main()
