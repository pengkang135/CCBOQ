#!/usr/bin/env python3
"""Build market inquiry material table from design institute BOQ.

Three-phase pipeline:
  Phase 1: Extract leaf items from BOQ Excel
  Phase 2: Consolidate by material category (keyword matching)
  Phase 3: Format to hierarchical inquiry template (xlsxwriter)
"""
import argparse, json, os, re, sys
from collections import OrderedDict
from datetime import date

import openpyxl
import xlsxwriter

# ── Phase 1: Extract ──────────────────────────────────────────────

def _col_letter_to_idx(letter):
    result = 0
    for ch in letter.upper():
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result


def detect_columns_from_ast(ast_path, sheet_name):
    """Auto-detect column mapping from document-ingest semantic_analysis JSON.
    Returns dict like {'item_no': 0, 'description': 1, 'unit': 2, 'quantity': 3, 'spec': 4}
    or None if detection fails.
    """
    with open(ast_path, 'r', encoding='utf-8') as f:
        ast = json.load(f)

    for s in ast.get('sheets', []):
        if s.get('sheet') != sheet_name:
            continue
        semantic = s.get('semantic', {})
        header_tree = semantic.get('header_tree', {})
        if not header_tree:
            continue

        col_map = {'item_no': 0, 'description': 1, 'unit': 2, 'quantity': 3, 'spec': 4}
        for node in header_tree.get('nodes', []):
            text = (node.get('text') or '').lower()
            col_letter = node.get('column', '')
            col_idx = _col_letter_to_idx(col_letter) - 1  # 0-based

            if any(kw in text for kw in ('item', 'no', '编号', '序号', 'code')):
                col_map['item_no'] = col_idx
            elif any(kw in text for kw in ('description', 'desc', '描述', '名称', '项目特征', '项目名称')):
                col_map['description'] = col_idx
            elif any(kw in text for kw in ('unit', '单位')):
                col_map['unit'] = col_idx
            elif any(kw in text for kw in ('quantity', 'qty', '数量', '工程量')):
                col_map['quantity'] = col_idx
            elif any(kw in text for kw in ('spec', '规格', '型号', '规格型号')):
                col_map['spec'] = col_idx

        return col_map
    return None

def _is_leaf(item_no, min_depth=2):
    s = str(item_no).strip() if item_no else ''
    parts = s.split('.')
    return len(parts) >= min_depth or 'ADD' in s.upper()


def _clean(v):
    if v is None:
        return ''
    return str(v).strip()


def _to_float(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(',', '').strip())
    except (ValueError, TypeError):
        return 0.0


def _should_exclude(desc, exclude_kw):
    if not exclude_kw:
        return False
    text = desc.lower()
    return any(kw.lower() in text for kw in exclude_kw)


def extract_items(source_path, config):
    """Phase 1: Extract leaf items from BOQ source Excel."""
    src = config['source']
    sheets = src.get('sheets', [])
    cols = src.get('columns', {})
    item_filter = src.get('item_filter', {})
    min_depth = item_filter.get('min_depth', 2)
    exclude_kw = item_filter.get('exclude_keywords', [])

    ci, cd, cu, cq1 = cols.get('item_no', 0), cols.get('description', 1), cols.get('unit', 2), cols.get('quantity', 3)
    cq2 = cols.get('quantity_alt', cq1 + 1)
    cs = cols.get('spec', cq2 + 1)

    wb = openpyxl.load_workbook(source_path, data_only=True)
    all_items = []
    sheet_stats = {}

    for sheet_name in sheets:
        if sheet_name not in wb.sheetnames:
            print(f'  [WARN] Sheet "{sheet_name}" not found, skipping')
            continue

        ws = wb[sheet_name]
        count = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            item_no = _clean(row[ci]) if len(row) > ci else ''
            desc = _clean(row[cd]) if len(row) > cd else ''
            unit = _clean(row[cu]) if len(row) > cu else ''
            qty = _to_float(row[cq1]) if len(row) > cq1 else 0.0
            spec = _clean(row[cs]) if len(row) > cs else ''

            if not _is_leaf(item_no, min_depth):
                continue
            if not desc:
                continue
            if _should_exclude(desc, exclude_kw):
                continue
            if _should_exclude(spec, exclude_kw):
                continue

            if qty <= 0 and len(row) > cq2:
                qty = _to_float(row[cq2])
            if qty <= 0:
                continue

            all_items.append({
                'item_no': item_no, 'desc': desc, 'spec': spec,
                'unit': unit, 'qty': qty, 'sheet': sheet_name,
            })
            count += 1

        sheet_stats[sheet_name] = count

    wb.close()

    for s, c in sheet_stats.items():
        print(f'  {s}: {c} items')
    print(f'  Total: {len(all_items)} leaf items extracted')

    return all_items


# ── Phase 2: Consolidate ───────────────────────────────────────────

def _match_item(item, group):
    text = ' '.join((item['desc'] + ' ' + item['spec']).lower().split())
    keywords = group.get('keywords', [])
    match_all = group.get('match_all', True)

    if match_all:
        return all(kw.lower() in text for kw in keywords)
    else:
        return any(kw.lower() in text for kw in keywords)


def consolidate_items(items, config):
    """Phase 2: Classify items into material groups and merge quantities."""
    groups = config['consolidation'].get('groups', [])
    qty_factor = config['consolidation'].get('quantity_factor', 1.05)

    grouped = OrderedDict()
    for g in groups:
        grouped[g['name']] = []

    unmatched = []
    for item in items:
        found = False
        for g in groups:
            if _match_item(item, g):
                grouped[g['name']].append(item)
                found = True
                break
        if not found:
            unmatched.append(item)

    if unmatched:
        print(f'  [WARN] {len(unmatched)} items unmatched, will be discarded')
        for it in unmatched[:10]:
            print(f'    - {it["item_no"]}: {it["desc"][:120]}')
        if len(unmatched) > 10:
            print(f'    ... and {len(unmatched) - 10} more')

    # Merge items within each group
    result = OrderedDict()
    seq = 0
    for g in groups:
        group_name = g['name']
        items_in_group = grouped.get(group_name, [])
        if not items_in_group:
            continue

        # Deduplicate by desc, accumulate qty
        merged = OrderedDict()
        for it in items_in_group:
            desc_key = it['desc'].strip().lower()
            if desc_key in merged:
                merged[desc_key]['qty'] += it['qty']
                if it['remark']:
                    if not merged[desc_key].get('remark'):
                        merged[desc_key]['remark'] = it['remark']
            else:
                unit = g.get('unit_override') or it['unit']
                merged[desc_key] = {
                    'desc': it['desc'],
                    'spec': it['spec'],
                    'unit': unit,
                    'qty': it['qty'],
                    'remark': g.get('note', ''),
                    'source_spec': it['spec'],
                }

        # Apply quantity factor and round
        mat_list = []
        for desc_key, m in merged.items():
            qty = m['qty'] * qty_factor
            m['qty'] = _round_qty(qty)
            mat_list.append(m)

        result[group_name] = mat_list
        seq += len(mat_list)

    # Assign M IDs
    mid = 1
    for group_name, mat_list in result.items():
        for m in mat_list:
            m['id'] = f'M{mid:03d}'
            mid += 1

    print(f'  {len(result)} material groups, {sum(len(v) for v in result.values())} consolidated items')
    return result


def _round_qty(qty):
    if qty >= 10000:
        return round(qty / 100) * 100
    if qty >= 1000:
        return round(qty / 50) * 50
    if qty >= 100:
        return round(qty / 10) * 10
    if qty >= 10:
        return round(qty)
    return round(qty, 1)


# ── Phase 3: Format ────────────────────────────────────────────────

# BOQ层级样式常量（引用 boq_hierarchy_rules.md）
BOQ_COLORS = {
    'l1_bg': '#C6D9F1',    # L1 section
    'l2_bg': '#EEF2FA',    # L2 class
    'l3_bg': '#FBE5D6',    # L3 sub-class
    'header_bg': '#4472C4',
    'border': '#BFBFBF',
}


def _build_hierarchy(consolidated, config):
    """Build L1 > L2 > items structure with optional L3."""
    l1_configs = config['hierarchy'].get('l1_groups', [])
    l3_threshold = config.get('l3_threshold', 30)

    l1_map = OrderedDict()
    for l1c in l1_configs:
        l1_map[l1c['name']] = OrderedDict()

    unmapped_l1 = OrderedDict()

    for l2_name, mat_list in consolidated.items():
        assigned = False
        for l1c in l1_configs:
            for kw in l1c.get('l2_keywords', []):
                if kw.lower() in l2_name.lower():
                    l1_map.setdefault(l1c['name'], OrderedDict())[l2_name] = mat_list
                    assigned = True
                    break
            if assigned:
                break
        if not assigned:
            unmapped_l1[l2_name] = mat_list

    if unmapped_l1:
        l1_map['其他 Miscellaneous'] = unmapped_l1

    empty = [k for k, v in l1_map.items() if not v]
    for k in empty:
        del l1_map[k]

    # L3 auto-detection
    needs_l3 = {}
    for l1_name, l2_sections in l1_map.items():
        for l2_name, items in l2_sections.items():
            if len(items) > l3_threshold:
                needs_l3[(l1_name, l2_name)] = True

    return l1_map, needs_l3


def _split_spec(desc):
    m = re.search(r'\[(.+?)\]', desc)
    if m:
        spec = m.group(1).strip()
        name = re.sub(r'\s*\[.+?\]', '', desc).strip()
        return name, spec
    return desc, ''


def format_output(consolidated, config, template_path, output_dir, title=None, no_md=False, no_xlsx=False):
    """Phase 3: Format consolidated materials into MD + xlsx output."""
    project = config.get('project', 'Project')
    today = date.today().isoformat()

    if title is None:
        title = f'人材机价格表 — {project} 市场询价'

    hierarchy, needs_l3 = _build_hierarchy(consolidated, config)

    if not no_md:
        md_path = os.path.join(output_dir, f'{today}_材料询价表.md')
        _write_md(hierarchy, title, project, today, md_path)
        print(f'  MD: {md_path}')

    if not no_xlsx:
        xlsx_path = os.path.join(output_dir, f'{today}_材料询价表.xlsx')
        _write_xlsx(hierarchy, needs_l3, title, xlsx_path)
        print(f'  xlsx: {xlsx_path}')


def _write_md(hierarchy, title, project, today, md_path):
    lines = [f'# {title}', '',
             f'> 来源：{project} Schedule of Prices',
             f'> 日期：{today}',
             f'> 说明：同类材料已跨部位合并，工程量为图纸量\xd71.05~1.10（施工损耗/搭接）取整。',
             '', '---', '']

    seq = 0
    for l1_name, l2_sections in hierarchy.items():
        lines.append(f'## {l1_name}')
        lines.append('')
        lines.append('| 编号 | 材料类别 | 材料名称/规格 | 单位 | 数量 | 币种 | 备注 |')
        lines.append('|------|---------|-------------|------|------|------|------|')
        for l2_name, items in l2_sections.items():
            lines.append(f'| | **{l2_name}** | | | | | |')
            for item in items:
                name, spec = _split_spec(item['desc'])
                if spec:
                    name = f'{name} [{spec}]'
                remark = item.get('remark', '') or ''
                lines.append(f'| {item["id"]} | | {name} | {item["unit"]} | {item["qty"]:,} | USD | {remark} |')
        lines.append('')

    lines.append(f'---')
    lines.append(f'**合计：{sum(len(items) for l2 in hierarchy.values() for items in l2.values())} 项材料**')

    with open(md_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))


def _write_xlsx(hierarchy, needs_l3, title, xlsx_path):
    wb = xlsxwriter.Workbook(xlsx_path)
    ws = wb.add_worksheet('市场询价表')

    # BOQ-standard formats
    fmt_l1 = wb.add_format({
        'bold': True, 'font_size': 11, 'font_name': 'Microsoft YaHei UI',
        'font_color': '#1A1A1A', 'bg_color': BOQ_COLORS['l1_bg'],
        'valign': 'vcenter', 'border': 0,
    })
    fmt_l2 = wb.add_format({
        'bold': True, 'font_size': 10, 'font_name': 'Microsoft YaHei UI',
        'font_color': '#1A1A1A', 'bg_color': BOQ_COLORS['l2_bg'],
        'valign': 'vcenter', 'border': 0,
    })
    fmt_l3 = wb.add_format({
        'bold': True, 'font_size': 10, 'font_name': 'Microsoft YaHei UI',
        'font_color': '#1A1A1A', 'bg_color': BOQ_COLORS['l3_bg'],
        'valign': 'vcenter', 'border': 0,
    })
    fmt_header = wb.add_format({
        'bold': True, 'font_size': 10, 'font_name': 'Microsoft YaHei UI',
        'bg_color': BOQ_COLORS['header_bg'], 'font_color': '#FFFFFF',
        'border': 1, 'text_wrap': True, 'valign': 'vcenter', 'align': 'center',
    })
    fmt_data = wb.add_format({
        'font_size': 9, 'font_name': 'Microsoft YaHei UI',
        'border': 1, 'valign': 'vcenter', 'border_color': BOQ_COLORS['border'],
    })
    fmt_num = wb.add_format({
        'font_size': 9, 'font_name': 'Microsoft YaHei UI',
        'border': 1, 'valign': 'vcenter', 'border_color': BOQ_COLORS['border'],
        'num_format': '#,##0.00',
    })
    fmt_num_qty = wb.add_format({
        'font_size': 9, 'font_name': 'Microsoft YaHei UI',
        'border': 1, 'valign': 'vcenter', 'border_color': BOQ_COLORS['border'],
        'num_format': '#,##0.0',
    })
    fmt_center = wb.add_format({
        'font_size': 9, 'font_name': 'Microsoft YaHei UI',
        'border': 1, 'valign': 'vcenter', 'border_color': BOQ_COLORS['border'],
        'align': 'center',
    })

    headers = ['编号', '专业', '名称', '项目特征', '单位', '数量',
               '除税单价', '税金', '含税单价', '日期', '币种',
               '供应商', '联系人', '电话', '地址', '备注']
    n_cols = len(headers)
    col_widths = [7, 14, 28, 24, 6, 9, 9, 6, 9, 10, 5, 10, 6, 10, 10, 18]

    for c, w in enumerate(col_widths):
        ws.set_column(c, c, w)

    ws.set_row(0, 22)
    for c, h in enumerate(headers):
        ws.write(0, c, h, fmt_header)
    ws.freeze_panes(1, 0)

    row = 1
    seq = 0

    for l1_name, l2_sections in hierarchy.items():
        if not l2_sections:
            continue

        # L1 header
        ws.write(row, 0, 'P', fmt_l1)
        ws.write(row, 1, '', fmt_l1)
        ws.write(row, 2, f'【{l1_name}】', fmt_l1)
        for c in range(3, n_cols):
            ws.write(row, c, '', fmt_l1)
        ws.set_row(row, 16.5, None, {'level': 0})
        row += 1

        for l2_name, items in l2_sections.items():
            if not items:
                continue

            key = (l1_name, l2_name)
            has_l3 = needs_l3.get(key, False)

            # L2 header
            ws.write(row, 0, '', fmt_l2)
            ws.write(row, 1, '', fmt_l2)
            ws.write(row, 2, f'《{l2_name}》', fmt_l2)
            for c in range(3, n_cols):
                ws.write(row, c, '', fmt_l2)
            ws.set_row(row, 14.5, None, {'level': 1})
            row += 1

            for item in items:
                name, spec = _split_spec(item['desc'])
                seq += 1
                ws.write(row, 0, item.get('id', ''), fmt_data)
                ws.write(row, 1, item.get('category', l2_name), fmt_data)
                ws.write(row, 2, name, fmt_data)
                ws.write(row, 3, spec or item.get('spec', ''), fmt_data)
                ws.write(row, 4, item.get('unit', ''), fmt_center)
                ws.write_number(row, 5, item.get('qty', 0), fmt_num_qty)
                ws.write(row, 6, '', fmt_num)
                ws.write(row, 7, '', fmt_num)
                ws.write(row, 8, '', fmt_num)
                ws.write(row, 9, '', fmt_center)
                ws.write(row, 10, 'USD', fmt_center)
                ws.write(row, 11, '', fmt_data)
                ws.write(row, 12, '', fmt_data)
                ws.write(row, 13, '', fmt_data)
                ws.write(row, 14, '', fmt_data)
                ws.write(row, 15, item.get('remark', ''), fmt_data)
                data_level = 3 if has_l3 else 2
                ws.set_row(row, 14.5, None, {'level': data_level})
                row += 1

    wb.close()


# ── CLI ────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description='Build market inquiry material table from BOQ')
    ap.add_argument('--source', help='BOQ source Excel path (Phase 1)')
    ap.add_argument('--config', required=True, help='JSON config file path')
    ap.add_argument('--ast', help='document-ingest semantic_analysis JSON for auto column detection (Phase 1)')
    ap.add_argument('--template', help='Reference template xlsx (Phase 3)')
    ap.add_argument('-o', '--output', default='.', help='Output directory')
    ap.add_argument('--phase', type=int, choices=[1, 2, 3], help='Run single phase only')
    ap.add_argument('--items', help='Phase 1 output JSON (Phase 2 input)')
    ap.add_argument('--consolidated', help='Phase 2 output JSON (Phase 3 input)')
    ap.add_argument('--title', help='Excel title override')
    ap.add_argument('--no-md', action='store_true', help='Skip MD output')
    ap.add_argument('--no-xlsx', action='store_true', help='Skip xlsx output')
    args = ap.parse_args()

    with open(args.config, 'r', encoding='utf-8') as f:
        config = json.load(f)

    # Override column config from document-ingest if provided
    if args.ast and args.source:
        col_map = detect_columns_from_ast(args.ast, config['source'].get('sheets', ['Sheet1'])[0])
        if col_map:
            cols = config['source'].setdefault('columns', {})
            for k in ('item_no', 'description', 'unit', 'quantity', 'spec'):
                if k not in cols or cols[k] == 0:
                    cols[k] = col_map.get(k, cols.get(k, 0))
            print(f'  Auto-detected columns from AST: {col_map}')

    os.makedirs(args.output, exist_ok=True)

    # Phase 1: Extract
    if args.phase is None or args.phase == 1:
        if not args.source:
            ap.error('--source required for Phase 1')
        print('Phase 1: Extracting leaf items...')
        items = extract_items(args.source, config)
        if args.phase == 1 or args.items:
            items_path = args.items or os.path.join(args.output, 'items.json')
            with open(items_path, 'w', encoding='utf-8') as f:
                json.dump(items, f, ensure_ascii=False, indent=2)
            print(f'  Saved: {items_path}')
        if args.phase == 1:
            return

    # Phase 2: Consolidate
    if args.phase is None or args.phase == 2:
        if args.phase == 2 and not args.items:
            ap.error('--items required for Phase 2 standalone')
        if args.items and args.phase == 2:
            with open(args.items, 'r', encoding='utf-8') as f:
                items = json.load(f)
        print('Phase 2: Consolidating by material category...')
        consolidated = consolidate_items(items, config)
        if args.phase == 2 or args.consolidated:
            cons_path = args.consolidated or os.path.join(args.output, 'consolidated.json')
            with open(cons_path, 'w', encoding='utf-8') as f:
                json.dump(consolidated, f, ensure_ascii=False, indent=2)
            print(f'  Saved: {cons_path}')
        if args.phase == 2:
            return

    # Phase 3: Format
    if args.phase is None or args.phase == 3:
        if args.phase == 3 and not args.consolidated:
            ap.error('--consolidated required for Phase 3 standalone')
        if args.consolidated and args.phase == 3:
            with open(args.consolidated, 'r', encoding='utf-8') as f:
                consolidated = json.load(f)
        print('Phase 3: Formatting output...')
        format_output(consolidated, config, args.template, args.output,
                      title=args.title, no_md=args.no_md, no_xlsx=args.no_xlsx)

    print('\nDone.')


if __name__ == '__main__':
    main()
