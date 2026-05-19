#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Build standardized 人材机价格表 Excel from quotation data.

Usage:
  python build_quotation_xlsx.py --data <data_file.py> [-o output.xlsx] [--title ...] [--subtitle ...]

The data file must define ALL_DATA as a list of tuples:
  (group, code, specialty, name, description, unit, price, date, currency, source, supplier, remark)
"""

import argparse, importlib.util, sys, os
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill


def load_data_module(path: str):
    spec = importlib.util.spec_from_file_location("quotation_data", path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod.ALL_DATA


def build_xlsx(data: list[tuple], output_path: str, title: str, subtitle: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quotation"

    # Styles
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_font = Font(name='Microsoft YaHei', size=11, bold=True)
    title_font = Font(name='Microsoft YaHei', size=14, bold=True)
    data_font = Font(name='Microsoft YaHei', size=10)
    group_font = Font(name='Microsoft YaHei', size=10, bold=True)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    group_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Column widths (A-P, 16 columns)
    col_widths = {'A': 16, 'B': 14, 'C': 40, 'D': 60, 'E': 8, 'F': 12, 'G': 10,
                  'H': 12, 'I': 14, 'J': 8, 'K': 22, 'L': 6, 'M': 6, 'N': 8, 'O': 30, 'P': 40}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # Row 1: Title
    ws.merge_cells('A1:P1')
    c = ws['A1']
    c.value = title
    c.font = title_font
    c.alignment = center_align

    # Row 2: Headers
    headers = ['编号', '专业', '名称', '项目特征', '单位', '除税单价', '税金', '含税单价',
               '日期', '币种', '供应商', '联系人', '电话', '地址', '备注', '来源']
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align

    # Row 3: Subtitle
    ws.merge_cells('A3:P3')
    c = ws['A3']
    c.value = subtitle
    c.font = Font(name='Microsoft YaHei', size=9, bold=True)
    c.alignment = left_align

    # Data rows (start from row 4)
    current_row = 4
    current_group = None

    for entry in data:
        n = len(entry)
        if n >= 12:
            group, code, specialty, name, desc, unit, price, date, currency, source, supplier, remark = entry[:12]
        else:
            group, code, specialty, name, desc, price, date, source = entry[:8]
            unit = '工日'
            currency = 'THB'
            supplier = '-'
            remark = ''

        if group != current_group:
            current_group = group
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=16)
            cell = ws.cell(row=current_row, column=1, value=f"《{group}》")
            cell.font = group_font
            cell.fill = group_fill
            cell.alignment = left_align
            for col in range(1, 17):
                ws.cell(row=current_row, column=col).border = thin_border
                ws.cell(row=current_row, column=col).fill = group_fill
            current_row += 1

        values = [code, specialty, name, desc, unit, price, None, price, date, currency,
                  supplier, '-', '-', '-', remark, source]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            if col_idx in (1, 5, 6, 7, 8, 9, 10, 12, 13, 14):
                cell.alignment = center_align
            else:
                cell.alignment = left_align

        current_row += 1

    # Summary rows
    current_row += 1
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=16)
    total = len(data)
    is_short = len(data[0]) < 12
    price_idx = 5 if is_short else 6
    source_idx = 7 if is_short else 9
    prices = [e[price_idx] for e in data]
    sources = sorted(set(e[source_idx] for e in data))
    cell = ws.cell(row=current_row, column=1)
    cell.value = f"数据来源：{', '.join(sources)}"
    cell.font = Font(name='Microsoft YaHei', size=9)
    cell.alignment = left_align
    current_row += 1

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=16)
    cell = ws.cell(row=current_row, column=1)
    cell.value = f"价格统计：共{total}项，最低{min(prices)}，最高{max(prices)}，平均{sum(prices)/total:.0f}"
    cell.font = Font(name='Microsoft YaHei', size=9)
    cell.alignment = left_align

    wb.save(output_path)
    return total, min(prices), max(prices)


def main():
    parser = argparse.ArgumentParser(description="Build standardized quotation Excel from data file")
    parser.add_argument('--data', required=True, help='Python file defining ALL_DATA list')
    parser.add_argument('-o', '--output', default=None, help='Output xlsx path')
    parser.add_argument('--title', default='人材机价格表', help='Main title in row 1')
    parser.add_argument('--subtitle', default='', help='Subtitle in row 3')
    args = parser.parse_args()

    data = load_data_module(args.data)
    output = args.output or Path(args.data).with_suffix('.xlsx').name

    total, lo, hi = build_xlsx(data, output, args.title, args.subtitle)
    print(f"Generated: {output}")
    print(f"  Entries: {total}, Range: {lo} - {hi}")


if __name__ == '__main__':
    main()
