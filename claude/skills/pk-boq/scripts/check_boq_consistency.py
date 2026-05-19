#!/usr/bin/env python3
"""BOQ 清单一致性校验 —— 验证提取/复制的清单工程量是否与源文件一致。

场景：询价包制作后，校验从设计院 BOQ 提取到询价包清单的工程量是否一致。

两种用法：

1. 自动匹配模式（两张表 sheet 名称相同时）：
   python check_boq_consistency.py target.xlsx source.xlsx --qty-col 5

2. 显式映射模式（多源文件、sheet 名不同时）：
   python check_boq_consistency.py target.xlsx \
       -m "TargetSheet|source.xlsx|SourceSheet|5" \
       -m "TargetSheet2|source2.xlsx|SourceSheet2|3"
"""

import argparse
import sys
from pathlib import Path

import openpyxl


class Mapping:
    def __init__(self, target_sheet: str, source_path: str, source_sheet: str, qty_col: int):
        self.target_sheet = target_sheet
        self.source_path = source_path
        self.source_sheet = source_sheet
        self.qty_col = qty_col  # 0-based column index for quantity


def parse_mapping(arg: str) -> Mapping:
    parts = arg.split("|")
    if len(parts) != 4:
        raise ValueError(f"Invalid mapping format: {arg!r}. Expected: TargetSheet|source.xlsx|SourceSheet|qty_col")
    return Mapping(parts[0].strip(), parts[1].strip(), parts[2].strip(), int(parts[3].strip()))


def _is_item_row(val) -> bool:
    """Check if cell value looks like a BOQ item code (e.g. E.1, E.1.1.2ADD)."""
    if val is None:
        return False
    s = str(val).strip()
    if not s:
        return False
    # Match patterns like "E.1", "E.1.1", "E.1.1.2ADD", "E1.2ADD", "ADD"
    import re
    return bool(re.match(r'^[A-Z][\.\d]', s))


def _resolve_sheet(wb, name: str):
    """Resolve sheet name with tolerance for trailing spaces, case, and substring matching."""
    if name in wb.sheetnames:
        return wb[name]
    stripped = name.strip()
    if stripped in wb.sheetnames:
        return wb[stripped]
    lowered = stripped.lower()
    for sn in wb.sheetnames:
        if sn.strip().lower() == lowered:
            return wb[sn]
    # Substring match as last resort (shortest match first to avoid false positives)
    candidates = [(len(sn), sn) for sn in wb.sheetnames if lowered in sn.strip().lower()]
    if candidates:
        candidates.sort()
        return wb[candidates[0][1]]
    raise KeyError(f"Worksheet {name!r} does not exist. Available: {wb.sheetnames}")


def _to_float(v) -> float | None:
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def compare_sheets(ws_target, ws_source, qty_col: int, threshold: float = 1.0):
    """Compare quantities between two sheets using sequential matching.

    Items are matched by code AND occurrence order: the Nth occurrence of a code
    in target is compared against the Nth occurrence of the same code in source.
    This handles sheets where the same item codes repeat across sub-sections (e.g.
    E.6.1 appearing under multiple scour protection options).
    """
    discrepancies = []

    # Build ordered lists of source rows per item code (preserve all occurrences).
    # Sequential matching is critical: the same item code may repeat across
    # sub-options within a sheet (e.g. FHDI E.6 Scour Protection). The Nth
    # occurrence in target is matched to the Nth occurrence in source.
    src_items = {}
    for row in range(1, ws_source.max_row + 1):
        item_code = ws_source.cell(row=row, column=1).value
        if _is_item_row(item_code):
            code = str(item_code).strip()
            src_items.setdefault(code, []).append(row)

    src_ptr = {code: 0 for code in src_items}
    target_counts = {}

    for row in range(1, ws_target.max_row + 1):
        item_code = ws_target.cell(row=row, column=1).value
        if not _is_item_row(item_code):
            continue
        code = str(item_code).strip()
        target_counts[code] = target_counts.get(code, 0) + 1

        qty_target = _to_float(ws_target.cell(row=row, column=qty_col + 1).value)
        if qty_target is None:
            continue

        src_rows = src_items.get(code, [])
        ptr = src_ptr.get(code, 0)

        if ptr >= len(src_rows):
            discrepancies.append({
                'item': code,
                'desc': str(ws_target.cell(row=row, column=2).value or '')[:80],
                'target_qty': qty_target,
                'source_qty': None,
                'diff': None,
                'issue': 'MISSING_IN_SOURCE',
            })
            continue

        src_row = src_rows[ptr]
        src_ptr[code] = ptr + 1

        qty_src = _to_float(ws_source.cell(row=src_row, column=qty_col + 1).value)
        if qty_src is None:
            continue

        diff = abs(qty_target - qty_src)
        if diff > threshold:
            discrepancies.append({
                'item': code,
                'desc': str(ws_target.cell(row=row, column=2).value or '')[:80],
                'target_qty': qty_target,
                'source_qty': qty_src,
                'diff': diff,
                'issue': 'QUANTITY_MISMATCH',
            })

    # Check for items in source with fewer target occurrences
    for code, src_rows in src_items.items():
        tgt_count = target_counts.get(code, 0)
        for i in range(tgt_count, len(src_rows)):
            src_row = src_rows[i]
            qty_src = _to_float(ws_source.cell(row=src_row, column=qty_col + 1).value)
            discrepancies.append({
                'item': code,
                'desc': str(ws_source.cell(row=src_row, column=2).value or '')[:80],
                'target_qty': None,
                'source_qty': qty_src,
                'diff': None,
                'issue': 'MISSING_IN_TARGET',
            })

    return discrepancies


def check_consistency(target_path: str, mappings: list[Mapping], threshold: float = 1.0):
    """Run consistency check across all mappings.

    Returns: dict with summary and details.
    """
    wb_target = openpyxl.load_workbook(target_path, data_only=True)
    source_wbs = {}

    total_items = 0
    total_mismatches = 0
    all_results = []

    for m in mappings:
        if m.source_path not in source_wbs:
            source_wbs[m.source_path] = openpyxl.load_workbook(m.source_path, data_only=True)

        ws_target = _resolve_sheet(wb_target, m.target_sheet)
        ws_source = _resolve_sheet(source_wbs[m.source_path], m.source_sheet)

        discrepancies = compare_sheets(ws_target, ws_source, m.qty_col, threshold)

        item_count = sum(1 for row in range(1, ws_target.max_row + 1)
                         if _is_item_row(ws_target.cell(row=row, column=1).value))

        total_items += item_count
        total_mismatches += len(discrepancies)

        all_results.append({
            'target_sheet': m.target_sheet,
            'source_file': Path(m.source_path).name,
            'source_sheet': m.source_sheet,
            'items_checked': item_count,
            'mismatches': len(discrepancies),
            'details': discrepancies,
        })

    wb_target.close()
    for wb in source_wbs.values():
        wb.close()

    return {
        'total_items': total_items,
        'total_mismatches': total_mismatches,
        'results': all_results,
        'all_match': total_mismatches == 0,
    }


def format_report(result: dict) -> str:
    lines = []
    lines.append("=" * 60)
    lines.append("BOQ Consistency Check Report")
    lines.append("=" * 60)

    for r in result['results']:
        status = "PASS" if r['mismatches'] == 0 else "FAIL"
        lines.append(f"\n[{status}] {r['target_sheet']}")
        lines.append(f"  Source: {r['source_file']} / {r['source_sheet']}")
        lines.append(f"  Items checked: {r['items_checked']}, Mismatches: {r['mismatches']}")

        if r['mismatches'] > 0:
            lines.append(f"  --- Discrepancies ---")
            for d in r['details'][:20]:  # cap at 20 per sheet
                def _fmt(v):
                    return f"{v:,.2f}" if v is not None else "MISSING"
                if d['issue'] == 'QUANTITY_MISMATCH':
                    lines.append(f"  {d['item']}: {d['desc']}")
                    lines.append(f"    Target={_fmt(d['target_qty'])}  Source={_fmt(d['source_qty'])}  Diff={_fmt(d['diff'])}")
                elif d['issue'] == 'MISSING_IN_SOURCE':
                    lines.append(f"  {d['item']}: {d['desc']}")
                    lines.append(f"    Target={_fmt(d['target_qty'])}  Source=MISSING")
                elif d['issue'] == 'MISSING_IN_TARGET':
                    lines.append(f"  {d['item']}: {d['desc']}")
                    lines.append(f"    Target=MISSING  Source={_fmt(d['source_qty'])}")
            if r['mismatches'] > 20:
                lines.append(f"  ... and {r['mismatches'] - 20} more discrepancies")

    lines.append(f"\n{'=' * 60}")
    if result['all_match']:
        lines.append("RESULT: All quantities match. No discrepancies found.")
    else:
        lines.append(f"RESULT: {result['total_mismatches']} discrepancies found across {result['total_items']} items.")
    lines.append("=" * 60)
    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser(
        description="BOQ Consistency Checker - verify extracted BOQ quantities match source files",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # JSON config mode (recommended for non-ASCII sheet names)
  python check_boq_consistency.py target.xlsx --config mappings.json

  # Explicit mappings via CLI
  python check_boq_consistency.py target.xlsx \\
      -m "Sheet A|source.xlsx|Sheet B|5"

  # Auto-match (same sheet names, single source)
  python check_boq_consistency.py target.xlsx source.xlsx --qty-col 5
        """,
    )
    parser.add_argument("target", help="Target BOQ file to check")
    parser.add_argument("source", nargs="?", default=None, help="Single source file for auto-match mode")
    parser.add_argument("-m", "--map", dest="mappings", action="append", default=[],
                        help="Mapping: TargetSheet|source.xlsx|SourceSheet|qty_col (0-based)")
    parser.add_argument("--config", type=str, default=None,
                        help="JSON config file with mappings (avoid encoding issues with non-ASCII sheet names)")
    parser.add_argument("--qty-col", type=int, default=None,
                        help="Quantity column index (0-based) for auto-match mode")
    parser.add_argument("-t", "--threshold", type=float, default=1.0,
                        help="Difference threshold to flag as discrepancy (default: 1.0)")
    parser.add_argument("--json", action="store_true", help="Output as JSON")

    args = parser.parse_args()

    mappings = [parse_mapping(m) for m in args.mappings]

    if args.config:
        import json
        with open(args.config, "r", encoding="utf-8") as f:
            config = json.load(f)
        for m in config.get("mappings", []):
            mappings.append(Mapping(m["target_sheet"], m["source_file"], m["source_sheet"], m["qty_col"]))
        # Allow overriding threshold from config
        if "threshold" in config and args.threshold == 1.0:
            args.threshold = config["threshold"]

    # Auto-match mode: find sheets with matching names
    if args.source and not mappings:
        if args.qty_col is None:
            print("Error: --qty-col required for auto-match mode", file=sys.stderr)
            sys.exit(1)

        wb_t = openpyxl.load_workbook(args.target, data_only=True)
        wb_s = openpyxl.load_workbook(args.source, data_only=True)

        # Build a lookup: for each target sheet, find best matching source sheet
        for ts in wb_t.sheetnames:
            try:
                src_ws = _resolve_sheet(wb_s, ts)
                mappings.append(Mapping(ts, args.source, src_ws.title, args.qty_col))
            except KeyError:
                pass

        wb_t.close()
        wb_s.close()

        if not mappings:
            print("Error: No matching sheet names found between target and source.", file=sys.stderr)
            print(f"  Target sheets: {wb_t.sheetnames}", file=sys.stderr)
            print(f"  Source sheets: {wb_s.sheetnames}", file=sys.stderr)
            sys.exit(1)

    if not mappings:
        print("Error: No mappings specified. Use --config, -m, or provide a source file for auto-match.", file=sys.stderr)
        sys.exit(1)

    result = check_consistency(args.target, mappings, args.threshold)

    if args.json:
        import json
        print(json.dumps(result, indent=2, ensure_ascii=False, default=str))
    else:
        print(format_report(result))

    sys.exit(0 if result['all_match'] else 1)


if __name__ == "__main__":
    main()
