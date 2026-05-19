# -*- coding: utf-8 -*-
"""多家设计院 BOQ 清单工程量对比分析

Usage:
    python compare_boq.py --wtcc <path> --fhdi <path> --sghcc <path> [-o output.md]

    也可作为模块导入:
    from compare_boq import BOQComparator
    comp = BOQComparator()
    comp.compare(wtcc_path, fhdi_path, sghcc_path, output_path)
"""
import argparse
import re
from datetime import datetime
from pathlib import Path
from collections import defaultdict
from difflib import SequenceMatcher

import openpyxl

# ── 正则 ──────────────────────────────────────────────
CLASS_RE = re.compile(r"^Class\s+([A-I])\b", re.IGNORECASE)
ITEM_CODE_RE = re.compile(r"^([A-I]\.\d+(?:\.\d+)*)\b")
OPTION_RE = re.compile(r"[（(]方案([一二三四五六七八九十]+)\s*[）)]")
ADD_RE = re.compile(r"\bADD\b", re.IGNORECASE)

# 三院简称
INSTITUTES = {"FHDI": "一航院", "SGHCC": "水规院", "WTCC": "三航院"}

def safe_float(v):
    """安全转 float"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v) if v == v else None  # NaN check
    try:
        return float(str(v).strip().replace(",", "").replace(" ", ""))
    except (ValueError, TypeError):
        return None


def normalize_unit(u):
    """归一化单位"""
    if u is None:
        return None
    u = str(u).strip().lower()
    mapping = {
        "m3": "m³", "m2": "m²", "no.": "No.", "nos": "No.", "nr.": "No.",
        "ton": "t", "tonne": "t", "item": "Item", "ls": "LS", "l.s.": "LS",
        "lump sum": "LS", "lin.m": "m", "lm": "m", "rm": "m",
        "kg": "kg", "set": "set", "month": "month", "day": "day",
        "hour": "hr", "week": "wk", "mm": "mm", "cm": "cm",
    }
    for k, v in mapping.items():
        if u == k:
            return v
    if u in ("m³", "m²", "m", "No.", "t", "LS", "Item", "kg", "set", "month", "day", "hr", "wk", "mm", "cm", "%"):
        return u
    return u


def parse_item_code(code_str):
    """解析 item code，返回 (cleaned_code, is_add)"""
    if code_str is None or not isinstance(code_str, str):
        return None, False
    code = code_str.strip()
    is_add = bool(ADD_RE.search(code))
    # 去掉 ADD 后缀用于匹配
    clean = ADD_RE.sub("", code).strip().rstrip(".-").rstrip()
    return clean, is_add


def get_option_number(desc):
    """从描述中提取方案编号"""
    if desc is None:
        return None
    m = OPTION_RE.search(str(desc))
    if not m:
        return None
    num_map = {"一": 1, "二": 2, "三": 3, "四": 4, "五": 5}
    return num_map.get(m.group(1), None)

# ═══════════════════════════════════════════════════════
# 文件读取与解析
# ═══════════════════════════════════════════════════════

def read_fhdi(path):
    """读取 FHDI 文件 (16列, A-P)"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["MergeSheet"]
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        vals = list(row) + [None] * (16 - len(row))
        item_code = str(vals[0]).strip() if vals[0] is not None else ""
        desc = str(vals[1]).strip() if vals[1] is not None else ""
        unit = str(vals[2]).strip() if vals[2] is not None else ""
        qty = safe_float(vals[3])

        # 跳过空行和表头行 ("Item" in col A)
        if not item_code and not desc and qty is None:
            continue
        if item_code.lower() == "item":
            continue

        rows.append({
            "row": row_idx, "item_code": item_code, "desc": desc,
            "unit": normalize_unit(unit), "qty": qty,
        })
    wb.close()
    return rows


def read_sghcc(path):
    """读取 SGHCC 文件 (19列, A-S)"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["MergeSheet"]
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        vals = list(row) + [None] * (19 - len(row))
        item_code = str(vals[0]).strip() if vals[0] is not None else ""
        desc = str(vals[1]).strip() if vals[1] is not None else ""
        unit = str(vals[2]).strip() if vals[2] is not None else ""
        qty = safe_float(vals[3])

        if not item_code and not desc and qty is None:
            continue
        if item_code.lower() == "item":
            continue

        rows.append({
            "row": row_idx, "item_code": item_code, "desc": desc,
            "unit": normalize_unit(unit), "qty": qty,
        })
    wb.close()
    return rows


def read_wtcc(path):
    """读取 WTCC 文件 (6列, A-F). F列=设计量, D列=报价量"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["MergeSheet"]
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        vals = list(row) + [None] * (6 - len(row))
        item_code = str(vals[0]).strip() if vals[0] is not None else ""
        desc = str(vals[1]).strip() if vals[1] is not None else ""
        unit = str(vals[2]).strip() if vals[2] is not None else ""
        bid_qty = safe_float(vals[3])    # D列: 报价量
        design_qty = safe_float(vals[5])  # F列: 设计量

        if not item_code and not desc and bid_qty is None and design_qty is None:
            continue
        if item_code.lower() == "item":
            continue

        rows.append({
            "row": row_idx, "item_code": item_code, "desc": desc,
            "unit": normalize_unit(unit), "qty": design_qty,  # 用设计量
            "bid_qty": bid_qty,
        })
    wb.close()
    return rows

# ═══════════════════════════════════════════════════════
# 行分类与结构化
# ═══════════════════════════════════════════════════════

def classify_rows(raw_rows, source):
    """分类所有行: CLASS header / Section header / BOQ item"""
    current_class = None
    current_section = None
    current_option = None  # FHDI only

    classified = []
    for r in raw_rows:
        code = r["item_code"]
        desc = r["desc"]
        unit = r["unit"]
        qty = r["qty"]

        # Class header
        cm = CLASS_RE.match(code)
        if cm:
            current_class = cm.group(1).upper()
            current_section = None
            # FHDI 方案检测
            if source == "FHDI":
                opt = get_option_number(desc)
                if opt:
                    current_option = opt
            classified.append({**r, "class_id": current_class, "section_id": None,
                               "is_class_header": True, "is_section_header": False,
                               "is_item": False, "option": current_option, "source": source})
            continue

        # 每类开始重置 option (FHDI)
        if source == "FHDI" and current_class:
            opt = get_option_number(desc)
            if opt:
                current_option = opt

        # 无 CLASS 则跳过
        if current_class is None:
            continue

        # Section / Sub-section header: 有层级code但无单位无数量
        code_match = ITEM_CODE_RE.match(code) if code else None
        is_header = (unit is None or unit == "") and qty is None
        is_without_children = code_match and not code_match.group(1).endswith("0")  # heuristic

        # 判断是否为 header
        if code and is_header:
            parts = code.split(".")
            if len(parts) <= 3:  # e.g. C.1, D.2.4
                current_section = code
                classified.append({**r, "class_id": current_class, "section_id": current_section,
                                   "is_class_header": False, "is_section_header": True,
                                   "is_item": False, "option": current_option, "source": source})
                continue

        # BOQ item
        clean_code, is_add = parse_item_code(code)
        classified.append({**r, "class_id": current_class, "section_id": current_section,
                           "is_class_header": False, "is_section_header": False,
                           "is_item": True, "option": current_option, "source": source,
                           "is_add": is_add, "clean_code": clean_code})

    return classified


def filter_items(classified):
    """只保留 is_item=True 的条目，FHDI 只保留方案一"""
    items = [r for r in classified if r["is_item"]]

    # FHDI 去重: 同一 item_code 如果有 option=1 和 option=2，只保留 option=1
    fhdi_items = [r for r in items if r["source"] == "FHDI"]
    other_items = [r for r in items if r["source"] != "FHDI"]

    seen = {}
    deduped_fhdi = []
    for r in fhdi_items:
        key = (r["clean_code"] or r["item_code"], r["class_id"])
        if key in seen:
            prev = seen[key]
            # 优先保留 option 1，其次保留有数量者
            if r.get("option") == 1 and prev.get("option") != 1:
                deduped_fhdi.remove(prev)
                deduped_fhdi.append(r)
                seen[key] = r
            elif r.get("qty") is not None and prev.get("qty") is None:
                deduped_fhdi.remove(prev)
                deduped_fhdi.append(r)
                seen[key] = r
        else:
            deduped_fhdi.append(r)
            seen[key] = r

    return deduped_fhdi + other_items

# ═══════════════════════════════════════════════════════
# 条目匹配引擎
# ═══════════════════════════════════════════════════════

def desc_similarity(a, b):
    """计算两个描述的相似度"""
    if not a or not b:
        return 0.0
    # 预处理
    def clean(s):
        s = s.lower()
        s = re.sub(r"[（(].*?[）)]", "", s)  # 去掉中文括号内容
        s = re.sub(r"\{.*?\}", "", s)         # 去掉花括号内容
        s = re.sub(r"[^\w\s]", " ", s)        # 标点变空格
        s = re.sub(r"\s+", " ", s).strip()
        return s
    return SequenceMatcher(None, clean(a), clean(b)).ratio()


def build_concept_key(item):
    """为条目建立概念键: class + 前3个描述词"""
    cls = item.get("class_id", "")
    desc = item.get("desc", "")
    words = re.findall(r"[a-zA-Z]{3,}", desc.lower())[:3]
    return f"{cls}:{'_'.join(words)}"


def match_items(wtcc_items, other_items, other_source):
    """
    将 other (FHDI/SGHCC) 条目匹配到 WTCC 条目。
    返回: {wtcc_idx: [(other_item, confidence), ...]} 和 unmatched other items
    """
    matches = defaultdict(list)
    unmatched = []

    # 按 (class_id, section_prefix) 分组
    wtcc_by_section = defaultdict(list)
    for i, item in enumerate(wtcc_items):
        cls = item["class_id"]
        sec = item.get("section_id", "") or ""
        # 用 section 的前两级作为 key
        sec_parts = sec.split(".") if sec else []
        sec_key = ".".join(sec_parts[:2]) if len(sec_parts) >= 2 else sec
        wtcc_by_section[(cls, sec_key)].append((i, item))

    for oitem in other_items:
        cls = oitem["class_id"]
        osec = oitem.get("section_id", "") or ""
        osec_parts = osec.split(".") if osec else []
        osec_key = ".".join(osec_parts[:2]) if len(osec_parts) >= 2 else osec

        # 候选集: 同 section key; 如为空则用整个 class
        key = (cls, osec_key)
        candidates = wtcc_by_section.get(key, [])
        if not candidates:
            candidates = wtcc_by_section.get((cls, ""), [])
        if not candidates:
            # 最后回退到整个 class
            for (c, sk), items in wtcc_by_section.items():
                if c == cls:
                    candidates.extend(items)
        if not candidates:
            unmatched.append(oitem)
            continue

        clean_ocode = oitem.get("clean_code", "")

        # 1. 精确 code 匹配
        best_match = None
        best_confidence = 0.0
        for wi, witem in candidates:
            wcode = witem.get("clean_code", "")
            if clean_ocode and wcode and clean_ocode == wcode:
                best_match = wi
                best_confidence = 1.0
                break

        # 2. 父级 code 匹配
        if best_match is None:
            for wi, witem in candidates:
                wcode = witem.get("clean_code", "")
                if clean_ocode and wcode and wcode.startswith(clean_ocode + "."):
                    best_match = wi
                    best_confidence = 0.85
                    break

        # 3. 描述相似度匹配 (提高阈值到 0.75)
        if best_match is None:
            for wi, witem in candidates:
                sim = desc_similarity(oitem.get("desc", ""), witem.get("desc", ""))
                if sim > best_confidence:
                    best_confidence = sim
                    best_match = wi

            if best_confidence < 0.75:
                best_match = None

        if best_match is not None:
            matches[best_match].append((oitem, best_confidence))
        else:
            unmatched.append(oitem)

    return matches, unmatched

# ═══════════════════════════════════════════════════════
# 报告生成
# ═══════════════════════════════════════════════════════

CLASS_NAMES = {
    "B": "Preparatory Works 准备工作",
    "C": "Dredging 疏浚工程",
    "D": "Reclamation, Ground Improvement and Earthworks 回填、地基处理与土方",
    "E": "Quay 码头工程",
    "F": "Yard Civil & Structures 堆场土建与结构",
    "G": "Electrical, IT & Mechanical 电气、IT与机械",
    "H": "Buildings 建筑物",
    "I": "Gate Complex 大门综合体",
}


def format_qty(v):
    """格式化数量显示"""
    if v is None:
        return "-"
    if isinstance(v, float):
        if abs(v) >= 1e6:
            return f"{v:,.0f}"
        elif abs(v) >= 1e3:
            return f"{v:,.1f}"
        elif abs(v) >= 1:
            return f"{v:,.2f}"
        else:
            return f"{v:.4f}"
    return str(v)


def calc_diff(vals):
    """计算差异百分比: (max-min)/max * 100，至少两个有效值"""
    valid = [v for v in vals if v is not None and v > 0]
    if len(valid) < 2:
        return None
    mx, mn = max(valid), min(valid)
    if mx == 0:
        return None
    return (mx - mn) / mx * 100


def build_comparison_rows(wtcc_items, fhdi_items, sghcc_items):
    """构建对比行"""
    # 匹配 FHDI → WTCC, SGHCC → WTCC
    fhdi_matches, fhdi_unmatched = match_items(wtcc_items, fhdi_items, "FHDI")
    sghcc_matches, sghcc_unmatched = match_items(wtcc_items, sghcc_items, "SGHCC")

    comp_rows = []

    # 遍历 WTCC 条目
    all_matched_wtcc = set(fhdi_matches.keys()) | set(sghcc_matches.keys())
    for wi, witem in enumerate(wtcc_items):
        if witem["qty"] is None:
            continue  # 跳过无设计量的条目

        cls = witem["class_id"]
        # 取最佳 FHDI 匹配
        fhdi_qty = None
        if wi in fhdi_matches:
            # 取 confidence 最高者
            best = max(fhdi_matches[wi], key=lambda x: x[1])
            fhdi_qty = best[0]["qty"]

        # 取最佳 SGHCC 匹配
        sghcc_qty = None
        if wi in sghcc_matches:
            best = max(sghcc_matches[wi], key=lambda x: x[1])
            sghcc_qty = best[0]["qty"]

        diff = calc_diff([fhdi_qty, sghcc_qty, witem["qty"]])
        is_add = witem.get("is_add", False) or "ADD" in (witem.get("item_code", ""))
        notes = []
        if is_add:
            notes.append("ADD")
        # 标记哪些设计院无此项
        missing = []
        if fhdi_qty is None:
            missing.append("FHDI无")
        if sghcc_qty is None:
            missing.append("SGHCC无")
        if missing:
            notes.append("; ".join(missing))

        comp_rows.append({
            "class_id": cls,
            "section_id": witem.get("section_id", ""),
            "item_code": witem.get("item_code", ""),
            "desc": witem.get("desc", ""),
            "unit": witem.get("unit", ""),
            "fhdi_qty": fhdi_qty,
            "sghcc_qty": sghcc_qty,
            "wtcc_qty": witem["qty"],
            "diff_pct": diff,
            "notes": ", ".join(notes),
            "is_add": is_add,
        })

    # 添加 FHDI 独有项 (WTCC 和 SGHCC 均无)
    for oitem in fhdi_unmatched:
        if oitem["qty"] is None:
            continue
        comp_rows.append({
            "class_id": oitem["class_id"],
            "section_id": oitem.get("section_id", ""),
            "item_code": oitem.get("item_code", ""),
            "desc": oitem.get("desc", ""),
            "unit": oitem.get("unit", ""),
            "fhdi_qty": oitem["qty"],
            "sghcc_qty": None,
            "wtcc_qty": None,
            "diff_pct": None,
            "notes": "FHDI独有",
            "is_add": False,
        })

    # 添加 SGHCC 独有项 (WTCC 和 FHDI 均无)
    for oitem in sghcc_unmatched:
        if oitem["qty"] is None:
            continue
        comp_rows.append({
            "class_id": oitem["class_id"],
            "section_id": oitem.get("section_id", ""),
            "item_code": oitem.get("item_code", ""),
            "desc": oitem.get("desc", ""),
            "unit": oitem.get("unit", ""),
            "fhdi_qty": None,
            "sghcc_qty": oitem["qty"],
            "wtcc_qty": None,
            "diff_pct": None,
            "notes": "SGHCC独有",
            "is_add": False,
        })

    return comp_rows

# ═══════════════════════════════════════════════════════
# Markdown 输出
# ═══════════════════════════════════════════════════════

def generate_summary_table(comp_rows):
    """生成总体概览表"""
    lines = []
    lines.append("## 1. 总体概览\n")
    lines.append("| Class | 分部名称 | FHDI条目 | SGHCC条目 | WTCC条目 | ADD项 | 差异>20%项 | 关键观察 |")
    lines.append("|-------|---------|----------|-----------|----------|-------|-----------|---------|")

    for cls in ["B", "C", "D", "E", "F", "G", "H", "I"]:
        crs = [r for r in comp_rows if r["class_id"] == cls]
        if not crs:
            continue
        fhdi_n = len([r for r in crs if r["fhdi_qty"] is not None])
        sghcc_n = len([r for r in crs if r["sghcc_qty"] is not None])
        wtcc_n = len([r for r in crs if r["wtcc_qty"] is not None])
        add_n = len([r for r in crs if r["is_add"]])
        gt20_n = len([r for r in crs if r["diff_pct"] is not None and r["diff_pct"] > 20])

        # 关键观察
        observations = []
        if gt20_n > 0:
            observations.append(f"{gt20_n}项差异>20%")
        if add_n > 0:
            observations.append(f"{add_n}项ADD")
        obs = "; ".join(observations) if observations else "基本一致"

        lines.append(f"| {cls} | {CLASS_NAMES.get(cls, cls)} | {fhdi_n} | {sghcc_n} | {wtcc_n} | {add_n} | {gt20_n} | {obs} |")

    return "\n".join(lines)


def generate_class_section(comp_rows, cls):
    """生成单个 Class 的对比表"""
    crs = [r for r in comp_rows if r["class_id"] == cls]
    if not crs:
        return ""

    lines = []
    title = CLASS_NAMES.get(cls, f"Class {cls}")
    lines.append(f"## {list(CLASS_NAMES.keys()).index(cls) + 2}. Class {cls} - {title}\n")

    # 统计
    fhdi_n = len([r for r in crs if r["fhdi_qty"] is not None])
    sghcc_n = len([r for r in crs if r["sghcc_qty"] is not None])
    wtcc_n = len([r for r in crs if r["wtcc_qty"] is not None])
    add_n = len([r for r in crs if r["is_add"]])

    lines.append(f"**该类条目统计**: FHDI {fhdi_n} 项 | SGHCC {sghcc_n} 项 | WTCC {wtcc_n} 项 | ADD {add_n} 项\n")

    # 表头
    lines.append("| # | Item Code | 项目描述 | Unit | FHDI 一航院 | SGHCC 水规院 | WTCC 三航院 | 差异% | 备注 |")
    lines.append("|---|-----------|---------|------|------------|-------------|------------|------|------|")

    row_num = 0
    for r in crs:
        row_num += 1
        code = r["item_code"]
        if r["is_add"] and "[ADD]" not in r["notes"]:
            code = f"{code} [ADD]"
        diff_str = f"**{r['diff_pct']:.0f}%**" if r["diff_pct"] is not None and r["diff_pct"] > 20 else (
            f"{r['diff_pct']:.0f}%" if r["diff_pct"] is not None else "-")
        lines.append(
            f"| {row_num} | {code} | {r['desc'][:80]} | {r['unit']} | "
            f"{format_qty(r['fhdi_qty'])} | {format_qty(r['sghcc_qty'])} | {format_qty(r['wtcc_qty'])} | "
            f"{diff_str} | {r['notes']} |"
        )

    return "\n".join(lines)


def generate_key_findings(comp_rows):
    """生成关键发现"""
    lines = []
    lines.append("## 10. 关键发现\n")

    # 差异最大的10项
    all_with_diff = [r for r in comp_rows if r["diff_pct"] is not None]
    all_with_diff.sort(key=lambda x: x["diff_pct"], reverse=True)
    top10 = all_with_diff[:10]

    if top10:
        lines.append("### 工程量差异最大的 10 项\n")
        lines.append("| # | Item Code | 项目描述 | Unit | FHDI | SGHCC | WTCC | 差异% |")
        lines.append("|---|-----------|---------|------|------|-------|------|------|")
        for i, r in enumerate(top10, 1):
            lines.append(
                f"| {i} | {r['item_code']} | {r['desc'][:60]} | {r['unit']} | "
                f"{format_qty(r['fhdi_qty'])} | {format_qty(r['sghcc_qty'])} | {format_qty(r['wtcc_qty'])} | "
                f"**{r['diff_pct']:.0f}%** |"
            )

    # 各 Class 独有项统计
    lines.append("\n### 各院独有清单项统计\n")
    lines.append("| Class | FHDI 独有 | SGHCC 独有 | WTCC 独有 |")
    lines.append("|-------|----------|-----------|----------|")
    for cls in ["B", "C", "D", "E", "F", "G", "H", "I"]:
        crs = [r for r in comp_rows if r["class_id"] == cls]
        fhdi_only = len([r for r in crs if "FHDI无" not in r["notes"] and r["sghcc_qty"] is None and r["wtcc_qty"] is None and r["fhdi_qty"] is not None])
        sghcc_only = len([r for r in crs if "SGHCC无" not in r["notes"] and r["fhdi_qty"] is None and r["wtcc_qty"] is None and r["sghcc_qty"] is not None])
        wtcc_only = len([r for r in crs if r["fhdi_qty"] is None and r["sghcc_qty"] is None and r["wtcc_qty"] is not None])
        if fhdi_only or sghcc_only or wtcc_only:
            lines.append(f"| {cls} | {fhdi_only} | {sghcc_only} | {wtcc_only} |")

    return "\n".join(lines)

# ═══════════════════════════════════════════════════════
# BOQComparator 类
# ═══════════════════════════════════════════════════════

FH_CN = {"FHDI": "一航院 (FHDI)", "SGHCC": "水规院 (SGHCC)", "WTCC": "三航院 (WTCC)"}

INSTITUTE_LABELS = {"FHDI": "一航院(FHDI)", "SGHCC": "水规院(SGHCC)", "WTCC": "三航院(WTCC)"}


class BOQComparator:
    """多家设计院 BOQ 清单对比分析器。

    Usage:
        comp = BOQComparator(project_name="My Project")
        comp.compare(wtcc_path, fhdi_path, sghcc_path, output_path)
    """

    def __init__(self, project_name: str = "", report_date: str = ""):
        self.project_name = project_name
        self.report_date = report_date or datetime.now().strftime("%Y-%m-%d")

    def compare(self, wtcc_path, fhdi_path, sghcc_path, output_path=None):
        """执行对比分析，生成 Markdown 报告。

        Args:
            wtcc_path: WTCC(三航院) BQMerge xlsx 文件路径（作为基准）
            fhdi_path: FHDI(一航院) BQMerge xlsx 文件路径
            sghcc_path: SGHCC(水规院) BQMerge xlsx 文件路径
            output_path: 输出 .md 报告路径，默认为 {cwd}/{date}_BOQ_Comparison_Report.md

        Returns:
            输出文件路径
        """
        if output_path is None:
            output_path = Path.cwd() / f"{self.report_date}_BOQ_Comparison_Report.md"
        else:
            output_path = Path(output_path)

        print("=" * 70)
        print("多家设计院 BOQ 清单工程量对比分析")
        print("=" * 70)

        # 1. 读取
        print("\n[1/5] 读取文件...")
        fhdi_raw = read_fhdi(fhdi_path)
        sghcc_raw = read_sghcc(sghcc_path)
        wtcc_raw = read_wtcc(wtcc_path)
        print(f"  FHDI={len(fhdi_raw)}, SGHCC={len(sghcc_raw)}, WTCC={len(wtcc_raw)}")

        # 2. 分类
        print("\n[2/5] 分类行...")
        fhdi_cls = classify_rows(fhdi_raw, "FHDI")
        sghcc_cls = classify_rows(sghcc_raw, "SGHCC")
        wtcc_cls = classify_rows(wtcc_raw, "WTCC")

        fhdi_items = filter_items(fhdi_cls)
        sghcc_items = filter_items(sghcc_cls)
        wtcc_items = filter_items(wtcc_cls)
        print(f"  BOQ条目: FHDI={len(fhdi_items)}, SGHCC={len(sghcc_items)}, WTCC={len(wtcc_items)}")

        for src, items in [("FHDI", fhdi_items), ("SGHCC", sghcc_items), ("WTCC", wtcc_items)]:
            cls_dist = defaultdict(int)
            for it in items:
                cls_dist[it["class_id"]] += 1
            dist_str = ", ".join(f"{k}:{v}" for k, v in sorted(cls_dist.items()))
            print(f"  {src} Class分布: {dist_str}")

        # 3. 构建对比行
        print("\n[3/5] 匹配条目...")
        comp_rows = build_comparison_rows(wtcc_items, fhdi_items, sghcc_items)
        print(f"  生成 {len(comp_rows)} 个对比行")

        cls_counts = defaultdict(int)
        for r in comp_rows:
            cls_counts[r["class_id"]] += 1
        for cls in ["B", "C", "D", "E", "F", "G", "H", "I"]:
            if cls_counts[cls]:
                print(f"  Class {cls}: {cls_counts[cls]} 对比行")

        # 4. 排序
        comp_rows.sort(key=lambda r: (r["class_id"] or "Z", r["section_id"] or "", r["item_code"] or ""))

        # 5. 生成报告
        print("\n[4/5] 生成 Markdown 报告...")
        report = self._build_report(comp_rows)

        content = "\n".join(report)
        output_path.write_text(content, encoding="utf-8")
        print(f"\n[5/5] 报告已生成: {output_path}")
        print(f"  文件大小: {output_path.stat().st_size:,} bytes")
        return str(output_path)

    def _build_report(self, comp_rows):
        report = []
        if self.project_name:
            report.append(f"# {self.project_name}")
        report.append("# 多家设计院 BOQ 清单工程量对比分析报告")
        report.append("")
        report.append(f"> **日期**: {self.report_date}")
        sources = set()
        for r in comp_rows:
            if r.get("fhdi_qty") is not None:
                sources.add(INSTITUTE_LABELS["FHDI"])
            if r.get("sghcc_qty") is not None:
                sources.add(INSTITUTE_LABELS["SGHCC"])
            if r.get("wtcc_qty") is not None:
                sources.add(INSTITUTE_LABELS["WTCC"])
        report.append(f"> **数据来源**: {', '.join(sorted(sources))} 初步设计方案工程量清单")
        report.append("> **比较基准**: 以 WTCC(三航院) 清单体系为基准，匹配 FHDI 和 SGHCC 对应条目")
        report.append("> **注意事项**: WTCC 数据采用 F列(设计工程量)；FHDI 采用方案一数据")
        report.append("")

        report.append(generate_summary_table(comp_rows))
        report.append("")

        for cls in ["B", "C", "D", "E", "F", "G", "H", "I"]:
            print(f"  生成 Class {cls}...")
            section = generate_class_section(comp_rows, cls)
            if section:
                report.append(section)
                report.append("")

        report.append(generate_key_findings(comp_rows))
        report.append("")

        report.append("---\n")
        report.append("## 附录：图例说明\n")
        report.append("| 标记 | 含义 |")
        report.append("|------|------|")
        report.append("| `[ADD]` | 设计院在标准BOQ外新增的清单条目 |")
        report.append("| `-` | 该设计院清单中无此条目或无数量 |")
        report.append("| **差异%** | 各院中(最大值-最小值)/最大值×100%，仅2个及以上有效值时计算 |")
        report.append("| `XXX独有` | 仅该院清单有此条目 |")
        report.append("| `XXX无` | 该院清单中无此条目 |")
        report.append("| 差异>20% 加粗 | 表示各院之间存在显著工程量偏差 |")
        return report


# ═══════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="多家设计院 BOQ 清单工程量对比分析",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python compare_boq.py --wtcc WTCC_BQMerge.xlsx --fhdi FHDI_BQMerge.xlsx --sghcc SGHCC_BQMerge.xlsx
  python compare_boq.py --wtcc WTCC.xlsx --fhdi FHDI.xlsx --sghcc SGHCC.xlsx -o report.md
  python compare_boq.py --wtcc WTCC.xlsx --fhdi FHDI.xlsx --sghcc SGHCC.xlsx --project "My Project"
        """,
    )
    parser.add_argument("--wtcc", required=True, help="WTCC(三航院) BQMerge xlsx 文件路径（作为基准）")
    parser.add_argument("--fhdi", required=True, help="FHDI(一航院) BQMerge xlsx 文件路径")
    parser.add_argument("--sghcc", required=True, help="SGHCC(水规院) BQMerge xlsx 文件路径")
    parser.add_argument("-o", "--output", default=None, help="输出 .md 报告路径（默认: {date}_BOQ_Comparison_Report.md）")
    parser.add_argument("--project", default="", help="项目名称（显示在报告标题）")
    parser.add_argument("--date", default="", help="报告日期（默认今日）")
    args = parser.parse_args()

    comp = BOQComparator(project_name=args.project, report_date=args.date)
    comp.compare(args.wtcc, args.fhdi, args.sghcc, args.output)


if __name__ == "__main__":
    main()
