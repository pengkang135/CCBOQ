#!/usr/bin/env python3
"""
Extract BOQ items matching a keyword from a merged BOQ file,
preserve hierarchy, and apply template styling with row grouping.

4-level hierarchy with collapsible outline groups:
  L1 【】section delimiter → L2 Class header → L3 subsection → L4 item

Usage:
    python extract_boq_by_keyword.py <source.xlsx> <keyword> <template.xlsx> [-o output.xlsx]
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color
from openpyxl.utils import get_column_letter
import re
import argparse
import os
from datetime import datetime

COL_MAP = {0: 0, 1: 1, 2: 2, 3: 3, 7: 4, 8: 5, 9: 6, 10: 7, 11: 8, 12: 9, 13: 10, 14: 11, 15: 12}
NUM_COLS = {'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M'}
NUM_FMT = '_ * #,##0.00_ ;_ * \\-#,##0.00_ ;_ * "-"??_ ;_ @_ '

# 5-level hierarchy: 【】 L1 → 《》 L2 → {} L3 → plain L4 → ADD/sub-item L5
# Outline: L1=0, L2=1, L3=2, L4=3, L5=3 (L4 and L5 both at deepest level)
TYPE_LEVEL = {'section': 1, 'class_header': 2, 'subsection': 3, 'item': 4}


def is_curly_header(val):
    """Check if B-column value is a {} third-level subsection header."""
    if not val:
        return False
    s = str(val).strip()
    return s.startswith('{')


def is_angle_header(val):
    """Check if B-column value is a 《》 second-level section header."""
    if not val:
        return False
    s = str(val).strip()
    return s.startswith('《')


def is_class_header(val):
    if not val:
        return False
    return bool(re.match(r'^Class\s+[A-Z]', str(val).strip()))


def is_valid_item(val):
    if not val:
        return False
    return bool(re.match(r'^[A-Z]\.\d+', str(val).strip()))


def is_section_delimiter(a_val, b_val):
    """Section delimiter: 【... in A or B."""
    for v in (a_val, b_val):
        if v and str(v).strip().startswith('【'):
            return True
    return False


def item_prefix_match(parent_a, child_a):
    if not parent_a or not child_a:
        return False
    p = parent_a.strip().replace('ADD', '')
    c = child_a.strip()
    return c.startswith(p + '.')


def count_dots(s):
    return s.count('.')


class BOQExtractor:
    def __init__(self, source_path, keyword, template_path):
        self.source_path = source_path
        self.keyword = keyword.lower()
        self.template_path = template_path
        self.ws_s = None
        self.output_rows = []

    def load_source(self):
        wb = openpyxl.load_workbook(self.source_path, data_only=True)
        self.ws_s = wb.active
        return self

    def find_markers(self):
        """Find all rows where B column matches keyword and A column is a valid item."""
        markers = []
        for row in range(1, self.ws_s.max_row + 1):
            a_val = self.ws_s.cell(row=row, column=1).value
            b_val = self.ws_s.cell(row=row, column=2).value
            if not b_val:
                continue
            if self.keyword not in str(b_val).strip().lower():
                continue
            if is_class_header(a_val):
                continue
            if not is_valid_item(a_val):
                continue
            markers.append((row, str(a_val).strip()))
        return markers

    def find_ancestors(self, marker_row, marker_a):
        """Find ancestor rows (section delimiter, class header, parent subsections)."""
        ancestors = []
        for row in range(marker_row - 1, 0, -1):
            a_val = self.ws_s.cell(row=row, column=1).value
            b_val = self.ws_s.cell(row=row, column=2).value

            # Section delimiter (check both cols)
            if is_section_delimiter(a_val, b_val):
                # Use B value for section name if A is a delimiter
                section_name = ''
                if a_val and str(a_val).strip().startswith('【'):
                    section_name = str(a_val).strip()
                elif b_val and str(b_val).strip().startswith('【'):
                    section_name = str(b_val).strip()
                ancestors.append(('section', row, section_name, str(b_val) if b_val else ''))
                break

            if not a_val:
                continue

            a_str = str(a_val).strip()

            # Class header: only include if class letter matches marker
            if is_class_header(a_str):
                class_letter = a_str.split()[-1]
                marker_letter = marker_a[0]
                if class_letter == marker_letter:
                    ancestors.append(('class_header', row, a_str, str(b_val) if b_val else ''))
                continue

            # Valid item: include only if it's a true prefix (parent) of marker
            if is_valid_item(a_str):
                if item_prefix_match(a_str, marker_a):
                    ancestors.append(('subsection', row, a_str, str(b_val) if b_val else ''))

        ancestors.reverse()
        return ancestors

    def find_children(self, marker_row, marker_a, next_marker_row=None):
        """Find all child items of a marker. Stop at next_marker_row or same/higher level item."""
        children = []
        end_row = next_marker_row if next_marker_row else self.ws_s.max_row + 1
        m_dots = count_dots(marker_a.replace('ADD', ''))

        for row in range(marker_row + 1, end_row):
            a_val = self.ws_s.cell(row=row, column=1).value
            if not a_val:
                continue
            a_str = str(a_val).strip()

            # Only process valid items and ADD items
            is_add = (a_str == 'ADD')
            if not is_add and not is_valid_item(a_str):
                continue

            if is_add:
                children.append((row, a_str))
                continue

            a_dots = count_dots(a_str.replace('ADD', ''))

            # Stop at same-level or higher items (different branch)
            if a_dots <= m_dots:
                break

            # Check if this is a descendant
            if item_prefix_match(marker_a, a_str):
                children.append((row, a_str))

        return children

    def extract(self):
        self.load_source()
        markers = self.find_markers()
        if not markers:
            print(f"No items found matching keyword: {self.keyword}")
            return self

        seen_rows = set()

        for i, (m_row, m_a) in enumerate(markers):
            next_m_row = markers[i + 1][0] if i + 1 < len(markers) else None

            ancestors = self.find_ancestors(m_row, m_a)

            # Add section delimiter
            for a_type, a_row, a_a, a_desc in ancestors:
                if a_type == 'section' and a_row not in seen_rows:
                    seen_rows.add(a_row)
                    vals = [None] * 13
                    # Display name from section delimiter
                    vals[1] = a_a
                    self.output_rows.append({
                        'type': 'section', 'src_row': a_row,
                        'values': vals, 'item_a': a_a
                    })

            # Add class and subsection headers
            for a_type, a_row, a_a, a_desc in ancestors:
                if a_type in ('class_header', 'subsection') and a_row not in seen_rows:
                    seen_rows.add(a_row)
                    vals = [None] * 13
                    vals[0] = a_a
                    vals[1] = a_desc
                    rtype = 'class_header' if a_type == 'class_header' else 'subsection'
                    self.output_rows.append({
                        'type': rtype, 'src_row': a_row,
                        'values': vals, 'item_a': a_a
                    })

            # Add marker row itself
            if m_row not in seen_rows:
                seen_rows.add(m_row)
                vals = self._map_row(m_row)
                b_val = vals[1]
                m_dots = count_dots(m_a.replace('ADD', ''))
                if b_val and is_angle_header(b_val) and m_dots <= 1:
                    m_type = 'subsection'
                elif b_val and is_curly_header(b_val):
                    m_type = 'subsection'
                else:
                    m_type = 'item'
                self.output_rows.append({
                    'type': m_type, 'src_row': m_row,
                    'values': vals, 'item_a': m_a
                })

            # Add child items (descendants)
            children = self.find_children(m_row, m_a, next_m_row)
            for c_row, c_a in children:
                if c_row not in seen_rows:
                    seen_rows.add(c_row)
                    vals = self._map_row(c_row)
                    self.output_rows.append({
                        'type': 'item', 'src_row': c_row,
                        'values': vals, 'item_a': c_a
                    })

        print(f"Extracted {len(self.output_rows)} rows for keyword '{self.keyword}'")
        return self

    def _map_row(self, src_row):
        vals = [None] * 13
        for s_col, t_col in COL_MAP.items():
            vals[t_col] = self.ws_s.cell(row=src_row, column=s_col + 1).value
        return vals

    def save(self, output_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = re.sub(r'[^a-zA-Z0-9_]', '_', self.keyword)[:31]

        # Styles
        h_font = Font(name='Microsoft YaHei UI', size=10, bold=True, color='FFFFFFFF')
        h_fill = PatternFill(start_color=Color(theme=4, tint=-0.249977111117893),
                             end_color=Color(indexed=64), fill_type='solid')
        h_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        sec_font = Font(name='Microsoft YaHei UI', size=11, bold=True, color='FF1A1A1A')
        sec_fill = PatternFill(start_color='FFC6D9F1', end_color='FFC6D9F1', fill_type='solid')
        sec_align = Alignment(vertical='center')

        cls_font = Font(name='Microsoft YaHei UI', size=10, bold=True, color='FF1A1A1A')
        cls_fill = PatternFill(start_color='FFEEF2FA', end_color='FFEEF2FA', fill_type='solid')
        cls_align = Alignment(vertical='center')

        sub3_font = Font(name='Microsoft YaHei UI', size=10, bold=True, color='FF1A1A1A')
        sub3_fill = PatternFill(start_color='FFFBE5D6', end_color='FFFBE5D6', fill_type='solid')
        sub3_align = Alignment(vertical='center')

        item_font = Font(name='Microsoft YaHei UI', size=9, bold=False, color='FF1A1A1A')
        item_align = Alignment(vertical='center')

        thin = Side(style='thin', color='FFBFBFBF')
        border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
        border_bottom = Border(bottom=thin)
        border_bl = Border(bottom=thin, left=thin)
        border_br = Border(bottom=thin, right=thin)

        widths = {'A': 11.09, 'B': 60.73, 'C': 8.73, 'D': 14.73,
                  'E': 12.73, 'F': 12.73, 'G': 12.73, 'H': 12.73,
                  'I': 12.73, 'J': 12.73, 'K': 12.73, 'L': 12.73, 'M': 14.73}
        for col_letter, w in widths.items():
            ws.column_dimensions[col_letter].width = w

        # Header row 1
        h1 = ['Item', 'Item Description', 'Unit', 'Quantity',
              'Unit Rate\n(in USD excl. Taxes)',
              'Unit Rate Breakdown (in USD)',
              None, None, None, None, None, None,
              'Total Price\n(in USD excl. Taxes)']
        for i, v in enumerate(h1):
            c = ws.cell(row=1, column=i + 1)
            c.value = v
            c.font = h_font
            c.fill = h_fill
            c.alignment = h_align
            c.border = border_all

        # Header row 2
        h2 = [None, None, None, None, None,
              'Labour', 'Plant', 'Material', 'Subcontractor',
              'Others / Consultants',
              'Off-Site Overheads\n(Local & Head Office)',
              'Local & Head Office Profit', None]
        for i, v in enumerate(h2):
            c = ws.cell(row=2, column=i + 1)
            c.value = v
            if v:
                c.font = h_font
                c.fill = h_fill
                c.alignment = h_align
            c.border = border_all

        ws.merge_cells('A1:A2')
        ws.merge_cells('B1:B2')
        ws.merge_cells('C1:C2')
        ws.merge_cells('D1:D2')
        ws.merge_cells('E1:E2')
        ws.merge_cells('F1:L1')
        ws.merge_cells('M1:M2')

        for r in [1, 2]:
            for c_idx in range(1, 14):
                ws.cell(row=r, column=c_idx).border = border_all

        ws.row_dimensions[2].height = 58

        # Freeze header rows (same as template)
        ws.freeze_panes = 'A3'

        # Enable row grouping (collapse +/- buttons)
        ws.sheet_properties.outlinePr.summaryBelow = True

        # Pre-compute which entries have child items
        def _is_descendant(parent_code, child_code):
            if not parent_code or not child_code:
                return False
            pc = parent_code.strip().replace('ADD', '')
            cc = child_code.strip().replace('ADD', '')
            if cc == '' or pc == '':
                return False
            return cc.startswith(pc + '.') and count_dots(cc) > count_dots(pc)

        has_children = [False] * len(self.output_rows)
        for i, entry in enumerate(self.output_rows):
            if entry['type'] in ('section', 'class_header'):
                has_children[i] = True
                continue
            code = entry.get('item_a', '')
            if not code:
                continue
            # Check subsequent entries for descendants
            for j in range(i + 1, len(self.output_rows)):
                ncode = self.output_rows[j].get('item_a', '')
                if _is_descendant(code, ncode):
                    has_children[i] = True
                    break
                # If next entry has same or fewer dots, stop looking
                cdots = count_dots(code.strip().replace('ADD', ''))
                ndots = count_dots(ncode.strip().replace('ADD', '')) if ncode else 0
                if ndots <= cdots:
                    break

        # Data rows
        out_row = 3
        for i, entry in enumerate(self.output_rows):
            rtype = entry['type']
            vals = entry['values']

            for col_idx, val in enumerate(vals):
                cell = ws.cell(row=out_row, column=col_idx + 1)
                cl = get_column_letter(col_idx + 1)

                # Determine if this row is a {} L3 subsection
                is_l3 = vals[1] and is_curly_header(vals[1])
                is_parent_item = (rtype == 'item' and has_children[i]
                                  and not is_l3
                                  and not (vals[1] and is_angle_header(vals[1])))

                # Wrap promoted parent item description with {}
                display_val = val
                if is_parent_item and col_idx == 1 and val is not None:
                    s = str(val).strip()
                    if not s.startswith('{'):
                        display_val = '{' + s + '}'

                cell.value = display_val

                if rtype == 'section':
                    cell.font = sec_font
                    cell.fill = sec_fill
                    cell.alignment = sec_align
                elif is_l3 or is_parent_item:
                    cell.font = sub3_font
                    cell.fill = sub3_fill
                    cell.alignment = sub3_align
                elif rtype in ('class_header', 'subsection'):
                    cell.font = cls_font
                    cell.fill = cls_fill
                    cell.alignment = cls_align
                else:
                    cell.font = item_font
                    cell.alignment = item_align
                cell.border = border_bl if col_idx == 0 else (border_br if col_idx == 12 else border_bottom)

                if cl in NUM_COLS and val is not None:
                    cell.number_format = NUM_FMT

            if rtype == 'section':
                ws.row_dimensions[out_row].height = 16.5
            elif rtype in ('class_header', 'subsection', 'item'):
                ws.row_dimensions[out_row].height = 14.5

            # Set row outline level for grouping
            # 【】L1=0 → 《》L2=1 → {} L3=2 → parent item L3=2 → leaf L4=3 → ADD L5=3
            ol = TYPE_LEVEL.get(rtype, 4) - 1
            b_val = vals[1]
            if rtype in ('subsection', 'item') and entry.get('item_a'):
                code = entry['item_a'].strip()
                cdots = count_dots(code.replace('ADD', ''))
                if code == 'ADD':
                    ol = 3
                elif b_val and is_angle_header(b_val) and cdots <= 1:
                    ol = 1  # 《》 L2 subsection
                elif b_val and is_curly_header(b_val):
                    ol = 2  # {} L3 subsection
                elif has_children[i]:
                    ol = 2  # parent item with children → L3
                else:
                    ol = 2 if cdots <= 1 else 3
            ws.row_dimensions[out_row].outline_level = ol

            out_row += 1

        # Bottom border on last data row
        last_r = out_row - 1
        for c_idx in range(1, 14):
            cell = ws.cell(row=last_r, column=c_idx)
            existing = cell.border
            cell.border = Border(left=existing.left, right=existing.right,
                                 top=existing.top, bottom=thin)

        # Outer border around entire table
        outer_side = Side(style='thin', color='FFBFBFBF')
        for r in range(1, last_r + 1):
            # Left edge
            cell_a = ws.cell(row=r, column=1)
            cell_a.border = Border(left=outer_side, right=cell_a.border.right,
                                   top=cell_a.border.top, bottom=cell_a.border.bottom)
            # Right edge
            cell_m = ws.cell(row=r, column=13)
            cell_m.border = Border(left=cell_m.border.left, right=outer_side,
                                   top=cell_m.border.top, bottom=cell_m.border.bottom)
        for c_idx in range(1, 14):
            # Top edge
            cell_top = ws.cell(row=1, column=c_idx)
            cell_top.border = Border(left=cell_top.border.left, right=cell_top.border.right,
                                     top=outer_side, bottom=cell_top.border.bottom)
            # Bottom edge
            cell_bot = ws.cell(row=last_r, column=c_idx)
            cell_bot.border = Border(left=cell_bot.border.left, right=cell_bot.border.right,
                                     top=cell_bot.border.top, bottom=outer_side)

        wb.save(output_path)
        wb.close()
        print(f"Saved: {output_path} ({out_row - 1} rows)")
        return self


def main():
    parser = argparse.ArgumentParser(description='Extract BOQ items by keyword with template styling')
    parser.add_argument('source', help='Path to merged BOQ xlsx file')
    parser.add_argument('keyword', help='Keyword to search in item descriptions (case-insensitive)')
    parser.add_argument('template', help='Path to template xlsx')
    parser.add_argument('-o', '--output', help='Output xlsx path (default: {date}_BOQ_{keyword}.xlsx)')
    args = parser.parse_args()

    if not args.output:
        date_str = datetime.now().strftime('%Y-%m-%d')
        safe_keyword = re.sub(r'[^\w]', '_', args.keyword)[:30]
        output_dir = os.path.dirname(args.source)
        args.output = os.path.join(output_dir, f'{date_str}_BOQ_{safe_keyword}.xlsx')

    extractor = BOQExtractor(args.source, args.keyword, args.template)
    extractor.extract().save(args.output)


if __name__ == '__main__':
    main()
