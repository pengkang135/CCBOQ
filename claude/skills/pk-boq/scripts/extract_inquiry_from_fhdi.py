"""从FHDI BOQ清单按原层级提炼材料询价表。

用法:
  python extract_inquiry_from_fhdi.py <FHDI清单.xlsx> [-o output.xlsx]
"""
import argparse, os, re
from collections import OrderedDict
import openpyxl
import xlsxwriter


EXCLUDE_L1 = ['施工单位', '材料关税', 'TOTAL']


def parse_boq_hierarchy(source_path):
    """解析FHDI BOQ，按原始层级提取叶节点条目。"""
    wb = openpyxl.load_workbook(source_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    current_l1 = None
    current_l2 = None
    current_l3 = None
    hierarchy = OrderedDict()
    l1_order = []
    seq = 0

    for r in range(3, ws.max_row + 1):
        name = ws.cell(r, 3).value
        if name is None:
            continue
        name = str(name).strip()
        no = str(ws.cell(r, 2).value or '').strip()
        desc = str(ws.cell(r, 4).value or '').strip()
        unit = str(ws.cell(r, 5).value or '').strip()
        qty_raw = ws.cell(r, 6).value
        qty = float(qty_raw) if qty_raw is not None else None

        # L1 header
        if name.startswith('【'):
            if any(kw in name for kw in EXCLUDE_L1):
                current_l1 = None
                current_l2 = None
                current_l3 = None
                continue
            current_l1 = name
            current_l2 = None
            current_l3 = None
            if current_l1 not in hierarchy:
                hierarchy[current_l1] = OrderedDict()
                l1_order.append(current_l1)
            continue

        if current_l1 is None:
            continue

        # L2 header
        if name.startswith('《'):
            current_l2 = name
            current_l3 = None
            if current_l2 not in hierarchy[current_l1]:
                hierarchy[current_l1][current_l2] = OrderedDict()
            continue

        if current_l2 is None:
            continue

        # L3 header
        if name.startswith('{'):
            current_l3 = name
            if current_l3 not in hierarchy[current_l1][current_l2]:
                hierarchy[current_l1][current_l2][current_l3] = []
            continue

        # Leaf item candidate
        if not no or '.' not in no:
            continue

        # Filter: skip LS items and zero-qty
        if unit.upper() == 'LS':
            continue
        if qty is None or qty <= 0:
            continue

        seq += 1
        item = {
            'id': f'M{seq:03d}',
            'no': no,
            'name': name,
            'desc': desc if desc and desc != '0' else '',
            'unit': unit if unit and unit != '0' else '',
            'qty': qty,
        }

        if current_l3:
            hierarchy[current_l1][current_l2].setdefault(current_l3, []).append(item)
        else:
            hierarchy[current_l1][current_l2].setdefault('_no_l3', []).append(item)

    wb.close()

    # Detect L3 sections (any L2 that has items in _no_l3 gets auto-L3 if > threshold)
    # But since we already have L3 markers, just keep them
    # Rebuild to remove empty sections and handle _no_l3
    clean = OrderedDict()
    for l1 in l1_order:
        if l1 not in hierarchy:
            continue
        clean[l1] = OrderedDict()
        l2_entries = hierarchy[l1]
        for l2_name, l3_groups in l2_entries.items():
            # Collect all items under this L2
            all_items = []
            has_real_l3 = False
            for l3_name, items in l3_groups.items():
                if l3_name == '_no_l3':
                    all_items.extend(items)
                else:
                    has_real_l3 = True
                    all_items.append(('l3_header', l3_name))
                    all_items.extend(items)
            if all_items:
                clean[l1][l2_name] = (has_real_l3, all_items)

    print(f'  L1 sections: {len(clean)}')
    l2_count = len(clean)
    # Count actual L2
    actual_l2 = 0
    actual_items = 0
    for l1, l2s in clean.items():
        actual_l2 += len(l2s)
        for _, (has_l3, items) in l2s.items():
            actual_items += sum(1 for x in items if isinstance(x, dict))
    print(f'  L2 sections: {actual_l2}')
    print(f'  Leaf items: {actual_items}')
    return clean


def write_xlsx(hierarchy, output_path, title=None):
    """输出BOQ格式询价表（与pk-boq-price-build相同样式）。"""
    wb = xlsxwriter.Workbook(output_path, {'strings_to_urls': False})
    ws = wb.add_worksheet('材料询价表')

    fmt_l1 = wb.add_format({
        'bold': True, 'font_size': 11, 'font_name': 'Microsoft YaHei UI',
        'font_color': '#1A1A1A', 'bg_color': '#C6D9F1',
        'valign': 'vcenter', 'border': 0,
    })
    fmt_l2 = wb.add_format({
        'bold': True, 'font_size': 10, 'font_name': 'Microsoft YaHei UI',
        'font_color': '#1A1A1A', 'bg_color': '#EEF2FA',
        'valign': 'vcenter', 'border': 0,
    })
    fmt_l3 = wb.add_format({
        'bold': True, 'font_size': 10, 'font_name': 'Microsoft YaHei UI',
        'font_color': '#1A1A1A', 'bg_color': '#FBE5D6',
        'valign': 'vcenter', 'border': 0,
    })
    fmt_header = wb.add_format({
        'bold': True, 'font_size': 10, 'font_name': 'Microsoft YaHei UI',
        'bg_color': '#4472C4', 'font_color': '#FFFFFF',
        'border': 1, 'text_wrap': True, 'valign': 'vcenter', 'align': 'center',
    })
    fmt_data = wb.add_format({
        'font_size': 9, 'font_name': 'Microsoft YaHei UI',
        'border': 1, 'valign': 'vcenter', 'border_color': '#BFBFBF',
    })
    fmt_num = wb.add_format({
        'font_size': 9, 'font_name': 'Microsoft YaHei UI',
        'border': 1, 'valign': 'vcenter', 'border_color': '#BFBFBF',
        'num_format': '#,##0.0',
    })
    fmt_center = wb.add_format({
        'font_size': 9, 'font_name': 'Microsoft YaHei UI',
        'border': 1, 'valign': 'vcenter', 'border_color': '#BFBFBF',
        'align': 'center',
    })

    headers = ['编号', '专业', '名称', '项目特征', '单位', '数量',
               '除税单价', '税金', '含税单价', '日期', '币种',
               '供应商', '联系人', '电话', '地址', '备注']
    n_cols = len(headers)
    col_widths = [7, 14, 32, 22, 6, 9, 9, 6, 9, 10, 5, 10, 6, 10, 10, 18]

    for c, w in enumerate(col_widths):
        ws.set_column(c, c, w)

    ws.set_row(0, 22)
    for c, h in enumerate(headers):
        ws.write(0, c, h, fmt_header)
    ws.freeze_panes(1, 0)

    # Extract L1 title text (remove 【】 markers)
    def strip_markers(text):
        text = text.strip()
        if text.startswith('【') and text.endswith('】'):
            text = text[1:-1]
        return text

    row = 1
    for l1_name, l2_sections in hierarchy.items():
        # L1 header
        ws.write(row, 0, 'P', fmt_l1)
        ws.write(row, 1, '', fmt_l1)
        ws.write(row, 2, l1_name, fmt_l1)
        for c in range(3, n_cols):
            ws.write(row, c, '', fmt_l1)
        ws.set_row(row, 16.5, None, {'level': 0})
        row += 1

        for l2_name, (has_l3, items) in l2_sections.items():
            # L2 header
            ws.write(row, 0, '', fmt_l2)
            ws.write(row, 1, '', fmt_l2)
            l2_display = l2_name if l2_name.startswith('《') else f'《{l2_name}》'
            ws.write(row, 2, l2_display, fmt_l2)
            for c in range(3, n_cols):
                ws.write(row, c, '', fmt_l2)
            ws.set_row(row, 14.5, None, {'level': 1})
            row += 1

            current_had_l3 = False
            for entry in items:
                if isinstance(entry, tuple) and entry[0] == 'l3_header':
                    # L3 header
                    l3_text = entry[1]
                    l3_display = l3_text if l3_text.startswith('{') else f'{{{l3_text}}}'
                    ws.write(row, 0, '', fmt_l3)
                    ws.write(row, 1, '', fmt_l3)
                    ws.write(row, 2, l3_display, fmt_l3)
                    for c in range(3, n_cols):
                        ws.write(row, c, '', fmt_l3)
                    ws.set_row(row, 14.5, None, {'level': 2})
                    row += 1
                    current_had_l3 = True
                elif isinstance(entry, dict):
                    item = entry
                    ws.write(row, 0, item['id'], fmt_data)
                    ws.write(row, 1, '', fmt_data)
                    ws.write(row, 2, item['name'], fmt_data)
                    ws.write(row, 3, item['desc'], fmt_data)
                    ws.write(row, 4, item['unit'], fmt_center)
                    ws.write_number(row, 5, item['qty'], fmt_num)
                    ws.write(row, 6, '', fmt_num)
                    ws.write(row, 7, '', fmt_num)
                    ws.write(row, 8, '', fmt_num)
                    ws.write(row, 9, '', fmt_center)
                    ws.write(row, 10, 'USD', fmt_center)
                    for c in range(11, n_cols):
                        ws.write(row, c, '', fmt_data)
                    data_level = 3 if current_had_l3 else 2
                    ws.set_row(row, 14.5, None, {'level': data_level})
                    row += 1

    wb.close()
    print(f'  Output: {output_path}')
    print(f'  Total rows (incl headers): {row}')


def main():
    parser = argparse.ArgumentParser(description='从FHDI BOQ清单按原层级提炼材料询价表')
    parser.add_argument('source', help='FHDI清单 xlsx 文件路径')
    parser.add_argument('-o', '--output', default=None, help='输出文件路径')
    parser.add_argument('-t', '--title', default=None, help='标题文字（未使用，保留接口）')
    args = parser.parse_args()

    if args.output is None:
        base = os.path.splitext(os.path.basename(args.source))[0]
        dirname = os.path.dirname(args.source) or '.'
        args.output = os.path.join(dirname, f'{base}_材料询价表.xlsx')

    print(f'Source: {args.source}')
    print('Parsing BOQ hierarchy...')
    hierarchy = parse_boq_hierarchy(args.source)

    print('Writing xlsx...')
    write_xlsx(hierarchy, args.output, args.title)
    print('Done.')


if __name__ == '__main__':
    main()
