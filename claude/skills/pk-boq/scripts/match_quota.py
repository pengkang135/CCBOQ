# -*- coding: utf-8 -*-
"""
概算定额匹配工具：将合并 BOQ 清单逐项匹配到概算定额库。

匹配规则：
1. 单位完全一致（首要条件）
2. 名称/描述关键词匹配（中英双语桥接）
3. 单位匹配前提下，名称匹配可适度放宽

Usage:
    python match_quota.py <boq_file> <quota_file> [-o output.xlsx]
"""
import re
import sys
from pathlib import Path
from difflib import SequenceMatcher

import openpyxl
import xlsxwriter

# ── 中英双语工程关键词映射 ──────────────────────────────
# 每个概念一组，中文词和英文词放在同一组即为等价

BILINGUAL_CONCEPTS = [
    # 准备工作 / Preparatory
    {"场地清理", "ground clearance", "clearance", "clearing", "site clearance", "site clearing"},
    {"拆除", "demolition", "demolish", "dismantl"},
    {"围墙", "fence", "wall", "boundary wall", "compound wall"},
    {"路面", "pavement", "road", "paving"},
    {"建筑物", "building", "structure", "built structure"},
    {"灯杆", "light tower", "light pole", "lighting mast"},
    {"海堤", "seawall", "sea wall", "sea defence"},
    {"护岸", "revetment", "coastal protection"},
    {"服务沟渠", "services trench", "service trench", "trench"},
    {"混凝土围栏", "concrete fence", "concrete barrier"},
    {"渡轮码头", "ferry terminal", "ferry ghat", "ferry"},

    # 测量勘察 / Surveys
    {"水深测量", "bathymetric survey", "bathymetry", "hydrographic survey"},
    {"地形测量", "topographic survey", "topography"},
    {"地质勘察", "geotechnical investigation", "soil investigation", "ground investigation", "borehole"},
    {"海图测绘", "nautical chart", "chart survey", "hydrographic chart"},
    {"水下障碍物", "underwater obstruction", "uxo", "unexploded ordnance", "obstruction"},
    {"环保防污帘", "silt curtain", "turbidity curtain", "environmental curtain"},

    # 疏浚 / Dredging
    {"疏浚", "dredging", "dredge", "dredged"},
    {"泊位", "berth", "berthing", "berthing pocket"},
    {"水域", "basin", "water area", "water"},
    {"调头区", "turning basin", "turning circle", "turning area"},
    {"航道", "channel", "navigation channel", "access channel"},
    {"进退场", "mobilisation", "mobilization", "demobilisation", "demobilization"},
    {"疏浚设备", "dredging equipment", "dredger", "dredging plant"},

    # 回填 / Fill & Reclamation
    {"回填", "backfill", "filling", "reclamation"},
    {"吹填", "hydraulic fill", "hydraulic backfill", "reclamation fill"},
    {"受控填料", "controlled fill", "engineered fill"},
    {"堆载", "surcharge", "preloading", "preload"},
    {"堆载料", "surcharge material", "surcharge fill"},
    {"排水垫层", "drainage blanket", "drainage layer", "draining blanket"},
    {"外购回填", "imported fill", "imported backfill", "purchased fill"},
    {"沉降补偿", "settlement compensation", "subsidence"},
    {"护岸后回填砂", "sand backfill", "backfill sand"},

    # 土工材料 / Geosynthetics
    {"土工布", "geotextile", "geofabric", "filter fabric"},
    {"土工格栅", "geogrid", "geosynthetic"},
    {"无纺土工布", "non-woven geotextile", "nonwoven geotextile"},
    {"软体排", "mattress", "geotextile mattress"},

    # 地基处理 / Ground Improvement
    {"塑料排水板", "pvd", "prefabricated vertical drain", "wick drain", "band drain"},
    {"强夯", "dynamic compaction", "heavy tamping", "drop weight"},
    {"整平碾压", "compaction", "rolling compaction", "grading"},
    {"DCM", "dcm", "deep cement mixing", "deep soil mixing", "cdm"},
    {"帷幕墙", "cutoff wall", "barrier wall", "containment"},
    {"砂桩", "sand compaction pile", "scp", "sand pile", "sand drain"},
    {"深层水泥搅拌桩", "deep cement mixing", "dcm", "cdm"},

    # 桩基 / Piling
    {"PHC管桩", "phc pile", "prestressed concrete pile", "spun pile", "precast concrete pile"},
    {"钢管桩", "steel pile", "steel tube pile", "steel pipe pile", "tubular steel pile"},
    {"桩芯混凝土", "pile core concrete", "pile plug", "concrete plug", "pile infill"},
    {"板桩", "混凝土板桩", "concrete sheet pile", "sheet pile", "precast sheet pile"},
    {"钻孔灌注桩", "bored pile", "bored cast-in-situ pile", "drilled pile"},

    # 上部结构 / Superstructure
    {"桩帽", "pile cap", "pilecap"},
    {"横梁", "beam", "cross beam", "transverse beam"},
    {"面板", "slab", "deck", "deck slab"},
    {"面层", "topping", "surface", "wearing course", "wearing surface"},
    {"靠船构件", "berthing member", "fender support", "dolphin"},
    {"梁系", "beam system", "beam grid", "longitudinal beam"},
    {"钢筋", "reinforcement", "rebar", "steel bar", "reinforcing steel"},
    {"模板", "formwork", "shuttering", "mould"},
    {"混凝土", "concrete"},

    # 附属设施 / Ancillary & Fittings
    {"护舷", "fender", "rubber fender", "fender system"},
    {"系船柱", "bollard", "mooring bollard", "mooring post"},
    {"岸桥轨道", "crane rail", "gantry rail", "quay crane rail", "crane track"},
    {"电缆沟", "cable trench", "cable duct", "cable tray"},
    {"岸桥缓冲器", "crane buffer", "crane stop", "end stop"},
    {"风暴系固", "storm mooring", "storm anchor", "typhoon mooring"},
    {"救生圈", "life buoy", "life ring", "life saving"},
    {"护轮坎", "wheel guard", "kerb", "curb"},
    {"护栏", "guardrail", "railing", "handrail", "safety barrier"},
    {"伸缩缝", "expansion joint", "movement joint"},
    {"标线", "marking", "line marking", "pavement marking"},
    {"岸电坑", "shore power pit", "shore power socket", "shore connection"},
    {"RTG风暴系固", "rtg storm", "rtg mooring", "rtg tie down"},

    # 防冲刷 / Scour Protection
    {"防冲刷", "scour protection", "scour prevention", "erosion protection"},
    {"袋装砂", "sand bag", "sand filled bag", "geobag", "geotextile bag"},
    {"护面块石", "armour rock", "armor stone", "riprap", "rock armour"},
    {"碎石垫层", "gravel bedding", "crushed stone", "granular bedding"},
    {"滤层", "filter layer", "filter", "granular filter"},
    {"倒滤层", "inverted filter", "reverse filter"},

    # 陆域 / Yard
    {"堆场", "container yard", "stacking yard", "storage yard", "yard"},
    {"道路", "road", "pavement", "access road", "roadway"},
    {"排水", "drainage", "storm water", "stormwater"},
    {"给水", "water supply", "potable water"},
    {"消防", "fire fighting", "fire hydrant", "fire protection"},
    {"污水", "sewerage", "sewage", "wastewater"},
    {"围网", "fencing", "security fence", "perimeter fence"},
    {"门", "gate", "door"},
    {"照明", "lighting", "illumination", "light"},
    {"铺面", "paving", "pavement", "block paving", "interlock"},
    {"路缘石", "kerb", "curb", "edge"},
    {"标志牌", "sign", "signage", "sign board"},

    # 建筑 / Buildings
    {"钢结构", "steel structure", "structural steel", "steel frame"},
    {"屋面", "roof", "roofing", "roof cover"},
    {"墙体", "wall", "masonry", "blockwork"},
    {"门窗", "door", "window", "doors and windows"},
    {"装修", "finishing", "fit out", "architectural finish"},
    {"电气", "electrical", "electric", "power"},
    {"给排水", "plumbing", "sanitary", "water supply and drainage"},
    {"暖通", "hvac", "ventilation", "air conditioning"},
    {"消防报警", "fire alarm", "fire detection", "fire safety"},
    {"电梯", "lift", "elevator"},

    # 机电 / MEP
    {"变压器", "transformer", "electrical transformer"},
    {"发电机", "generator", "diesel generator", "standby generator"},
    {"配电柜", "switchboard", "distribution board", "panel board"},
    {"电缆", "cable", "power cable", "electrical cable"},
    {"管道", "pipe", "pipeline", "piping", "duct"},
    {"水泵", "pump", "water pump"},
    {"阀门", "valve"},
    {"灭火器", "fire extinguisher", "extinguisher"},

    # 场地形成 / Site Formation
    {"回填砂", "sand fill", "sand backfill"},
    {"场地整平", "site grading", "leveling", "site formation"},
    {"碾压", "compaction", "rolling", "compacted"},

    # 其他通用
    {"监测", "monitoring", "instrumentation", "observation"},
    {"检测", "testing", "inspection", "test"},
    {"临时工程", "temporary works", "temporary", "temp"},
    {"交通疏导", "traffic management", "traffic control"},
    {"安全防护", "safety protection", "safety barrier", "safety"},

    # 桩基检测
    {"静载试验", "static load test", "pile load test", "static test"},
    {"动测", "dynamic test", "pda test", "pile dynamic test"},
    {"PIT检测", "pit test", "pile integrity test", "integrity test"},
    {"声测管", "sonic logging", "cross hole", "sonic tube"},

    # CDM/DCM (补充)
    {"CDM桩", "cdm pile", "cdm column", "deep mixing column"},
    {"CDM非注浆区", "cdm non-grouted", "non-grouted zone"},
    {"CDM地基处理", "cdm ground improvement", "cdm treatment"},

    # 水下障碍物/管线
    {"海底电缆", "submarine cable", "subsea cable", "marine cable"},
    {"管线", "pipeline", "pipelines", "services"},
    {"障碍物调查", "obstruction survey", "obstruction investigation"},
    {"试挖", "trial pit", "trial trench", "test pit"},

    # 测试检测 (补充)
    {"测试检测", "testing and inspection", "testing", "ndt", "non-destructive"},
    {"填方检测", "testing for fill", "fill testing", "earthwork testing"},
    {"护岸检测", "testing for revetment", "revetment testing"},
    {"地基处理检测", "ground treatment testing", "ground improvement testing"},

    # 临时工程
    {"临时支撑", "temporary support", "temporary works", "temporary prop"},
    {"临时围堰", "temporary cofferdam", "cofferdam"},
    {"临时排水", "temporary drainage", "dewatering"},

    # 模板/临时
    {"模板", "formwork", "shuttering", "formwork system"},
    {"脚手架", "scaffolding", "scaffold", "access platform"},

    # 钢结构/铁件
    {"钢结构", "steel structure", "structural steel", "steelwork", "steel frame"},
    {"预埋件", "embedded part", "embed", "cast-in fitting", "sole plate", "baseplate", "base plate"},
    {"螺栓", "bolt", "screw", "anchor bolt", "holding down bolt"},
    {"镀锌", "galvanised", "galvanized", "hot dip", "hdg"},

    # 围栏/安全
    {"刺绳", "barbed wire", "concertina coil", "razor wire"},
    {"围栏", "fencing", "fence", "perimeter fence", "security fence"},

    # 道路/闸口
    {"标志牌", "signage", "sign", "sign board", "traffic sign"},
    {"地磅", "weighbridge", "truck scale", "weigh scale"},
    {"道闸", "barrier gate", "boom gate", "boom barrier"},
    {"雨棚", "canopy", "shed", "shelter", "roof cover"},

    # 电气面板/插座
    {"插座", "socket", "outlet", "power socket", "plug socket"},
    {"电缆桥架", "cable tray", "cable ladder", "cable rack"},
    {"冷藏箱", "reefer", "refrigerated container", "reefer container"},

    # 雨棚/天棚
    {"雨棚钢结构", "canopy steel", "canopy structure"},
    {"雨棚电气", "canopy electrical", "canopy lighting"},

    # 码头面层
    {"现浇面板", "cast in situ slab", "in-situ slab", "insitu deck"},
    {"现浇混凝土", "cast in situ concrete", "in-situ concrete", "insitu concrete"},
    {"预制构件", "precast element", "precast unit", "prefabricated"},

    # 轨道
    {"岸桥轨道", "crane rail", "gantry rail", "crane beam"},
    {"紧固系统", "fixing system", "fastening", "rail clamp"},

    # 门机 / Gate Complex
    {"大门", "gate", "main gate", "entrance gate"},
    {"门卫", "guard house", "security booth", "gate house"},
    {"围墙", "boundary wall", "compound wall", "perimeter wall"},
    {"地磅", "weighbridge", "truck scale", "weigh bridge"},
]

# 构建概念查找：每个词 → 概念组索引
_CONCEPT_INDEX = {}
for i, concept_set in enumerate(BILINGUAL_CONCEPTS):
    for term in concept_set:
        key = term.lower().strip()
        if key not in _CONCEPT_INDEX:
            _CONCEPT_INDEX[key] = []
        _CONCEPT_INDEX[key].append(i)

# 通用概念组（太泛，单独不足以锚定匹配）
# 如"混凝土"出现在主体结构、附属设施、桩基等各处；"排水"横跨岩土、管道、道路
_GENERIC_CONCEPT_GROUPS = {
    i for i, s in enumerate(BILINGUAL_CONCEPTS)
    if s in (
        {'混凝土', 'concrete'},
        {'排水', 'drainage', 'storm water', 'stormwater'},
    )
}


def _tokenize_cn(text):
    """中文分词：2-gram + 3-gram 字符 + 整词"""
    text = text.lower()
    text = re.sub(r'[（）()\s,，.。、/·:：;；\-+]+', ' ', text)
    tokens = set()
    # 整词（空格分隔）
    for w in text.split():
        w = w.strip()
        if len(w) >= 1:
            tokens.add(w)
    # 2-gram + 3-gram
    clean = text.replace(' ', '')
    for i in range(len(clean) - 1):
        tokens.add(clean[i:i+2])
    for i in range(len(clean) - 2):
        tokens.add(clean[i:i+3])
    return tokens


def _tokenize_en(text):
    """英文分词"""
    # 分类号前缀如 B.1.1 先移除
    text = re.sub(r'\b[A-I]\.\d+(?:\.\d+)*\b', '', text)
    text = text.lower()
    text = re.sub(r'[^\w\s/]', ' ', text)
    tokens = set()
    for w in text.split():
        w = w.strip()
        if len(w) >= 2:
            tokens.add(w)
    return tokens


def extract_concepts(text, is_chinese):
    """提取文本中涉及的概念组 ID 列表"""
    tokens = _tokenize_cn(text) if is_chinese else _tokenize_en(text)
    concepts = set()
    for token in tokens:
        # 精确匹配
        if token in _CONCEPT_INDEX:
            concepts.update(_CONCEPT_INDEX[token])

    # 部分匹配（仅英语）
    if not is_chinese:
        # 清除标点，与 _tokenize_en 一致
        cleaned = re.sub(r'[^\w\s/]', ' ', text.lower())
        text_words = set(cleaned.split())
        for key, idxs in _CONCEPT_INDEX.items():
            if len(key) <= 5:
                continue
            if ' ' in key:
                # 多词 key：连续出现 或 所有词分散在原文中均可匹配
                if key in cleaned:
                    concepts.update(idxs)
                else:
                    key_words = key.split()
                    if all(kw in text_words for kw in key_words):
                        concepts.update(idxs)
            else:
                # 单词 key：key 必须是 token 的子串
                for token in tokens:
                    if len(token) >= 5 and key in token:
                        concepts.update(idxs)
                        break
    return concepts


def name_similarity(quota_name, boq_desc, section_concepts=None):
    """计算概算定额名称与 BOQ 描述的概念重叠度 (0~1)。
    返回 (jaccard, intersection_count) 以便调用方判断是否有实质重叠。
    section_concepts: 分部/子分部标题提取的概念，作为 BOQ 上下文补充"""
    cn_concepts = extract_concepts(quota_name, is_chinese=True)
    en_concepts = extract_concepts(boq_desc, is_chinese=False)

    # 融入分部/子分部标题上下文：分部标题概念加入 BOQ 概念集
    # 使得与分部主题不符的定额匹配被稀释
    if section_concepts:
        en_concepts = en_concepts | section_concepts

    if not cn_concepts and not en_concepts:
        return 0.0, 0, 0
    if not cn_concepts or not en_concepts:
        return 0.0, 0, 0

    intersection = cn_concepts & en_concepts
    union = cn_concepts | en_concepts
    jaccard = len(intersection) / len(union) if union else 0.0
    # 区分度：非重叠概念数越多，匹配越不可靠
    non_overlap = len(union) - len(intersection)
    # 通用概念重叠惩罚：仅有通用概念（如「混凝土」「排水」）重叠不足以锚定匹配
    if intersection and intersection.issubset(_GENERIC_CONCEPT_GROUPS):
        jaccard *= 0.5
    return jaccard, len(intersection), non_overlap


def normalize_unit(u):
    """标准化单位"""
    if u is None:
        return None
    u = str(u).strip().lower()
    u = re.sub(r'\s+', '', u)
    # 去掉尾部句点
    u = u.rstrip('.')
    # 统一常见写法
    mapping = {
        'm3': 'm³', 'm2': 'm²', 'm³': 'm³', 'm²': 'm²',
        '㎡': 'm²',
        'no': 'no.', 'nos': 'no.',
        'nr': 'no.',
        'ton': 't', 'tons': 't', 'tonne': 't', 'tonnes': 't',
        'lumpsum': 'ls', 'ls': 'ls', 'l.s': 'ls',
        'item': 'item',
        'sqm': 'm²', 'sq.m': 'm²',
        'cum': 'm³', 'cu.m': 'm³',
        'meter': 'm', 'meters': 'm',
        'm.': 'm', 'm': 'm',
        'kg': 'kg',
        'set': '套', 'sets': '套',
        'piece': 'no.', 'pieces': 'no.',
        '个': 'no.',
        'hr': 'hr', 'hour': 'hr', 'hours': 'hr',
        'day': 'day', 'days': 'day',
        'month': 'month', 'months': 'month',
        'week': 'week', 'weeks': 'week',
        'lot': 'lot',
        'l.s.': 'ls', 'l.s': 'ls',
        'mm': 'mm', 'cm': 'cm', 'km': 'km',
        'ha': 'ha',
        '栋': '栋', '座': '座', '台': '台',
        '坑': '坑',
        '套': '套',
        'points': 'no.', 'point': 'no.',
    }
    return mapping.get(u, u)

# 单位桥接：BOQ单位 → 可匹配的定额单位列表
UNIT_BRIDGE = {
    'kg': ['t', 'kg'],       # 钢筋 kg ↔ t 可互通
    '套': ['套', 'no.', 'lot'],
    'no.': ['no.', '套', '座', '台', '栋', '坑'],
    'hr': ['hr'], 'day': ['day'], 'month': ['month'],
}


def read_quota_db(path):
    """读取概算定额库，返回 [(定额编号, 项目名称, 单位), ...]"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row) + [None] * 4
        code = str(vals[0]).strip() if vals[0] is not None else ""
        name = str(vals[2]).strip() if vals[2] is not None else ""
        unit = str(vals[3]).strip() if vals[3] is not None else ""
        # 跳过 Class 标题 (【), Section 标题(《), 方案标记({), 合价估算行
        if not code or code.startswith('【') or code.startswith('《') or code.startswith('{'):
            continue
        if not name:
            continue
        if '合价估算' in name:
            continue
        entries.append((code, name, normalize_unit(unit)))
    wb.close()
    return entries


def code_prefix_depth(boq_code, quota_code):
    """计算 BOQ 编号与定额编号的层级匹配深度。
    1 = 同 Class 字母, 2 = 同 Class + 同一级数字, 3+ = 更深匹配"""
    if not boq_code or not quota_code:
        return 0
    b_class = boq_code[0].upper()
    q_class = quota_code[0].upper()
    if b_class != q_class:
        return 0
    b_nums = re.findall(r'\d+', boq_code)
    q_nums = re.findall(r'\d+', quota_code)
    depth = 1  # class 匹配
    for i in range(min(len(b_nums), len(q_nums))):
        if int(b_nums[i]) == int(q_nums[i]):
            depth += 1
        else:
            break
    return depth


def match_boq_to_quota(boq_rows, quota_entries, min_score=0.30):
    """
    逐行匹配 BOQ → 概算定额。
    boq_rows: [(item_code, desc, unit, section_concepts), ...]
    返回: [(定额编号, 定额名称, match_score), ...]  未匹配为 (None, None, 0)
    """
    results = []
    for item_code, desc, unit, section_concepts in boq_rows:
        if not desc or not unit:
            results.append((None, None, 0.0))
            continue

        norm_unit = normalize_unit(unit)
        if not norm_unit:
            results.append((None, None, 0.0))
            continue

        # 可接受的单位列表（含桥接）
        acceptable_units = {norm_unit}
        if norm_unit in UNIT_BRIDGE:
            acceptable_units.update(UNIT_BRIDGE[norm_unit])

        # 过滤单位匹配的定额项（精确+桥接）
        unit_matches = [(qcode, qname, qunit) for qcode, qname, qunit in quota_entries
                       if qunit in acceptable_units]

        if not unit_matches:
            results.append((None, None, 0.0))
            continue

        # 计算每个候选的相似度
        scored = []
        for qcode, qname, qunit in unit_matches:
            sim, intersect_count, non_overlap = name_similarity(qname, desc, section_concepts)

            # 必须要有实质概念重叠，否则直接跳过
            if intersect_count == 0:
                continue

            # 非重叠概念惩罚：定额或清单独有的概念越多，匹配置信度越低
            sim -= 0.12 * non_overlap

            # 分部上下文不匹配惩罚：分部有明确主题但定额完全未涉及
            if section_concepts:
                cn_concepts = extract_concepts(qname, is_chinese=True)
                if not (section_concepts & cn_concepts):
                    sim -= 0.15

            # 层级匹配加分：按编号前缀深度递进
            prefix_depth = code_prefix_depth(item_code, qcode)
            if prefix_depth == 1:
                sim += 0.03
            elif prefix_depth == 2:
                sim += 0.06
            elif prefix_depth >= 3:
                sim += 0.09

            # 精确单位匹配微调加分
            if qunit == norm_unit:
                sim += 0.05

            scored.append((qcode, qname, sim))

        if not scored:
            results.append((None, None, 0.0))
            continue

        scored.sort(key=lambda x: x[2], reverse=True)
        best = scored[0]

        if best[2] >= min_score:
            results.append((best[0], best[1], best[2]))
        else:
            results.append((None, None, 0.0))

    return results


def main():
    import argparse
    parser = argparse.ArgumentParser(description="匹配概算定额到 BOQ 合并清单")
    parser.add_argument("boq", help="BOQ 合并清单 xlsx")
    parser.add_argument("quota", help="概算定额库 xlsx")
    parser.add_argument("-o", "--output", default=None, help="输出路径")
    parser.add_argument("--min-score", type=float, default=0.30,
                       help="最低匹配分数阈值 (default: 0.30)")
    args = parser.parse_args()

    boq_path = Path(args.boq)
    quota_path = Path(args.quota)

    if args.output:
        output_path = Path(args.output)
    else:
        output_path = boq_path.parent / f"{boq_path.stem}_with_quota.xlsx"

    print("=" * 60)
    print("概算定额匹配")
    print("=" * 60)

    # 1. 读取定额库
    print("\n[1/4] 读取概算定额库...")
    quota_entries = read_quota_db(quota_path)
    print(f"  有效定额条目: {len(quota_entries)}")

    # 2. 读取 BOQ
    print("\n[2/4] 读取 BOQ 合并清单...")
    wb = openpyxl.load_workbook(boq_path, data_only=True)
    ws = wb["MergedBOQ"]

    boq_data = []  # [(item, desc, unit, ...row_data), ...]
    boq_rows_for_match = []  # [(item, desc, unit, section_concepts), ...]

    # 跟踪分部/子分部标题上下文
    current_class = ""     # 【...】
    current_section = ""   # 《...》
    current_subsection = ""  # {...}

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        vals = list(row) + [None] * (10 - len(row))
        item = str(vals[0]).strip() if vals[0] is not None else ""
        desc = str(vals[1]).strip() if vals[1] is not None else ""
        unit = str(vals[2]).strip() if vals[2] is not None else ""

        boq_data.append({
            'row': row_idx,
            'item': item,
            'desc': desc,
            'unit': unit,
            'rest': list(vals[3:10])  # D-J columns
        })

        # 需要匹配的行：有单位且非标题行（非【】、《》、{}开头的描述）
        is_header = desc.startswith('【') or desc.startswith('《') or desc.startswith('{')
        if row_idx == 1:
            is_header = True  # 表头行

        # 更新分部/子分部标题上下文
        if is_header and row_idx > 1:
            if desc.startswith('【'):
                current_class = desc
                current_section = ""
                current_subsection = ""
            elif desc.startswith('《'):
                current_section = desc
                current_subsection = ""
            elif desc.startswith('{'):
                current_subsection = desc

        # 提取分部标题的概念（英文，因为标题多为中英混合）
        section_text = " ".join(filter(None, [current_class, current_section, current_subsection]))
        section_concepts = extract_concepts(section_text, is_chinese=False) if section_text else set()

        if is_header or not unit:
            boq_rows_for_match.append((item, desc, None, set()))  # 不参与匹配
        else:
            boq_rows_for_match.append((item, desc, unit, section_concepts))

    wb.close()
    print(f"  BOQ 总行数: {len(boq_data)}")

    # 3. 匹配
    print("\n[3/4] 执行匹配...")
    match_results = match_boq_to_quota(boq_rows_for_match, quota_entries, min_score=args.min_score)

    matched = sum(1 for r in match_results if r[0] is not None)
    total_matchable = sum(1 for r in boq_rows_for_match if r[2] is not None)
    print(f"  匹配成功: {matched}/{total_matchable} ({matched/total_matchable*100:.1f}%)")

    # 4. 写输出
    print("\n[4/4] 写入 Excel...")
    out_wb = xlsxwriter.Workbook(str(output_path))
    out_ws = out_wb.add_worksheet("MergedBOQ")

    # 样式
    header_fill = out_wb.add_format({
        'bold': True, 'font_color': '#FFFFFF', 'bg_color': '#1F4E79',
        'font_name': 'Microsoft YaHei UI', 'font_size': 9,
        'border': 1, 'border_color': '#D9D9D9',
        'text_wrap': True, 'valign': 'vcenter', 'align': 'center',
    })
    data_fmt = out_wb.add_format({
        'font_name': 'Microsoft YaHei UI', 'font_size': 9,
        'border': 1, 'border_color': '#D9D9D9',
        'valign': 'vcenter', 'font_color': '#1A1A1A',
    })
    num_fmt = out_wb.add_format({
        'font_name': 'Microsoft YaHei UI', 'font_size': 9,
        'border': 1, 'border_color': '#D9D9D9',
        'valign': 'vcenter', 'font_color': '#1A1A1A',
        'num_format': '#,##0.00',
    })
    title_fmt = out_wb.add_format({
        'font_name': 'Microsoft YaHei UI', 'font_size': 11,
        'bold': True, 'font_color': '#1A1A1A',
        'bg_color': '#C6D9F1',
        'border': 1, 'border_color': '#D9D9D9',
        'valign': 'vcenter',
    })
    section_fmt = out_wb.add_format({
        'font_name': 'Microsoft YaHei UI', 'font_size': 10,
        'bold': True, 'font_color': '#1A1A1A',
        'bg_color': '#EEF2FA',
        'border': 1, 'border_color': '#D9D9D9',
        'valign': 'vcenter',
    })
    sub_fmt = out_wb.add_format({
        'font_name': 'Microsoft YaHei UI', 'font_size': 9,
        'font_color': '#1A1A1A',
        'bg_color': '#FBE5D6',
        'border': 1, 'border_color': '#D9D9D9',
        'valign': 'vcenter',
    })
    low_score_fmt = out_wb.add_format({
        'font_name': 'Microsoft YaHei UI', 'font_size': 9,
        'font_color': '#1A1A1A',
        'bg_color': '#FFF2CC',
        'border': 1, 'border_color': '#D9D9D9',
        'valign': 'vcenter',
    })

    # 新列头
    new_headers = ["定额编号", "概算定额名称", "Item", "Item Description", "Unit",
                   "招标标准清单", "WTCC Design", "WTCC Bid",
                   "FHDI 方案一", "FHDI 方案二", "SGHCC", "Main Specification"]

    for c, h in enumerate(new_headers):
        out_ws.write(0, c, h, header_fill)
    out_ws.set_row(0, 28)

    # 列宽
    col_widths = [14, 30, 18, 55, 8, 14, 14, 14, 14, 14, 14, 40]
    for c, w in enumerate(col_widths):
        out_ws.set_column(c, c, w)

    # 写数据
    for i, (data, (qcode, qname, score)) in enumerate(zip(boq_data, match_results)):
        r = i + 1
        desc = data['desc']

        # 确定行样式
        if i == 0:  # 表头已在上面写了，跳过
            continue

        is_num_col = lambda c: c >= 5  # D列以后是数值列

        # 分级判断
        if desc.startswith('【'):
            row_fmt = title_fmt
            row_height = 24
        elif desc.startswith('《'):
            row_fmt = section_fmt
            row_height = 20
        elif desc.startswith('{'):
            row_fmt = sub_fmt
            row_height = 18
        else:
            if score > 0 and score < 0.25:
                row_fmt = low_score_fmt
            else:
                row_fmt = data_fmt
            row_height = 16

        out_ws.set_row(r, row_height)

        # 定额编号 (A)
        out_ws.write(r, 0, qcode if qcode else "", row_fmt)
        # 概算定额名称 (B)
        out_ws.write(r, 1, qname if qname else "", row_fmt)
        # Item (C)
        out_ws.write(r, 2, data['item'], row_fmt)
        # Description (D)
        out_ws.write(r, 3, desc, row_fmt)
        # Unit (E)
        out_ws.write(r, 4, data['unit'], row_fmt)

        # 数值列 (F-L, 原 D-J)
        for ci, val in enumerate(data['rest']):
            col = 5 + ci
            if val is None:
                continue
            if isinstance(val, (int, float)):
                out_ws.write_number(r, col, val, num_fmt)
            else:
                out_ws.write(r, col, val, row_fmt)

    # 冻结 + 筛选
    out_ws.freeze_panes(1, 0)
    out_ws.autofilter(0, 0, len(boq_data), len(new_headers) - 1)

    out_wb.close()
    print(f"\n输出: {output_path}")
    print(f"匹配率: {matched}/{total_matchable} ({matched/total_matchable*100:.1f}%)")


if __name__ == "__main__":
    main()
