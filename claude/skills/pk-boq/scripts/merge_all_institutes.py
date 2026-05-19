# -*- coding: utf-8 -*-
"""全量合并多家 BOQ 清单 → 统一 Excel

Usage:
    python merge_all_institutes.py --wtcc <path> --fhdi <path> --sghcc <path> --tender <path> [-o output.xlsx]

    也可作为模块导入:
    from merge_all_institutes import InstituteMerger
    merger = InstituteMerger()
    merger.merge(wtcc_path, fhdi_path, sghcc_path, tender_path, output_path)
"""
import argparse
import re
from datetime import datetime
from pathlib import Path
from collections import defaultdict, OrderedDict, Counter
from difflib import SequenceMatcher

import openpyxl
import xlsxwriter

# ── 正则 ──────────────────────────────────────────────
CLASS_RE = re.compile(r"^Class\s+([A-I])\b", re.IGNORECASE)
ITEM_CODE_RE = re.compile(r"^([A-I]\.\d+(?:\.\d+)*)\b")
OPTION_RE = re.compile(r"[（(]方案([一二三四五六七八九十]+)\s*[）)]")
ADD_RE = re.compile(r"\bADD\b", re.IGNORECASE)
TENDER_CLASS_RE = re.compile(r"^CLASS\s+([A-I])\b", re.IGNORECASE)

# ── 工具函数 ──────────────────────────────────────────
def safe_float(v):
    if v is None: return None
    if isinstance(v, (int, float)): return float(v) if v == v else None
    try: return float(str(v).strip().replace(",", "").replace(" ", ""))
    except (ValueError, TypeError): return None

def safe_qty(v):
    """保留数值或文本，不丢弃非数字信息（如 'Customized as required'）"""
    if v is None: return None
    if isinstance(v, (int, float)):
        return float(v) if v == v else None
    s = str(v).strip()
    if not s: return None
    try:
        f = float(s.replace(",", "").replace(" ", ""))
        return int(f) if f == int(f) and "." not in s.replace(",", "").replace(" ", "") else f
    except (ValueError, TypeError):
        return s  # 保留原始文本

def normalize_unit(u):
    if u is None: return None
    u = str(u).strip().lower()
    mapping = {"m3":"m³","m2":"m²","no.":"No.","nos":"No.","nr.":"No.",
               "ton":"t","item":"Item","ls":"LS","l.s.":"LS","lump sum":"LS"}
    return mapping.get(u, u)

def parse_item_code(code_str):
    if code_str is None or not isinstance(code_str, str): return None, False
    code = code_str.strip()
    is_add = bool(ADD_RE.search(code))
    clean = ADD_RE.sub("", code).strip().rstrip(".-").rstrip()
    return clean, is_add

def get_option_number(desc):
    if desc is None: return None
    m = OPTION_RE.search(str(desc))
    if not m: return None
    num_map = {"一":1,"二":2,"三":3,"四":4}
    return num_map.get(m.group(1), None)

def desc_similarity(a, b):
    if not a or not b: return 0.0
    def clean(s):
        s = s.lower()
        s = re.sub(r"[（(].*?[）)]", "", s)
        s = re.sub(r"\{.*?\}", "", s)
        s = re.sub(r"[^\w\s]", " ", s)
        s = re.sub(r"\s+", " ", s).strip()
        return s
    return SequenceMatcher(None, clean(a), clean(b)).ratio()

# ═══════════════════════════════════════════════════════
# 文件读取
# ═══════════════════════════════════════════════════════

def read_wtcc(path):
    """WTCC: A=Item, B=Desc, C=Unit, D=BidQty, E=Spec, F=DesignQty"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["MergeSheet"]
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        vals = list(row) + [None] * (6 - len(row))
        code = str(vals[0]).strip() if vals[0] is not None else ""
        desc = str(vals[1]).strip() if vals[1] is not None else ""
        unit = str(vals[2]).strip() if vals[2] is not None else ""
        bid_qty = safe_qty(vals[3])
        spec = str(vals[4]).strip() if vals[4] is not None else ""
        design_qty = safe_qty(vals[5])
        if not code and not desc and bid_qty is None and design_qty is None: continue
        if code.lower() == "item": continue
        rows.append({"row":row_idx, "code":code, "desc":desc, "unit":normalize_unit(unit),
                     "bid_qty":bid_qty, "design_qty":design_qty, "spec":spec, "source":"WTCC"})
    wb.close()
    return rows


def read_fhdi(path):
    """FHDI: A=Item, B=Desc, C=Unit, D=Qty"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["MergeSheet"]
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        vals = list(row) + [None] * (16 - len(row))
        code = str(vals[0]).strip() if vals[0] is not None else ""
        desc = str(vals[1]).strip() if vals[1] is not None else ""
        unit = str(vals[2]).strip() if vals[2] is not None else ""
        qty = safe_qty(vals[3])
        if not code and not desc and qty is None: continue
        if code.lower() == "item": continue
        rows.append({"row":row_idx, "code":code, "desc":desc, "unit":normalize_unit(unit),
                     "qty":qty, "source":"FHDI"})
    wb.close()
    return rows


def read_sghcc(path):
    """SGHCC: A=Item, B=Desc, C=Unit, D=Qty"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["MergeSheet"]
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        vals = list(row) + [None] * (19 - len(row))
        code = str(vals[0]).strip() if vals[0] is not None else ""
        desc = str(vals[1]).strip() if vals[1] is not None else ""
        unit = str(vals[2]).strip() if vals[2] is not None else ""
        qty = safe_qty(vals[3])
        if not code and not desc and qty is None: continue
        if code.lower() == "item": continue
        rows.append({"row":row_idx, "code":code, "desc":desc, "unit":normalize_unit(unit),
                     "qty":qty, "source":"SGHCC"})
    wb.close()
    return rows


def read_tender(path):
    """招标标准清单: 支持两种格式
    - BQMerge (MergeSheet): A=Item, B=Desc, C=Unit, D=Qty
    - 原始 Sheet1: A=Item, B=Desc_EN, C=Desc_ZH, D=Unit, E=Qty
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    if "MergeSheet" in wb.sheetnames:
        ws = wb["MergeSheet"]
        rows = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            vals = list(row) + [None] * (5 - len(row))
            code = str(vals[0]).strip() if vals[0] is not None else ""
            desc = str(vals[1]).strip() if vals[1] is not None else ""
            unit = str(vals[2]).strip() if vals[2] is not None else ""
            qty = safe_qty(vals[3])
            if not code and not desc and qty is None: continue
            if code.lower() == "item": continue
            rows.append({"row":row_idx, "code":code, "desc":desc, "desc_zh":"",
                         "unit":normalize_unit(unit), "qty":qty, "source":"TENDER"})
    else:
        ws = wb["Sheet1"]
        rows = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            vals = list(row) + [None] * (5 - len(row))
            code = str(vals[0]).strip() if vals[0] is not None else ""
            desc_en = str(vals[1]).strip() if vals[1] is not None else ""
            desc_zh = str(vals[2]).strip() if vals[2] is not None else ""
            unit = str(vals[3]).strip() if vals[3] is not None else ""
            qty = safe_qty(vals[4])
            if not code and not desc_en and qty is None: continue
            if code.lower() == "item": continue
            rows.append({"row":row_idx, "code":code, "desc":desc_en, "desc_zh":desc_zh,
                         "unit":normalize_unit(unit), "qty":qty, "source":"TENDER"})
    wb.close()
    return rows

# ═══════════════════════════════════════════════════════
# 行分类
# ═══════════════════════════════════════════════════════

def classify_rows(raw_rows, source):
    """分类：CLASS标题 / Section标题 / BOQ条目，标注 class_id, section_id, section_desc, option"""
    current_class = None; current_section = None; current_section_desc = ""
    current_option = None
    classified = []
    for r in raw_rows:
        code = r["code"]; desc = r.get("desc", "")
        unit = r.get("unit")
        # qty可能是0，不能用 or 链（0是falsy）
        qty = r.get("qty")
        if qty is None: qty = r.get("design_qty")
        if qty is None: qty = r.get("bid_qty")

        # Class header (WTCC/SGHCC/FHDI: "Class X", TENDER: "CLASS X")
        cm = CLASS_RE.match(code) if code else None
        if not cm and source == "TENDER":
            cm = TENDER_CLASS_RE.match(code) if code else None
        if cm:
            current_class = cm.group(1).upper()
            current_section = None; current_section_desc = ""
            current_option = None  # 换Class时重置option
            if source == "FHDI":
                opt = get_option_number(desc)
                if opt: current_option = opt
            classified.append({**r, "class_id":current_class, "section_id":None,
                               "section_desc":"", "row_type":"CLASS_HEADER",
                               "option":current_option})
            continue

        # FHDI: 在当前class内检测方案标记（仅D/E类有方案切换）
        if source == "FHDI" and current_class in ("D", "E"):
            opt = get_option_number(desc)
            if opt: current_option = opt

        if current_class is None:
            current_class = "A" if source == "TENDER" else None
            if current_class is None: continue

        # Section header: 有层级code + 无单位 + 无数量
        code_match = ITEM_CODE_RE.match(code) if code else None
        has_qty = qty is not None

        if code and not has_qty and unit in (None, "", "None"):
            parts = code.split(".")
            if len(parts) <= 3:
                current_section = code
                current_section_desc = desc
                classified.append({**r, "class_id":current_class,
                                   "section_id":current_section,
                                   "section_desc":desc,
                                   "row_type":"SECTION_HEADER", "option":current_option})
                continue

        # BOQ item
        # 过滤范本说明文字（无code + 长描述匹配范本关键词）
        if not code:
            desc_lower = desc.lower()
            if any(kw in desc_lower for kw in [
                "bill of quantities is based on",
                "all quantities stated are to be verified",
                "employer will accept no liability",
                "tenderer's responsibility",
                "tenderer shall separately list",
            ]):
                continue

        clean_code, is_add = parse_item_code(code)
        classified.append({**r, "class_id":current_class, "section_id":current_section,
                           "section_desc":current_section_desc,
                           "row_type":"ITEM", "option":current_option,
                           "is_add":is_add, "clean_code":clean_code or code})
    return classified


def get_item_rows(classified):
    """只取 ITEM 行 + SECTION_HEADER 行，FHDI不去重（方案一/二分别填不同列）"""
    return [r for r in classified if r["row_type"] in ("ITEM", "SECTION_HEADER")]

# ═══════════════════════════════════════════════════════
# 匹配引擎
# ═══════════════════════════════════════════════════════

def match_to_wtcc(src_items, wtcc_items):
    """将 src_items 匹配到 wtcc_items。返回 {wtcc_idx: (src_item, confidence)} 和 unmatched。
    同一wtcc行已被匹配时，后续项归入unmatched（不覆写、不丢失）。"""
    matches = {}
    unmatched = []
    # 按 class 分组 wtcc
    wtcc_by_class = defaultdict(list)
    for i, item in enumerate(wtcc_items):
        wtcc_by_class[item.get("class_id","")].append((i, item))

    for sitem in src_items:
        cls = sitem.get("class_id","")
        candidates = wtcc_by_class.get(cls, [])
        if not cls:
            for c_items in wtcc_by_class.values():
                candidates.extend(c_items)

        scode = sitem.get("clean_code","")
        best_idx = None; best_conf = 0.0

        for wi, witem in candidates:
            wcode = witem.get("clean_code","")
            # 精确匹配
            if scode and wcode and scode == wcode:
                # 同时要求描述有一定相似度（防止同code不同子项全匹配到同一行）
                sim = desc_similarity(sitem.get("desc",""), witem.get("desc",""))
                if sim >= 0.5:
                    best_idx = wi; best_conf = 1.0; break
                else:
                    best_idx = wi; best_conf = 0.6  # 弱匹配，允许后续被覆盖
            # 父级匹配
            if scode and wcode and wcode.startswith(scode + "."):
                if 0.85 > best_conf: best_idx = wi; best_conf = 0.85
        # 描述匹配
        if best_idx is None and cls:
            for wi, witem in candidates:
                sim = desc_similarity(sitem.get("desc",""), witem.get("desc",""))
                if sim > best_conf: best_conf = sim; best_idx = wi
            if best_conf < 0.75: best_idx = None

        if best_idx is not None:
            if best_idx in matches:
                # 该WTCC行已有匹配 → 当前项放入unmatched（不覆写，不丢失）
                unmatched.append(sitem)
            else:
                matches[best_idx] = (sitem, best_conf)
        else:
            unmatched.append(sitem)
    return matches, unmatched

# ═══════════════════════════════════════════════════════
# 统一行构建
# ═══════════════════════════════════════════════════════

class UnifiedRow:
    """统一行: 10列数据 + 元信息"""
    __slots__ = ("item","desc","unit","tender_qty","wtcc_design","wtcc_bid",
                 "fhdi_opt1","fhdi_opt2","sghcc_qty","spec","row_type","class_id",
                 "section_id","is_add","outline_level","source_note","fhdi_opt1_unmatched",
                 "fhdi_opt2_unmatched","sghcc_unmatched","tender_unmatched")
    def __init__(self, item="", desc="", unit="", row_type="ITEM", class_id="", section_id=""):
        self.item = item; self.desc = desc; self.unit = unit
        self.tender_qty = None; self.wtcc_design = None; self.wtcc_bid = None
        self.fhdi_opt1 = None; self.fhdi_opt2 = None; self.sghcc_qty = None
        self.spec = ""; self.row_type = row_type; self.class_id = class_id
        self.section_id = section_id or ""; self.is_add = False
        self.outline_level = 0; self.source_note = ""


def build_unified_rows(wtcc_cls, fhdi_cls, sghcc_cls, tender_cls):
    """构建统一行列表"""
    wtcc_items = [r for r in wtcc_cls if r["row_type"]=="ITEM"]
    wtcc_all = [r for r in wtcc_cls]

    fhdi_all = [r for r in fhdi_cls if r["row_type"]=="ITEM"]
    fhdi_items = [r for r in fhdi_all if r.get("option") != 2]
    fhdi_opt2_items = [r for r in fhdi_all if r.get("option") == 2]
    sghcc_items = [r for r in sghcc_cls if r["row_type"]=="ITEM"]
    tender_items = [r for r in tender_cls if r["row_type"]=="ITEM"]

    print(f"  [匹配] 各源候选: FHDI方案一={len(fhdi_items)}, FHDI方案二={len(fhdi_opt2_items)}, SGHCC={len(sghcc_items)}, Tender={len(tender_items)}")

    # ── 匹配 ──
    fhdi_matches, fhdi_unmatched = match_to_wtcc(fhdi_items, wtcc_items)
    fhdi_opt2_matches, fhdi_opt2_unmatched = match_to_wtcc(fhdi_opt2_items, wtcc_items)
    sghcc_matches, sghcc_unmatched = match_to_wtcc(sghcc_items, wtcc_items)
    tender_matches, tender_unmatched = match_to_wtcc(tender_items, wtcc_items)

    print(f"  [匹配] FHDI方案一 matched={len(fhdi_matches)} unmatched={len(fhdi_unmatched)}")
    print(f"  [匹配] FHDI方案二 matched={len(fhdi_opt2_matches)} unmatched={len(fhdi_opt2_unmatched)}")
    print(f"  [匹配] SGHCC matched={len(sghcc_matches)} unmatched={len(sghcc_unmatched)}")
    print(f"  [匹配] Tender matched={len(tender_matches)} unmatched={len(tender_unmatched)}")

    # ── 构建 section 查找表 ──
    def _build_section_map(classified_rows):
        """{class_id: OrderedDict[section_id → section_desc]}"""
        result = defaultdict(OrderedDict)
        for r in classified_rows:
            if r["row_type"] == "SECTION_HEADER":
                sid = r.get("section_id","")
                if sid:
                    result[r["class_id"]][sid] = r.get("desc","")
        return result

    wtcc_sec_map = _build_section_map(wtcc_cls)
    fhdi_sec_map = _build_section_map(fhdi_cls)
    sghcc_sec_map = _build_section_map(sghcc_cls)
    tender_sec_map = _build_section_map(tender_cls)

    SRC_MAPS = {"FHDI":fhdi_sec_map, "FHDI-OPT2":fhdi_sec_map,
                "SGHCC":sghcc_sec_map, "TENDER":tender_sec_map}

    # ── Section 匹配 ──
    def find_target_section(item, class_id):
        """3级匹配: 精确code→描述相似→None（需新建section）"""
        wtcc_secs = wtcc_sec_map.get(class_id, {})
        if not wtcc_secs:
            return None
        code = item.get("code","")
        src_sec_id = item.get("section_id","")

        # 策略1: code前缀匹配 (F.2.1.1 → F.2)
        parts = code.split(".")
        if len(parts) >= 2:
            prefix = ".".join(parts[:2])
            if prefix in wtcc_secs:
                return prefix
        if src_sec_id and src_sec_id in wtcc_secs:
            return src_sec_id

        # 策略2: section描述相似度
        src_sec_desc = item.get("section_desc","")
        if src_sec_desc:
            best_sim = 0.5; best_sec = None
            for sec_id, sec_desc in wtcc_secs.items():
                sim = desc_similarity(src_sec_desc, sec_desc)
                if sim > best_sim:
                    best_sim = sim; best_sec = sec_id
            if best_sec:
                return best_sec
        return None

    # ── 辅助: 创建独有项行 ──
    def _make_unmatched_row(it, class_id, source_label):
        ur = UnifiedRow(item=it["code"], desc=it["desc"], unit=it.get("unit",""),
                        row_type="ITEM", class_id=class_id,
                        section_id=it.get("section_id",""))
        ur.source_note = f"{source_label}独有"
        if source_label == "FHDI":
            ur.fhdi_opt1 = it.get("qty")
        elif source_label == "FHDI-OPT2":
            ur.fhdi_opt2 = it.get("qty")
            ur.source_note = "FHDI独有(方案二)"
        elif source_label == "SGHCC":
            ur.sghcc_qty = it.get("qty")
        elif source_label == "TENDER":
            ur.tender_qty = it.get("qty")
        return ur

    def _write_section_unmatched(unified_list, class_id, section_id, source_label, items):
        """在指定section末尾写独有项三级标题+条目"""
        if not items:
            return
        sep = UnifiedRow(item="", desc=f"{{{source_label}独有项}}",
                         row_type="SECTION_HEADER", class_id=class_id,
                         section_id=section_id)
        sep.outline_level = 2  # {} = 三级标题
        unified_list.append(sep)
        for it in items:
            unified_list.append(_make_unmatched_row(it, class_id, source_label))

    def _create_new_sections(unified_list, class_id, source_label, items, src_sec_map):
        """为无法匹配的项按源section分组新建《》"""
        if not items:
            return
        by_src_sec = defaultdict(list)
        for it in items:
            sid = it.get("section_id","") or "__no_sec__"
            by_src_sec[sid].append(it)
        for src_sec_id, sec_items in by_src_sec.items():
            sec_desc = src_sec_map.get(class_id, {}).get(src_sec_id, "")
            if src_sec_id == "__no_sec__":
                label = f"《{source_label}独有项》"
            elif sec_desc:
                label = f"《{sec_desc} — {source_label}独有》"
            else:
                label = f"《{src_sec_id} — {source_label}独有》"
            new_sec = UnifiedRow(item="", desc=label, row_type="SECTION_HEADER",
                                class_id=class_id, section_id=src_sec_id)
            new_sec.outline_level = 2
            new_sec.source_note = f"{source_label}独有"
            unified_list.append(new_sec)
            for it in sec_items:
                unified_list.append(_make_unmatched_row(it, class_id, source_label))

    # ── 预处理: 未匹配项按 (class_id, target_section) 分组 ──
    def _group_unmatched_by_section(unmatched_items):
        """{class_id: {target_section_id|None: [items]}}"""
        result = defaultdict(lambda: defaultdict(list))
        for it in unmatched_items:
            cid = it.get("class_id","")
            target = find_target_section(it, cid)
            result[cid][target].append(it)
        return result

    fhdi_un = _group_unmatched_by_section(fhdi_unmatched)
    fhdi_o2_un = _group_unmatched_by_section(fhdi_opt2_unmatched)
    sghcc_un = _group_unmatched_by_section(sghcc_unmatched)
    tender_un = _group_unmatched_by_section(tender_unmatched)

    def _process_class_sections(unified_list, cid, cls_rows, is_tender_only=False):
        """处理一个CLASS: 按section分组WTCC行，逐section插入独有项"""
        # 分离 CLASS_HEADER / SECTION_HEADER / ITEM
        class_header = None
        sections = OrderedDict()  # section_id → {"header": row, "items": [rows]}
        current_sec = None
        for r in cls_rows:
            if r["row_type"] == "CLASS_HEADER":
                class_header = r
            elif r["row_type"] == "SECTION_HEADER":
                current_sec = r.get("section_id","")
                if current_sec not in sections:
                    sections[current_sec] = {"header": r, "items": []}
            elif r["row_type"] == "ITEM":
                if current_sec not in sections:
                    sections[current_sec] = {"header": None, "items": []}
                sections[current_sec]["items"].append(r)

        # 写 CLASS_HEADER
        if class_header:
            ur = UnifiedRow(item=class_header["code"], desc=class_header["desc"],
                           row_type="CLASS_HEADER", class_id=cid)
            ur.outline_level = 1
            unified_list.append(ur)

        # 逐 section 处理
        for sec_id, sec_data in sections.items():
            # 写 SECTION_HEADER
            if sec_data["header"]:
                ur = UnifiedRow(item=sec_data["header"]["code"],
                               desc=sec_data["header"]["desc"],
                               row_type="SECTION_HEADER", class_id=cid,
                               section_id=sec_id)
                ur.outline_level = 2
                unified_list.append(ur)

            # 写该 section 下的 WTCC/TENDER 条目
            for r in sec_data["items"]:
                if is_tender_only:
                    ur = UnifiedRow(item=r["code"], desc=r.get("desc",""),
                                    unit=r.get("unit",""), row_type="ITEM",
                                    class_id=cid, section_id=r.get("section_id",""))
                    ur.tender_qty = r.get("qty")
                else:
                    wtcc_idx = wtcc_items.index(r) if r in wtcc_items else -1
                    ur = UnifiedRow(item=r["code"], desc=r["desc"],
                                    unit=r.get("unit",""), row_type="ITEM",
                                    class_id=cid, section_id=r.get("section_id",""))
                    ur.wtcc_design = r.get("design_qty")
                    ur.wtcc_bid = r.get("bid_qty")
                    ur.spec = r.get("spec","")
                    ur.is_add = r.get("is_add", False)
                    if ur.is_add:
                        ur.source_note = "WTCC ADD"
                    if wtcc_idx in fhdi_matches:
                        ur.fhdi_opt1 = fhdi_matches[wtcc_idx][0].get("qty")
                    if wtcc_idx in fhdi_opt2_matches:
                        ur.fhdi_opt2 = fhdi_opt2_matches[wtcc_idx][0].get("qty")
                    if wtcc_idx in sghcc_matches:
                        ur.sghcc_qty = sghcc_matches[wtcc_idx][0].get("qty")
                    if wtcc_idx in tender_matches:
                        ur.tender_qty = tender_matches[wtcc_idx][0].get("qty")
                unified_list.append(ur)

            # 插入映射到此 section 的独有项（按 source 顺序）
            if not is_tender_only:
                for src_label, un_dict in [("TENDER",tender_un), ("FHDI",fhdi_un),
                                           ("FHDI-OPT2",fhdi_o2_un), ("SGHCC",sghcc_un)]:
                    if cid in un_dict and sec_id in un_dict[cid]:
                        items = un_dict[cid].pop(sec_id)
                        _write_section_unmatched(unified_list, cid, sec_id, src_label, items)

        # ── 该 CLASS 剩余无法匹配的 → 新建《》section ──
        if not is_tender_only:
            for src_label, un_dict, src_map in [
                ("TENDER", tender_un, tender_sec_map),
                ("FHDI", fhdi_un, fhdi_sec_map),
                ("FHDI-OPT2", fhdi_o2_un, fhdi_sec_map),
                ("SGHCC", sghcc_un, sghcc_sec_map),
            ]:
                if cid in un_dict:
                    remaining = []
                    for items in un_dict[cid].values():
                        remaining.extend(items)
                    if remaining:
                        _create_new_sections(unified_list, cid, src_label, remaining, src_map)
                    del un_dict[cid]

    unified = []

    # ── 1. Class A（招标独有，WTCC无）──
    tender_a_rows = [r for r in tender_cls if r.get("class_id")=="A"]
    if tender_a_rows:
        # 把 TENDER 未匹配的 A 类项合并进去
        if "A" in tender_un:
            for target, items in tender_un["A"].items():
                for it in items:
                    tender_a_rows.append(it)
            del tender_un["A"]
        _process_class_sections(unified, "A", tender_a_rows, is_tender_only=True)

    # ── 2. Classes B-I (WTCC 模板) ──
    wtcc_by_class = OrderedDict()
    for r in wtcc_all:
        cid = r.get("class_id","")
        if cid not in wtcc_by_class:
            wtcc_by_class[cid] = []
        wtcc_by_class[cid].append(r)

    for cid in wtcc_by_class:
        _process_class_sections(unified, cid, wtcc_by_class[cid])

    # ── 3. 兜底：遗留未匹配项 ──
    for src_label, un_dict, src_map in [
        ("TENDER", tender_un, tender_sec_map),
        ("FHDI", fhdi_un, fhdi_sec_map),
        ("FHDI-OPT2", fhdi_o2_un, fhdi_sec_map),
        ("SGHCC", sghcc_un, sghcc_sec_map),
    ]:
        for cid in list(un_dict.keys()):
            remaining = []
            for items in un_dict[cid].values():
                remaining.extend(items)
            if remaining:
                _create_new_sections(unified, cid, src_label, remaining, src_map)

    return unified

# ═══════════════════════════════════════════════════════
# xlsxwriter 输出
# ═══════════════════════════════════════════════════════

FONT_NAME = "Microsoft YaHei UI"
HEADER_FILL = "#1F4E79"
HEADER_FONT = "#FFFFFF"
CLASS_FILL = "#C6D9F1"
SECTION_FILL = "#EEF2FA"
ADD_FILL = "#FBE5D6"
ONLY_FILL = "#FFF2CC"
NUM_FORMAT = '#,##0.00'

COL_WIDTHS = [22, 55, 8, 14, 14, 14, 14, 14, 14, 40]
COL_HEADERS = [
    "Item", "Item Description", "Unit",
    "招标标准清单", "WTCC Design", "WTCC Bid",
    "FHDI 方案一", "FHDI 方案二", "SGHCC",
    "Main Specification"
]

# 四级分级样式 (merge_boq.py 配色)
LEVEL_STYLES = {
    1: {'bold':True,  'bg_color':'#C6D9F1', 'font_size':11, 'row_height':24},  # 【】
    2: {'bold':True,  'bg_color':'#EEF2FA', 'font_size':10, 'row_height':20},  # 《》
    3: {'bold':False, 'bg_color':'#FBE5D6', 'font_size':9,  'row_height':18},  # {}
    4: {'bold':False, 'bg_color':None,       'font_size':9,  'row_height':16},  # 分项
}


def _detect_level(desc):
    """从描述前缀检测分级: 【→1, 《→2, {→3, 其他→4"""
    if not desc: return 4
    if desc.startswith('【'): return 1
    if desc.startswith('《'): return 2
    if desc.startswith('{'): return 3
    return 4


def write_xlsx(unified_rows, output_path):
    wb = xlsxwriter.Workbook(str(output_path))
    ws = wb.add_worksheet("MergedBOQ")

    # 预计算所有行的显示级别
    levels = [_detect_level(ur.desc) for ur in unified_rows]

    # 构建每级格式 (level, is_number) → format
    cell_fmts = {}
    for lv in [1, 2, 3, 4]:
        s = LEVEL_STYLES[lv]
        base = {
            'font_name': FONT_NAME, 'font_color': '#1A1A1A',
            'bold': s['bold'], 'font_size': s['font_size'],
            'border': 1, 'border_color': '#D9D9D9',
            'valign': 'vcenter',
        }
        if s['bg_color']:
            base['bg_color'] = s['bg_color']
        cell_fmts[(lv, False)] = wb.add_format(base)
        cell_fmts[(lv, True)] = wb.add_format({**base, 'num_format': NUM_FORMAT})

    # 特殊格式
    add_fmt = wb.add_format({
        'font_name': FONT_NAME, 'font_size': 9, 'font_color': '#1A1A1A',
        'border': 1, 'border_color': '#D9D9D9', 'valign': 'vcenter',
        'num_format': NUM_FORMAT, 'bg_color': '#FBE5D6',
    })
    only_fmt = wb.add_format({
        'font_name': FONT_NAME, 'font_size': 9, 'font_color': '#1A1A1A',
        'border': 1, 'border_color': '#D9D9D9', 'valign': 'vcenter',
        'num_format': NUM_FORMAT, 'bg_color': '#FFF2CC',
    })
    hdr_fmt = wb.add_format({
        'bold': True, 'font_color': '#FFFFFF', 'bg_color': '#1F4E79',
        'font_name': FONT_NAME, 'font_size': 9, 'border': 1,
        'border_color': '#D9D9D9', 'text_wrap': True, 'valign': 'vcenter',
        'align': 'center',
    })
    bold_fmt = wb.add_format({
        'bold': True, 'font_name': FONT_NAME, 'font_size': 9, 'font_color': '#1A1A1A',
        'border': 1, 'border_color': '#D9D9D9', 'valign': 'vcenter',
    })

    # 表头
    ws.freeze_panes(1, 0)
    ws.autofilter(0, 0, len(unified_rows), len(COL_HEADERS) - 1)
    for c, h in enumerate(COL_HEADERS):
        ws.write(0, c, h, hdr_fmt)
    ws.set_row(0, 28)

    # 列宽
    for c, w in enumerate(COL_WIDTHS):
        ws.set_column(c, c, w)

    # 写数据行
    for i, ur in enumerate(unified_rows):
        r = i + 1
        lv = levels[i]
        s = LEVEL_STYLES[lv]
        outline_lv = lv - 1  # Excel行分组: 【→0, 《→1, {→2, 分项→3

        # 行高 + 分组级别
        ws.set_row(r, s['row_height'], None, {'level': outline_lv})

        # 选择格式
        is_add = ur.is_add
        is_only = bool(ur.source_note) and not is_add
        is_header = lv <= 2  # 【】和《》是标题行

        for ci in range(len(COL_HEADERS)):
            val = None
            if ci == 0: val = ur.item
            elif ci == 1: val = ur.desc
            elif ci == 2: val = ur.unit
            elif ci == 3: val = ur.tender_qty
            elif ci == 4: val = ur.wtcc_design
            elif ci == 5: val = ur.wtcc_bid
            elif ci == 6: val = ur.fhdi_opt1
            elif ci == 7: val = ur.fhdi_opt2
            elif ci == 8: val = ur.sghcc_qty
            elif ci == 9: val = ur.spec if not is_header else ""

            if val is None:
                continue

            is_num_col = ci >= 3
            is_num_val = isinstance(val, (int, float))

            if is_add:
                fmt = add_fmt
            elif is_only and not is_header:
                fmt = only_fmt
            elif is_header and ci == 0:
                fmt = cell_fmts[(lv, False)]
            else:
                fmt = cell_fmts[(lv, is_num_col and is_num_val)]

            if is_num_col and is_num_val:
                ws.write_number(r, ci, val, fmt)
            else:
                ws.write(r, ci, val, fmt)

    wb.close()
    return output_path

# ═══════════════════════════════════════════════════════
# InstituteMerger 类
# ═══════════════════════════════════════════════════════

SOURCE_LABELS = {
    "TENDER": "招标", "WTCC": "WTCC", "FHDI": "FHDI",
    "FHDI-OPT2": "FHDI-OPT2", "SGHCC": "SGHCC",
}


class InstituteMerger:
    """多家设计院 BOQ 全量合并器。

    Usage:
        merger = InstituteMerger()
        merger.merge(wtcc_path, fhdi_path, sghcc_path, tender_path, output_path)
    """

    def __init__(self, report_date: str = ""):
        self.report_date = report_date or datetime.now().strftime("%Y-%m-%d")

    def merge(self, wtcc_path, fhdi_path, sghcc_path, tender_path, output_path=None):
        """执行全量合并，生成统一 Excel。

        Args:
            wtcc_path: WTCC(三航院) BQMerge xlsx 文件路径（作为模板基准）
            fhdi_path: FHDI(一航院) BQMerge xlsx 文件路径
            sghcc_path: SGHCC(水规院) BQMerge xlsx 文件路径
            tender_path: 招标标准清单 xlsx 文件路径
            output_path: 输出 .xlsx 路径，默认为 {cwd}/{date}_BOQ_Merged_All_Institutes.xlsx

        Returns:
            输出文件路径
        """
        if output_path is None:
            # 输出目录选择：所有输入文件在同一目录 → 该目录；否则 → WTCC（基准）所在目录
            input_dirs = {Path(p).parent.resolve() for p in [wtcc_path, fhdi_path, sghcc_path, tender_path]}
            if len(input_dirs) == 1:
                out_dir = input_dirs.pop()
            else:
                out_dir = Path(wtcc_path).parent.resolve()
            output_path = out_dir / f"{self.report_date}_BOQ_Merged_All_Institutes.xlsx"
        else:
            output_path = Path(output_path)

        print("=" * 60)
        print("全量合并 BOQ 清单 → 统一 Excel")
        print("=" * 60)

        # 1. 读取
        print("\n[1/4] 读取文件...")
        wtcc_raw = read_wtcc(wtcc_path)
        fhdi_raw = read_fhdi(fhdi_path)
        sghcc_raw = read_sghcc(sghcc_path)
        tender_raw = read_tender(tender_path)
        print(f"  WTCC={len(wtcc_raw)}, FHDI={len(fhdi_raw)}, SGHCC={len(sghcc_raw)}, Tender={len(tender_raw)}")

        # 2. 分类
        print("\n[2/4] 分类 + 分级...")
        wtcc_cls = classify_rows(wtcc_raw, "WTCC")
        fhdi_cls = classify_rows(fhdi_raw, "FHDI")
        sghcc_cls = classify_rows(sghcc_raw, "SGHCC")
        tender_cls = classify_rows(tender_raw, "TENDER")

        fhdi_items_all = get_item_rows(fhdi_cls)
        fhdi_headers = [r for r in fhdi_cls if r["row_type"] == "CLASS_HEADER"]
        fhdi_cls = fhdi_headers + fhdi_items_all
        fhdi_item_count = sum(1 for r in fhdi_items_all if r["row_type"] == "ITEM")
        fhdi_qty_count = sum(1 for r in fhdi_items_all if r.get("qty") is not None)
        print(f"  FHDI保留全部: {fhdi_item_count} ITEM条目, {fhdi_qty_count} 有工程量")

        for src, data in [("WTCC", wtcc_cls), ("FHDI", fhdi_cls), ("SGHCC", sghcc_cls), ("Tender", tender_cls)]:
            items = [r for r in data if r["row_type"] == "ITEM"]
            headers = [r for r in data if r["row_type"] != "ITEM"]
            cls_dist = defaultdict(int)
            for it in items:
                cls_dist[it["class_id"]] += 1
            dist_str = ", ".join(f"{k}:{v}" for k, v in sorted(cls_dist.items()))
            print(f"  {src}: {len(items)} items + {len(headers)} headers, classes: {dist_str}")

        # 3. 构建统一行
        print("\n[3/4] 构建统一行...")
        unified = build_unified_rows(wtcc_cls, fhdi_cls, sghcc_cls, tender_cls)
        print(f"  共 {len(unified)} 行")
        types = Counter(r.row_type for r in unified)
        print(f"  CLASS标题={types.get('CLASS_HEADER',0)}, Section标题={types.get('SECTION_HEADER',0)}, 条目={types.get('ITEM',0)}")

        tender_n = sum(1 for r in unified if r.tender_qty is not None)
        wtcc_d_n = sum(1 for r in unified if r.wtcc_design is not None)
        wtcc_b_n = sum(1 for r in unified if r.wtcc_bid is not None)
        fhdi1_n = sum(1 for r in unified if r.fhdi_opt1 is not None)
        fhdi2_n = sum(1 for r in unified if r.fhdi_opt2 is not None)
        sghcc_n = sum(1 for r in unified if r.sghcc_qty is not None)
        print(f"  数据覆盖: 招标={tender_n}, WTCC-设计={wtcc_d_n}, WTCC-报价={wtcc_b_n}, FHDI-方案一={fhdi1_n}, FHDI-方案二={fhdi2_n}, SGHCC={sghcc_n}")

        # 4. 写 Excel
        print("\n[4/4] 写入 Excel...")
        write_xlsx(unified, output_path)
        print(f"  输出: {output_path}")
        print(f"  文件大小: {output_path.stat().st_size:,} bytes")
        return str(output_path)


# ═══════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="全量合并多家 BOQ 清单 → 统一 Excel",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python merge_all_institutes.py --wtcc WTCC.xlsx --fhdi FHDI.xlsx --sghcc SGHCC.xlsx --tender Tender.xlsx
  python merge_all_institutes.py --wtcc WTCC.xlsx --fhdi FHDI.xlsx --sghcc SGHCC.xlsx --tender Tender.xlsx -o merged.xlsx
  python merge_all_institutes.py --wtcc WTCC.xlsx --fhdi FHDI.xlsx --sghcc SGHCC.xlsx --tender Tender.xlsx --date 2026-05-04
        """,
    )
    parser.add_argument("--wtcc", required=True, help="WTCC(三航院) BQMerge xlsx 文件路径（作为模板基准）")
    parser.add_argument("--fhdi", required=True, help="FHDI(一航院) BQMerge xlsx 文件路径")
    parser.add_argument("--sghcc", required=True, help="SGHCC(水规院) BQMerge xlsx 文件路径")
    parser.add_argument("--tender", required=True, help="招标标准清单 xlsx 文件路径")
    parser.add_argument("-o", "--output", default=None, help="输出 .xlsx 路径（默认: {date}_BOQ_Merged_All_Institutes.xlsx）")
    parser.add_argument("--date", default="", help="输出日期前缀（默认今日）")
    args = parser.parse_args()

    merger = InstituteMerger(report_date=args.date)
    merger.merge(args.wtcc, args.fhdi, args.sghcc, args.tender, args.output)


if __name__ == "__main__":
    main()
