"""
工程量清单（BOQ）合并工具

将 Schedule of Prices 工作簿中的分部分项工程量清单合并到单一汇总表。
- 筛选分部分项清单 sheet（排除开办费、汇总、计日工、报价说明等）
- 自动检测列布局（标准/移位），归一化到统一列结构
- 取消合并单元格，保留所有原始数据
- 删除空行、Excel错误行、汇总行（SUBTOTAL/TOTAL/% COST/DITTO 等）
- 自动分级：一级【】、二级《》、三级{}、四级分项
- 层级配色 + 可折叠行分组 + 数字会计格式
- 输出端使用 xlsxwriter（Office 原生兼容，无 COM/分组/fill 问题）
"""
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
import xlsxwriter
from openpyxl.worksheet.worksheet import Worksheet


# ── 样式常量 ──────────────────────────────────────────────

FONT_NAME = "Microsoft YaHei UI"
FONT_COLOR = "#1A1A1A"
BORDER_COLOR = "#D9D9D9"
HEADER_FILL = "#1F4E79"
NUM_FORMAT = '#,##0.00'

LEVEL_FILLS = {1: "#C6D9F1", 2: "#EEF2FA", 3: "#FBE5D6", 4: None}
LEVEL_BOLD = {1: True, 2: True, 3: False, 4: False}
LEVEL_SIZE = {1: 11, 2: 10, 3: 9, 4: 9}
LEVEL_HEIGHT = {1: 24, 2: 20, 3: 18, 4: 16}


# ── 配置 ──────────────────────────────────────────────

SKIP_SHEET_PATTERNS: list[str] = [
    "Preamble", "Cashflow", "Standby Rates", "Dayworks",
    "Grand Summary", "BoQ Grand Summary",
]

PRELIMS_PREFIXES: tuple[str, ...] = ("A_", "A0", "A.", "Prelims", "Preliminaries")

REMOVE_PATTERNS: re.Pattern = re.compile(
    r"^\s*(SUBTOTAL|TOTAL|% COST|CARRIED FORWARD|BROUGHT FORWARD|DITTO|PAGE\s*TOTAL)\b",
    re.IGNORECASE,
)

HEADER_SKIP_KEYWORDS: tuple[str, ...] = (
    "Project:", "Contract:", "Subject:", "Part II", "Item B105",
    "Schedule of Prices", "Tender No",
)

DISCLAIMER_KEYWORDS: tuple[str, ...] = (
    "Bill of Quantities is based on",
    "Employer's Requirements",
    "Tenderer's responsibility",
    "All quantities stated are to be verified",
    "The Employer will accept no liability",
    "Tenderer shall separately list",
)

EXCEL_ERRORS: frozenset[str] = frozenset({
    "#REF!", "#VALUE!", "#N/A", "#DIV/0!", "#NAME?", "#NULL!",
})

COL_ITEM = 1
COL_DESC = 2
COL_UNIT = 3
COL_QTY = 4

CLASS_RE = re.compile(r"^Class\s+[A-Z]\b", re.IGNORECASE)
ITEM_CODE_RE = re.compile(r"^([A-Za-z]+\.\d+(?:\.\d+)*)\b")

OPTION_RE = re.compile(r"[（(]方案[一二三四五六七八九十]+\s*[）)]|\bOption\s*\d+\b", re.IGNORECASE)



# ── 辅助函数 ──────────────────────────────────────────

def _unmerge_and_fill(ws: Worksheet) -> None:
    merged = list(ws.merged_cells.ranges)
    for mrange in merged:
        top_left_val = ws.cell(row=mrange.min_row, column=mrange.min_col).value
        ws.unmerge_cells(str(mrange))
        for r in range(mrange.min_row, mrange.max_row + 1):
            for c in range(mrange.min_col, mrange.max_col + 1):
                if r == mrange.min_row and c == mrange.min_col:
                    continue
                ws.cell(row=r, column=c).value = top_left_val


def _cell_str(ws: Worksheet, row: int, col: int) -> str:
    val = ws.cell(row=row, column=col).value
    return str(val).strip() if val is not None else ""


def _is_excel_error(val: str) -> bool:
    return val in EXCEL_ERRORS


def _row_all_empty(ws: Worksheet, row: int, max_col: int) -> bool:
    for c in range(1, max_col + 1):
        if _cell_str(ws, row, c):
            return False
    return True


def _code_depth(code: str) -> int:
    return code.count(".")


def _is_remove_row(text: str) -> bool:
    return bool(REMOVE_PATTERNS.search(text))


def _looks_like_unit(val: str) -> bool:
    if not val:
        return False
    common_units = {
        "m", "m2", "m3", "m²", "m³", "㎡", "㎥",
        "kg", "t", "No.", "No", "NOS", "Sets",
        "LS", "ls", "Sum", "sum", "hr", "hour", "day", "week",
        "month", "nr", "pcs", "set", "sets", "each", "item", "lot",
        "km", "mm", "cm", "ha", "nos", "nos.", "hrs", "hours", "days",
        "weeks", "months", "tonne", "tonnes", "ton", "tons",
        "meter", "meters", "points", "sq.m", "sqm",
    }
    clean = val.strip()
    clean_lower = clean.lower()
    if clean_lower in {u.lower() for u in common_units}:
        return True
    # 去掉尾部句点再试（如 "m." → "m"、"kg." → "kg"、"sq.m." → "sq.m"）
    if clean_lower.endswith("."):
        stripped = clean_lower.rstrip(".")
        if stripped in {u.lower() for u in common_units}:
            return True
    return False



def _looks_like_qty(val: str) -> bool:
    if not val:
        return False
    clean = val.strip().replace(",", "").replace(" ", "")
    try:
        float(clean)
        return True
    except ValueError:
        pass
    return False


def _to_number(val: str):
    if not val:
        return val
    clean = val.strip().replace(",", "").replace(" ", "")
    if not clean:
        return val
    try:
        f = float(clean)
        return int(f) if f == int(f) and "." not in clean else f
    except ValueError:
        return val


def _find_unit_rate_col(ws: Worksheet, header_row: int, max_col: int) -> int:
    """返回 'Unit Rate' 表头的 0-based 列索引（默认 4 = E 列）。"""
    for row in (header_row, header_row + 1):
        for c in range(4, max_col + 1):
            val = _cell_str(ws, row, c + 1).lower()
            if "unit rate" in val and "breakdown" not in val:
                return c
    return 4


def _find_design_qty_col(ws: Worksheet, header_row: int, max_col: int) -> int:
    """返回 Design Quantity 列的 1-based 索引，若不存在则返回 0。"""
    for row in (header_row, header_row + 1):
        for c in range(1, max_col + 1):
            val = _cell_str(ws, row, c).lower()
            if "design" in val and "quantity" in val:
                return c
    return 0


def _read_headers(ws: Worksheet, header_row: int, max_col: int) -> list[str]:
    """从源表读取动态表头，处理双行表头合并。"""
    headers = []
    for c in range(1, max_col + 1):
        h1 = _cell_str(ws, header_row, c)
        h2 = _cell_str(ws, header_row + 1, c) if header_row + 1 <= (ws.max_row or 50) else ""
        if h1 and h2 and h1.lower() != h2.lower():
            headers.append(f"{h1}\n{h2}")
        elif h1:
            headers.append(h1)
        elif h2:
            headers.append(h2)
        else:
            headers.append("")
    while headers and not headers[-1]:
        headers.pop()
    return headers


def _extract_alias(name: str) -> str:
    parts = name.split("_", 1)
    if len(parts) == 2:
        return parts[1].strip()
    return name.strip()


def _extract_option_label(name: str) -> str:
    m = OPTION_RE.search(name)
    return m.group(0) if m else ""


# ── 主类 ──────────────────────────────────────────────

class BOQMerger:
    def __init__(self, source_path: str | Path):
        self.source_path = Path(source_path)
        self.wb = openpyxl.load_workbook(self.source_path, data_only=True)
        self.sheet_aliases: dict[str, str] = {}
        self.sheet_options: dict[str, str] = {}

    # ── Sheet 筛选 ──

    def _is_boq_sheet(self, name: str) -> bool:
        for pat in SKIP_SHEET_PATTERNS:
            if pat.lower() in name.lower():
                return False
        return True

    def _is_prelims(self, name: str) -> bool:
        for prefix in PRELIMS_PREFIXES:
            if name.lower().startswith(prefix.lower()):
                return True
        return False

    # ── 表头定位 ──

    def _find_header_row(self, ws: Worksheet, max_col: int) -> int:
        for r in range(1, min(ws.max_row or 200, 200) + 1):
            for c in range(1, max_col + 1):
                if "Item Description" in _cell_str(ws, r, c):
                    return r
        return 0

    def _find_data_start(self, ws: Worksheet, max_col: int) -> int:
        header_row = self._find_header_row(ws, max_col)
        if header_row == 0:
            return 1
        for r in range(header_row + 1, min(ws.max_row or 500, 500) + 1):
            val_a = _cell_str(ws, r, COL_ITEM)
            val_b = _cell_str(ws, r, COL_DESC)
            if not val_a and not val_b:
                continue
            if _is_excel_error(val_a) or _is_excel_error(val_b):
                continue
            if self._is_info_header(val_a) or self._is_info_header(val_b):
                continue
            if "Item Description" in val_a or "Item Description" in val_b:
                continue
            return r
        return header_row + 1

    def _is_info_header(self, row_text: str) -> bool:
        for kw in HEADER_SKIP_KEYWORDS:
            if kw.lower() in row_text.lower():
                return True
        return False

    def _is_disclaimer(self, val_b: str) -> bool:
        for kw in DISCLAIMER_KEYWORDS:
            if kw.lower() in val_b.lower():
                return True
        return False

    # ── 行分类 ──

    def _classify_row(
        self, ws: Worksheet, row: int, max_col: int,
        qty_col: int = COL_QTY,
    ) -> dict:
        val_a = _cell_str(ws, row, COL_ITEM)
        val_b = _cell_str(ws, row, COL_DESC)
        val_c = _cell_str(ws, row, COL_UNIT)
        val_d = _cell_str(ws, row, qty_col)

        if not val_a and not val_b and not val_c and not val_d:
            return {"action": "remove"}

        if _is_excel_error(val_a) or _is_excel_error(val_b):
            return {"action": "remove"}

        if self._is_info_header(val_a) or self._is_info_header(val_b):
            return {"action": "remove"}

        if "Item Description" in val_a or "Item Description" in val_b:
            return {"action": "remove"}

        if _is_remove_row(val_a) or _is_remove_row(val_b):
            return {"action": "remove"}

        if self._is_disclaimer(val_b):
            return {"action": "remove"}
        if not val_a and not val_c and not val_d:
            word_count = len(val_b.split())
            if word_count > 25:
                return {"action": "remove"}

        code_match = ITEM_CODE_RE.match(val_a)
        code = code_match.group(1) if code_match else ""
        depth = _code_depth(code) if code else 0

        has_unit = bool(val_c) and _looks_like_unit(val_c)
        has_qty = bool(val_d) and _looks_like_qty(val_d)

        # 核心规则：单位列和工程量列同时非空 → 一定是分部分项清单条目，不是标题
        # （标题行不会有工程量，即使 LS/No./Sum 带数量也是实际报价条目）
        if bool(val_c) and bool(val_d):
            return {"action": "keep", "level": 4,
                    "code": val_a, "desc": val_b,
                    "unit": val_c, "quantity": val_d}

        if CLASS_RE.match(val_a) or CLASS_RE.match(val_b):
            return {"action": "keep", "level": 2,
                    "code": val_a, "desc": val_b,
                    "unit": val_c, "quantity": val_d}

        if code:
            # Depth≥3 with raw unit or qty → Level 4 (sub-items follow Level 3 titles)
            if has_unit or has_qty:
                return {"action": "keep", "level": 4,
                        "code": val_a, "desc": val_b,
                        "unit": val_c, "quantity": val_d}
            elif depth == 1:
                return {"action": "keep", "level": 2,
                        "code": val_a, "desc": val_b,
                        "unit": val_c, "quantity": val_d}
            elif depth >= 3:
                # Deep code without detected unit/qty — check raw values
                if bool(val_c) or bool(val_d):
                    return {"action": "keep", "level": 4,
                            "code": val_a, "desc": val_b,
                            "unit": val_c, "quantity": val_d}
                return {"action": "keep", "level": 3,
                        "code": val_a, "desc": val_b,
                        "unit": val_c, "quantity": val_d}
            else:
                # depth == 2: Level 3 title (followed by depth≥3 sub-items)
                return {"action": "keep", "level": 3,
                        "code": val_a, "desc": val_b,
                        "unit": val_c, "quantity": val_d}

        if val_a:
            if has_unit or has_qty:
                return {"action": "keep", "level": 4,
                        "code": val_a, "desc": val_b,
                        "unit": val_c, "quantity": val_d}
            else:
                return {"action": "keep", "level": 3,
                        "code": val_a, "desc": val_b,
                        "unit": val_c, "quantity": val_d}

        if val_b:
            if has_unit or has_qty:
                return {"action": "keep", "level": 4,
                        "code": "", "desc": val_b,
                        "unit": val_c, "quantity": val_d}
            # Raw unit or qty present → Level 4 (preserve the data)
            if bool(val_c) or bool(val_d):
                return {"action": "keep", "level": 4,
                        "code": "", "desc": val_b,
                        "unit": val_c, "quantity": val_d}
            word_count = len(val_b.split())
            if word_count <= 10 and not val_b.endswith("."):
                return {"action": "keep", "level": 3,
                        "code": "", "desc": val_b,
                        "unit": val_c, "quantity": val_d}
            return {"action": "keep", "level": 4,
                    "code": "", "desc": val_b,
                    "unit": val_c, "quantity": val_d}

        return {"action": "keep", "level": 0,
                "code": val_a, "desc": val_b,
                "unit": val_c, "quantity": val_d}

    # ── 格式化 ──

    def _format_desc(self, row_info: dict, sheet_name: str) -> str:
        desc = row_info["desc"]
        level = row_info["level"]
        if level == 1:
            text = desc.strip()
            option_label = self.sheet_options.get(sheet_name, "")
            if option_label:
                return f"【{text} — {option_label}】"
            return f"【{text}】"
        elif level == 2:
            return f"《{desc.strip()}》"
        elif level == 3:
            return f"{{{desc.strip()}}}"
        else:
            return desc

    # ── 主流程 ──

    def merge(self, output_path: Optional[str | Path] = None,
              keep_source_sheets: bool = False,
              use_outline: bool = True,
              use_write_blank: bool = False) -> Path:
        if output_path is None:
            stem = self.source_path.stem
            output_path = self.source_path.parent / f"{datetime.now().strftime('%Y-%m-%d')}_BQMerge_{stem}.xlsx"
        else:
            output_path = Path(output_path)

        # ── Phase 1: 读取 & 分类 (openpyxl) ──
        all_sheets = self.wb.sheetnames
        boq_sheets: list[str] = []
        for name in all_sheets:
            if self._is_boq_sheet(name) and not self._is_prelims(name):
                boq_sheets.append(name)
                self.sheet_aliases[name] = _extract_alias(name)
                self.sheet_options[name] = _extract_option_label(name)

        # 动态检测表头和列数：遍历所有 sheet 取最大列数和最佳表头
        dynamic_headers: list[str] = []
        global_unit_rate_col = 4  # 默认 E 列

        for name in boq_sheets:
            ws = self.wb[name]
            sheet_max_col = ws.max_column or 50
            header_row = self._find_header_row(ws, sheet_max_col)
            if header_row:
                hdrs = _read_headers(ws, header_row, sheet_max_col)
                if len(hdrs) > len(dynamic_headers):
                    dynamic_headers = hdrs
                urc = _find_unit_rate_col(ws, header_row, sheet_max_col)
                if urc > global_unit_rate_col:
                    global_unit_rate_col = urc

        if not dynamic_headers:
            dynamic_headers = ["Item", "Item Description", "Unit", "Quantity",
                               "Unit Rate\n(in USD excl. Taxes)",
                               "Labour", "Plant", "Material", "Subcontractor",
                               "Others / Consultants",
                               "Off-Site Overheads (Local & Head Office)",
                               "Local & Head Office Profit",
                               "Total Price\n(in USD excl. Taxes)"]

        global_max_col = len(dynamic_headers)
        # 统一数量列表头为 "Quantity"
        if global_max_col > 3:
            dynamic_headers[3] = "Quantity"
        all_rows: list[tuple[list, int]] = []  # (row_data, level)

        for sheet_name in boq_sheets:
            ws = self.wb[sheet_name]
            _unmerge_and_fill(ws)

            sheet_max_col = ws.max_column or global_max_col
            header_row = self._find_header_row(ws, sheet_max_col)

            start_row = self._find_data_start(ws, sheet_max_col)
            if start_row == 1:
                continue

            # 检测本 sheet 的数量列：优先 Design Quantity，其次默认 COL_QTY
            design_qty_col = 0
            if header_row:
                design_qty_col = _find_design_qty_col(ws, header_row, sheet_max_col)
            qty_col = design_qty_col if design_qty_col else COL_QTY

            title_row = [None] * global_max_col
            title_row[1] = f"【{sheet_name}】"
            all_rows.append((title_row, 1))

            for r in range(start_row, (ws.max_row or 500) + 1):
                if _row_all_empty(ws, r, sheet_max_col):
                    continue

                info = self._classify_row(ws, r, sheet_max_col, qty_col)
                if info["action"] == "remove":
                    continue

                level = info["level"]
                if level == 0:
                    level = 4
                desc = self._format_desc(info, sheet_name)

                # 保留全部源列：A-D(分类用) + E 起全部数据列
                raw_row = [info["code"], desc, info["unit"], info["quantity"]]
                for c in range(5, sheet_max_col + 1):
                    if c == design_qty_col:
                        raw_row.append(None)
                    else:
                        val = _cell_str(ws, r, c)
                        if _is_excel_error(val):
                            val = ""
                        raw_row.append(val)

                # 补齐到统一列数
                while len(raw_row) < global_max_col:
                    raw_row.append("")

                # 数值转换（Quantity 列起，含 Tender/Design Qty、Unit Rate 等）
                for ci in range(3, len(raw_row)):
                    if raw_row[ci]:
                        raw_row[ci] = _to_number(raw_row[ci])
                    elif raw_row[ci] == "":
                        raw_row[ci] = None

                all_rows.append((raw_row, level))

        self.wb.close()

        # ── Phase 2: 写入 (xlsxwriter) ──
        out_wb = xlsxwriter.Workbook(str(output_path), {'constant_memory': False})
        out_ws = out_wb.add_worksheet("MergeSheet")

        cell_fmts = {}
        for lvl in [1, 2, 3, 4]:
            base = {
                'font_name': FONT_NAME,
                'font_color': FONT_COLOR,
                'bold': LEVEL_BOLD[lvl],
                'font_size': LEVEL_SIZE[lvl],
                'bottom': 1,
                'bottom_color': BORDER_COLOR,
                'valign': 'vcenter',
            }
            if LEVEL_FILLS[lvl]:
                base['bg_color'] = LEVEL_FILLS[lvl]

            cell_fmts[(lvl, False)] = out_wb.add_format(base)
            cell_fmts[(lvl, True)] = out_wb.add_format({**base, 'num_format': NUM_FORMAT})

        header_fmt = out_wb.add_format({
            'font_name': FONT_NAME,
            'bold': True,
            'font_color': '#FFFFFF',
            'bg_color': HEADER_FILL,
            'border': 1,
            'border_color': BORDER_COLOR,
            'align': 'center',
            'valign': 'vcenter',
            'text_wrap': True,
            'font_size': 10,
        })

        used_cols = global_max_col

        # 写表头（动态）
        out_ws.set_row(0, 32, header_fmt)
        for ci in range(used_cols):
            hdr = dynamic_headers[ci] if ci < len(dynamic_headers) else ""
            out_ws.write(0, ci, hdr, header_fmt)

        # 计算 outline level
        outline_lv = [0] * len(all_rows)
        if use_outline:
            for i, (_, lvl) in enumerate(all_rows):
                outline_lv[i] = lvl - 1  # 【1→0, 《2→1, {3→2, 分项4→3

        # 写数据行
        for i, (row_data, level) in enumerate(all_rows):
            xl_row = i + 2
            if outline_lv[i] > 0:
                out_ws.set_row(xl_row, LEVEL_HEIGHT[level], None,
                               {'level': outline_lv[i]})
            else:
                out_ws.set_row(xl_row, LEVEL_HEIGHT[level])
            for ci, val in enumerate(row_data):
                if ci >= used_cols:
                    break
                is_num = ci >= 3
                fmt = cell_fmts[(level, is_num)]
                if val is None:
                    if use_write_blank:
                        out_ws.write_blank(xl_row, ci, None, fmt)
                else:
                    out_ws.write(xl_row, ci, val, fmt)

        # 列宽
        col_widths = {0: 18, 1: 60, 2: 8}
        for c in range(3, used_cols):
            col_widths[c] = 14
        for c, w in col_widths.items():
            out_ws.set_column(c, c, w)

        # 冻结 + 筛选
        out_ws.freeze_panes(1, 0)
        out_ws.autofilter(0, 0, len(all_rows) + 1, used_cols - 1)

        # 源分表（纯数据，无格式）
        if keep_source_sheets:
            for name in boq_sheets:
                src_ws = self.wb[name] if name in self.wb.sheetnames else None
                if src_ws is None:
                    tmp_wb = openpyxl.load_workbook(self.source_path, data_only=True)
                    src_ws = tmp_wb[name]
                else:
                    tmp_wb = None

                safe_name = name[:31]
                dst_ws = out_wb.add_worksheet(safe_name)
                _unmerge_and_fill(src_ws)
                for r in range(1, (src_ws.max_row or 100) + 1):
                    for c in range(0, (src_ws.max_column or 20)):
                        val = src_ws.cell(row=r, column=c + 1).value
                        if val is not None:
                            dst_ws.write(r - 1, c, val)

                if tmp_wb:
                    tmp_wb.close()

        out_wb.close()
        return output_path


def main():
    import argparse
    parser = argparse.ArgumentParser(description="合并工程量清单 (BOQ)")
    parser.add_argument("source", help="源 Excel 文件路径")
    parser.add_argument("-o", "--output", default=None, help="输出文件路径")
    parser.add_argument("--keep-source-sheets", action="store_true",
                        help="在输出中保留原始分表（默认不保留）")
    parser.add_argument("--no-outline", action="store_true",
                        help="禁用行分组/大纲")
    parser.add_argument("--no-write-blank", action="store_true",
                        help="不对空单元格写入格式（减少 XML 体积）")
    args = parser.parse_args()

    merger = BOQMerger(args.source)
    out = merger.merge(
        output_path=args.output,
        keep_source_sheets=args.keep_source_sheets,
        use_outline=not args.no_outline,
        use_write_blank=not args.no_write_blank,
    )
    print(f"Merged → {out}")


if __name__ == "__main__":
    main()
