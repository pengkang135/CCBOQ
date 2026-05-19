"""
分包询价 BOQ 清单拆分工具

从设计院 BOQ 清单中按专业分包范围提取对应 sheet，以基础模板为骨架组装独立询价清单。
- 以 combine.xlsx 模板为骨架（保留 B101-B105、A_Prelims 等通用 sheet）
- 从各设计院清单中按 --match 模式匹配目标 sheet
- 完整逐行复制（值 + 字体 + 边框 + 填充 + 数字格式 + 对齐 + 列宽 + 行高 + 合并单元格）
- 自动修复 FHDI 等设计院的表头 #REF! 错误
- 输出端保持 openpyxl 原生格式，不破坏源表样式

用法:
  python split_inquiry_boq.py \
      --template "combine.xlsx" \
      --source "WTCC.xlsx" --label WTCC \
      --source "FHDI.xlsx" --label FHDI --fix-ref \
      --match "E_Quay" \
      --output "output.xlsx"
"""

import argparse
import os
import re
import sys
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import MergedCell


# ── FHDI 表头 #REF! 修复映射 ──────────────────────────────

FHDI_HEADER_FIX = {
    (1, 1): "Project:",
    (1, 2): "LALDIA CONTAINER TERMINAL",
    (3, 1): "Contract:",
    (3, 2): "MAIN CONSTRUCTION WORKS",
    (4, 1): "Subject:",
    (4, 2): "Part II - Volume B - Schedules",
    (5, 2): "Item B105 - Schedule of Prices & Unit Rate Breakdowns",
}


# ── 核心函数 ──────────────────────────────────────────────

def copy_sheet(src_ws, dst_wb, new_name):
    """完整复制 worksheet（值 + 全部样式 + 列宽行高 + 合并单元格）"""
    dst_ws = dst_wb.create_sheet(title=new_name)

    # 合并单元格
    for mc in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(mc))

    # 列宽
    for col_letter, col_dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = col_dim.width
        dst_ws.column_dimensions[col_letter].hidden = col_dim.hidden

    # 行高
    for row_num, row_dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row_num].height = row_dim.height
        dst_ws.row_dimensions[row_num].hidden = row_dim.hidden

    # 单元格（值 + 样式）
    for row in src_ws.iter_rows():
        for src_cell in row:
            if isinstance(src_cell, MergedCell):
                continue
            dst_cell = dst_ws.cell(row=src_cell.row, column=src_cell.column)
            dst_cell.value = src_cell.value
            if src_cell.has_style:
                dst_cell.font = copy(src_cell.font)
                dst_cell.border = copy(src_cell.border)
                dst_cell.fill = copy(src_cell.fill)
                dst_cell.number_format = src_cell.number_format
                dst_cell.alignment = copy(src_cell.alignment)
                dst_cell.protection = copy(src_cell.protection)

    dst_ws.sheet_properties.tabColor = src_ws.sheet_properties.tabColor
    return dst_ws


def fix_fhdi_header(ws, fix_map=None):
    """修复 FHDI 表头 #REF! 错误"""
    if fix_map is None:
        fix_map = FHDI_HEADER_FIX
    fixed = 0
    for (row, col), val in fix_map.items():
        cell = ws.cell(row=row, column=col)
        if cell.value is not None and "#REF!" in str(cell.value):
            cell.value = val
            fixed += 1
    return fixed


def count_ref_errors(wb):
    """统计工作簿中所有 #REF! 错误"""
    total = 0
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                if cell.value is not None and isinstance(cell.value, str) and "#REF!" in cell.value:
                    total += 1
    return total


def extract_option(src_sheet_name, match_pattern):
    """从源 sheet 名称中提取方案标识。

    WTCC: "E_Quay(OPT. 1 - STEEL PILES)" → "OPT1"
    FHDI: "FHDI_E_Quay（方案一）" → "方案一"
    """
    name = src_sheet_name.strip()
    # Remove common institute prefixes if present
    for prefix in ["FHDI_", "SHCC_", "WTCC_"]:
        if name.startswith(prefix):
            name = name[len(prefix):]

    # Remove the class part (match_pattern)
    if name.startswith(match_pattern):
        name = name[len(match_pattern):]

    # Clean up separators
    name = name.strip("() （）-_ ")

    # Shorten long descriptions
    if "STEEL" in name.upper() and "PILE" in name.upper():
        name = "OPT1"
    elif "PHC" in name.upper():
        name = "OPT2"

    return name if name else "F1"


def find_matched_sheets(wb, pattern):
    """在 workbook 中查找名称包含 pattern 的 sheet，返回排序列表"""
    matched = [s for s in wb.sheetnames if pattern in s]
    return sorted(matched)


def remove_blank_sheet(wb):
    """删除 openpyxl 默认创建的空白 Sheet"""
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]


# ── 主流程 ──────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="分包询价 BOQ 清单拆分 — 从设计院清单提取目标分部，组合到基础模板"
    )
    parser.add_argument("--template", required=True, help="基础模板 xlsx（含 B101-B105, A_Prelims 等通用 sheet）")
    parser.add_argument("--source", action="append", default=[], dest="sources",
                        help="设计院 BOQ 源文件路径（可重复指定）")
    parser.add_argument("--label", action="append", default=[], dest="labels",
                        help="设计院标签，与 --source 一一对应（如 WTCC/FHDI/SHCC）")
    parser.add_argument("--fix-ref", action="append", default=[], dest="fix_refs",
                        help="对指定 --source 启用表头 #REF! 修复（值为 0-based index）")
    parser.add_argument("--match", required=True, help="sheet 名称匹配模式（如 E_Quay）")
    parser.add_argument("--output", required=True, help="输出 xlsx 路径")
    parser.add_argument("--header-fix", type=int, default=1,
                        help="表头修复模式: 0=不修复, 1=仅 FHDI 表头 (默认)")
    args = parser.parse_args()

    # 校验 --source 和 --label 数量一致
    if len(args.sources) != len(args.labels):
        print("Error: --source 与 --label 数量必须一致", file=sys.stderr)
        sys.exit(1)

    # 解析 --fix-ref: 支持 "0,1" 或 "0"
    fix_ref_indices = set()
    if args.fix_refs:
        for item in args.fix_refs:
            for part in item.split(","):
                fix_ref_indices.add(int(part.strip()))

    print("=== BOQ Inquiry Package Split ===")
    print()

    # 加载模板
    print(f"Loading template: {args.template}")
    wb = openpyxl.load_workbook(args.template)
    print(f"  Template sheets: {wb.sheetnames}")

    # 处理每个源文件
    seq_counter = 0  # 全局 sheet 序号

    for idx, (src_path, label) in enumerate(zip(args.sources, args.labels)):
        print(f"\n--- Processing {label} ({src_path}) ---")
        wb_src = openpyxl.load_workbook(src_path)
        matched = find_matched_sheets(wb_src, args.match)
        print(f"  Matched sheets: {matched}")

        if not matched:
            print(f"  WARNING: No sheets matching '{args.match}' found in {label}")

        for name in matched:
            seq_counter += 1
            src_ws = wb_src[name]
            option = extract_option(name, args.match)
            new_name = f"{args.match}({seq_counter}.{label}-{option})"
            copy_sheet(src_ws, wb, new_name)

            # 修复 FHDI 表头 #REF!
            if idx in fix_ref_indices:
                n = fix_fhdi_header(wb[new_name])
                if n:
                    print(f"    Fixed {n} #REF! in header")

        wb_src.close()

    remove_blank_sheet(wb)

    ref_count = count_ref_errors(wb)
    print(f"\n#REF! errors in output: {ref_count}")
    print(f"Final sheets: {wb.sheetnames}")

    os.makedirs(os.path.dirname(args.output), exist_ok=True)
    wb.save(args.output)
    print(f"\nSaved: {args.output}")
    print("Done!")


if __name__ == "__main__":
    main()
