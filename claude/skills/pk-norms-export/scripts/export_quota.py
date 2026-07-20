
"""
Export quota items from SQLite to 3.3土建定额 Excel format.
Column layout: A-BQ (69 cols), chapter > section > items.
Formulas in K(单价)=L+M+N, L(人工)=labour_qty*price, M(材料)=SUMPRODUCT, N(机械)=SUMPRODUCT.
"""
import json, sys, sqlite3
from pathlib import Path

_PARENT_SKILL = Path(__file__).resolve().parent.parent.parent / "pk-norms-import" / "scripts"
if str(_PARENT_SKILL) not in sys.path:
    sys.path.insert(0, str(_PARENT_SKILL))
from db_clean import clean_unit, extract_unit_from_work_content

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# Load subsection grouping config
_GROUPS_PATH = Path(__file__).resolve().parent.parent / "config" / "subsection_groups.json"
_SUBSECTION_GROUPS = {}
if _GROUPS_PATH.exists():
    with open(_GROUPS_PATH, "r", encoding="utf-8") as f:
        _SUBSECTION_GROUPS = json.load(f)

DB = Path(r"F:\BaiduSyncdisk\2.清单定额\Norms-AI\output\db\norms_jts276-1-2019_excel.sqlite")
N_CONSUMPTION = 55
MAX_COL = 14 + N_CONSUMPTION

EN_HEADERS = ["code", "item_name", "unit", "description", "type", "content", "rule",
              "english_name", "amount", "quantity", "rate", "labour", "materials", "mech"]
CN_HEADERS = ["定额编号", "项目名称", "单位", "项目特征", "分类指标", "工作内容", "计算规则",
              "英文名称", "成本合价", "工程量", "单价", "人工", "材料", "机械"]

CHAPTER_META = {
    "1": ("第一章 土石方工程", "Earthwork & Stonework"),
    "2": ("第二章 基础工程", "Foundation Engineering"),
    "3": ("第三章 混凝土及钢筋混凝土构件预制安装工程", "Precast Concrete Works"),
    "4": ("第四章 现浇混凝土及钢筋混凝土工程", "Cast-in-situ Concrete Works"),
    "5": ("第五章 钢结构制作及安装工程", "Steel Structure Works"),
    "6": ("第六章 其他工程", "Miscellaneous Works"),
}

# Section (节) assignment for chapters without L2 hierarchy or needing reassignment.
# Key: chapter digit. Value: {subsection_number_prefix: L2_section_title}.
# Subsection number is the leading Chinese/Arabic number in the subsection name.
SECTION_ASSIGN = {
    "3": {
        "3": "第三节 钢筋工程",
    },
    "5": {
        "三": "第一节 钢结构制作",
        "十四": "第二节 钢结构防腐",
        "十五": "第二节 钢结构防腐",
        "十六": "第二节 钢结构防腐",
    },
    "6": {
        "十六": "第一节 道路与堆场工程",
    },
}


def _extract_sub_num(name):
    """Extract leading number from subsection name: '三、钢管桩制作' → '三', '3. 夯坑' → '3'."""
    import re
    if not name: return None
    m = re.match(r'^([一二三四五六七八九十]+)[、,\s]', name)
    if m: return m.group(1)
    m = re.match(r'^(\d+)\s*[.、]', name)
    if m: return m.group(1)
    return None


def classify(cost_item):
    if cost_item == "人工": return "labour"
    if cost_item == "潜水组": return "labour"
    if cost_item in ("其他材料",): return "materials"
    if cost_item in ("其他船机",): return "mech"
    for kw in ("机", "船", "车", "泵", "钻", "搅拌站", "搅拌船", "搅拌机",
               "起重机", "装载机", "推土机", "压路机", "打桩", "挖掘机",
               "自卸", "拖轮", "驳", "发电机组", "空压机", "电焊机", "对焊机",
               "打桩门架", "潜孔钻", "抓斗"):
        if kw in cost_item: return "mech"
    return "materials"


def _build_quota_name(section_title, attrs):
    """Build descriptive quota name: section_title + attr_label: attr_value pairs."""
    import re
    sec_name = (section_title or "").strip()
    # Strip leading number/roman numeral prefix: "八、" or "3.  " or "Ⅰ  "
    sec_name = re.sub(r'^[一-鿿\dⅠⅡⅢⅣⅤⅥⅦⅧⅨⅩ]+[、.\s]*', '', sec_name).strip()

    parts = [sec_name] if sec_name else []
    for value, label in attrs:
        if not value or not value.strip(): continue
        v = value.strip()
        lbl = (label or "").strip()
        if lbl and lbl != v:
            parts.append(f"{lbl}: {v}")
        else:
            parts.append(v)
    return " ".join(parts)


# clean_unit and extract_unit_from_work_content imported from pk-norms-import.scripts.db_clean


def parse_unit_factor(unit_str):
    """Parse unit factor and base unit. e.g. '100m³' → (100, 'm³'), '10根' → (10, '根'), 'm³' → (1, 'm³')."""
    import re
    if not unit_str: return 1, unit_str
    u = unit_str.strip()
    m = re.match(r'^([\d.]+)\s*(\D.*)$', u)
    if m:
        return float(m.group(1)), m.group(2).strip()
    return 1, u


_CN_NUMS = {"1": "一", "2": "二", "3": "三", "4": "四", "5": "五", "6": "六"}

def _get_merged_group(chapter_name, l2_section, subsection_name):
    """Look up merged group name for a subsection. Returns group name or original subsection name."""
    if not _SUBSECTION_GROUPS or not subsection_name:
        return subsection_name
    for key, val in _SUBSECTION_GROUPS.items():
        key_parts = key.split("|", 1)
        if len(key_parts) != 2:
            continue
        key_ch, key_sec = key_parts
        # Chapter match: try both Arabic (第1章) and Chinese (第一章) forms
        cn = _CN_NUMS.get(key_ch, key_ch)
        ch_match = (chapter_name or "").startswith(f"第{key_ch}章") or \
                   (chapter_name or "").startswith(f"第{cn}章")
        sec_match = key_sec in (l2_section or "")
        if ch_match and sec_match:
            for group in val.get("groups", []):
                for member in group["members"]:
                    if subsection_name.startswith(member):
                        return group["name"]
    return subsection_name


def lookup_quota(db_path, code):
    conn = sqlite3.connect(str(db_path))
    meta = conn.execute("""
        SELECT work_item, attr_level1, attr_level2, attr_level3, attr_level4,
               attr1_label, attr2_label, attr3_label, attr4_label,
               section_title, l2_section, chapter_title, work_content, table_unit
        FROM v_quota_name WHERE quota_code = ? LIMIT 1
    """, (code,)).fetchone()
    if not meta: conn.close(); return None

    ch_key = code[0] if code and code[0] in CHAPTER_META else "6"
    ch_cn, ch_en = CHAPTER_META[ch_key]

    # Build descriptive quota name
    section_title = meta[9] or ""
    attrs = [
        (meta[1], meta[5]),   # attr_level1, attr1_label
        (meta[2], meta[6]),   # attr_level2, attr2_label
        (meta[3], meta[7]),   # attr_level3, attr3_label
        (meta[4], meta[8]),   # attr_level4, attr4_label
    ]
    quota_name = _build_quota_name(section_title, attrs)

    # Unit: prefer table_unit, fallback to work_content extraction
    table_unit = meta[13] or ""
    if not table_unit:
        table_unit = extract_unit_from_work_content(meta[12] or "")

    sub_items = conn.execute(
        "SELECT cost_item, cost_item_unit, amount FROM norms_item WHERE norms_code = ? ORDER BY sort_order",
        (code,)).fetchall()

    resources = []
    for r in sub_items:
        name = (r[0] or "").strip()
        if not name or name == "基价": continue
        unit = (r[1] or "").strip()
        cls = classify(name)
        resources.append({"name": name, "unit": unit, "amount": r[2], "cls": cls,
                          "label": f"{name}\n{unit}" if unit else name})

    l2_sec = meta[10] or ""
    conn.close()
    return {
        "code": code, "chapter_cn": ch_cn, "chapter_en": ch_en,
        "l2_section": l2_sec,
        "subsection": section_title,
        "work_item": meta[0] or "",
        "quota_name": quota_name,
        "work_content": meta[12] or "详见定额原文",
        "table_unit": table_unit,
        "resources": resources,
    }


def _cn_num(s):
    """Convert Chinese number string to int. Supports 一~十, 二十一 etc."""
    if not s: return 99
    cn = {'一':1,'二':2,'三':3,'四':4,'五':5,'六':6,'七':7,'八':8,'九':9,'十':10}
    s = s.strip()
    if s in cn: return cn[s]
    if s.startswith('十'): return 10 + cn.get(s[1:], 0) if len(s) > 1 else 10
    if '十' in s:
        parts = s.split('十')
        tens = cn.get(parts[0], 0)
        ones = cn.get(parts[1], 0) if len(parts) > 1 else 0
        return tens * 10 + ones
    return cn.get(s, 99)

def sec_sort_key(name):
    import re
    m = re.match(r'^(第([一二三四五六七八九十]+))节', name)
    if m: return _cn_num(m.group(2))
    m2 = re.match(r'^([一二三四五六七八九十]+)、', name)
    if m2: return _cn_num(m2.group(1))
    # Roman numeral subsection headers like "Ⅰ  构件预制工程"
    m3 = re.match(r'^[ⅠⅡⅢⅣⅤⅥⅦⅧⅨⅩ]+', name)
    if m3:
        roman = {'Ⅰ':1,'Ⅱ':2,'Ⅲ':3,'Ⅳ':4,'Ⅴ':5,'Ⅵ':6,'Ⅶ':7,'Ⅷ':8,'Ⅸ':9,'Ⅹ':10}
        return roman.get(m3.group(0)[0], 50)
    if '附录' in name: return 90
    return 50

def _sub_sort_key(name):
    """Sort key for subsection (分项工程) names like '三、人力挖岸坡土方' or '3.  钢筋笼加工'."""
    import re
    # Chinese number prefix: "三、" or "二十一、"
    m = re.match(r'^([一二三四五六七八九十]+)、', name)
    if m: return _cn_num(m.group(1))
    # Arabic number prefix: "3.  " or "3.   "
    m = re.match(r'^(\d+)\s*[.、]', name)
    if m: return int(m.group(1))
    # Roman numeral prefix
    m = re.match(r'^[ⅠⅡⅢⅣⅤⅥⅦⅧⅨⅩ]+', name)
    if m:
        roman = {'Ⅰ':1,'Ⅱ':2,'Ⅲ':3,'Ⅳ':4,'Ⅴ':5,'Ⅵ':6,'Ⅶ':7,'Ⅷ':8,'Ⅸ':9,'Ⅹ':10}
        return roman.get(m.group(0)[0], 50)
    return 50

def build_workbook(quota_codes, db_path=DB, normalize_unit=False):
    # ── Build 3-level hierarchy: chapter → L2 section → subsection → items ──
    chapters = {}
    for entry in quota_codes:
        if isinstance(entry, str):
            info = lookup_quota(db_path, entry)
        else:
            info = lookup_quota(db_path, entry["code"])
            if info and "chapter" in entry: info["chapter_cn"] = entry["chapter"]
            if info and "section" in entry: info["subsection"] = entry["section"]
        if not info: continue

        ch = info["chapter_cn"]
        l2 = info["l2_section"] or ch
        sub = info["subsection"] or ""

        # Apply subsection grouping BEFORE section overrides (uses original DB l2)
        merged_sub = _get_merged_group(ch, l2, sub)

        # Apply section assignment overrides
        ch_key = info["code"][0]
        if ch_key in SECTION_ASSIGN:
            sub_num = _extract_sub_num(sub)
            if sub_num and sub_num in SECTION_ASSIGN[ch_key]:
                l2 = SECTION_ASSIGN[ch_key][sub_num]

        if ch not in chapters:
            chapters[ch] = {"meta": info, "l2_sections": {}}
        if l2 not in chapters[ch]["l2_sections"]:
            chapters[ch]["l2_sections"][l2] = {"meta": info, "subsections": {}}
        if merged_sub not in chapters[ch]["l2_sections"][l2]["subsections"]:
            chapters[ch]["l2_sections"][l2]["subsections"][merged_sub] = {"meta": info, "items": [], "all_resources": []}

        ss = chapters[ch]["l2_sections"][l2]["subsections"][merged_sub]
        ss["items"].append(info)
        existing = {r["label"] for r in ss["all_resources"]}
        for res in info["resources"]:
            if res["label"] not in existing:
                existing.add(res["label"])
                ss["all_resources"].append({"label": res["label"], "cls": res["cls"]})

    # Sort resources per subsection: labour -> materials -> mech
    cls_order = {"labour": 0, "materials": 1, "mech": 2}
    for ch_data in chapters.values():
        for l2_data in ch_data["l2_sections"].values():
            for sub_data in l2_data["subsections"].values():
                sub_data["all_resources"].sort(key=lambda r: cls_order.get(r["cls"], 9))

    wb = Workbook()
    ws = wb.active
    ws.title = "沿海港口水工建筑工程定额"

    thin = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    hdr_fill = PatternFill("solid", fgColor="D9E1F2")
    ch_fill = PatternFill("solid", fgColor="B4C6E7")
    l2_fill = PatternFill("solid", fgColor="BDD7EE")
    sub_fill = PatternFill("solid", fgColor="DCE6F1")

    # Row 1: English headers
    for ci, h in enumerate(EN_HEADERS, 1):
        c = ws.cell(row=1, column=ci, value=h); c.font = Font(bold=True, size=9); c.fill = hdr_fill; c.border = thin
    for ci in range(15, MAX_COL + 1):
        c = ws.cell(row=1, column=ci, value=f"qty{ci - 14}"); c.font = Font(bold=True, size=8); c.fill = hdr_fill; c.border = thin

    # Row 2: Chinese headers
    for ci, h in enumerate(CN_HEADERS, 1):
        c = ws.cell(row=2, column=ci, value=h); c.font = Font(bold=True, size=9); c.fill = hdr_fill; c.border = thin
        c.alignment = Alignment(wrap_text=True, vertical="center")
    for ci in range(15, MAX_COL + 1):
        c = ws.cell(row=2, column=ci, value=f"消耗量{ci - 14}"); c.font = Font(bold=True, size=8); c.fill = hdr_fill; c.border = thin

    row = 3
    ch_order = {"第一章 土石方工程": 1, "第二章 基础工程": 2,
                "第三章 混凝土及钢筋混凝土构件预制安装工程": 3,
                "第四章 现浇混凝土及钢筋混凝土工程": 4,
                "第五章 钢结构制作及安装工程": 5, "第六章 其他工程": 6}
    ordered_chapters = sorted(chapters.items(), key=lambda x: ch_order.get(x[0], 99))

    for ch_name, ch_data in ordered_chapters:
        # ── Chapter header ──
        ws.cell(row=row, column=1, value=ch_name[:3]).font = Font(bold=True, size=10)
        ws.cell(row=row, column=2, value=f"【{ch_name.split(maxsplit=1)[-1]}】").font = Font(bold=True, size=10)
        ws.cell(row=row, column=8, value=f"【{ch_data['meta']['chapter_en']}】").font = Font(bold=True, size=9)
        for ci in range(1, MAX_COL + 1):
            ws.cell(row=row, column=ci).fill = ch_fill; ws.cell(row=row, column=ci).border = thin
        row += 1

        ordered_l2 = sorted(ch_data["l2_sections"].items(),
                            key=lambda x: sec_sort_key(x[0]))
        for l2_name, l2_data in ordered_l2:
            has_l2 = bool(l2_data["meta"]["l2_section"])

            # ── L2 Section header (节) ──
            if has_l2:
                ws.cell(row=row, column=2, value=f"《{l2_name}》").font = Font(bold=True, size=9)
                for ci in range(1, MAX_COL + 1):
                    ws.cell(row=row, column=ci).fill = l2_fill; ws.cell(row=row, column=ci).border = thin
                row += 1

            ordered_subs = sorted(l2_data["subsections"].items(),
                                  key=lambda x: _sub_sort_key(x[0]))
            for sub_name, sub_data in ordered_subs:
                resources = sub_data["all_resources"]
                n_res = len(resources)

                labour_cols = []; material_cols = []; mech_cols = []
                for ri, res in enumerate(resources):
                    col = 15 + ri
                    if res["cls"] == "labour": labour_cols.append(col)
                    elif res["cls"] == "materials": material_cols.append(col)
                    elif res["cls"] == "mech": mech_cols.append(col)

                # ── Subsection header (分项工程) + resource labels ──
                if sub_name:
                    ws.cell(row=row, column=2, value=f"{{{sub_name}}}").font = Font(bold=True, size=9)
                for ri, res in enumerate(resources):
                    col = 15 + ri
                    if col > MAX_COL: break
                    ws.cell(row=row, column=col, value=res["label"])
                    ws.cell(row=row, column=col).font = Font(size=7)
                    ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="center")
                for ci in range(1, MAX_COL + 1):
                    ws.cell(row=row, column=ci).fill = sub_fill; ws.cell(row=row, column=ci).border = thin
                    if ci <= 14: ws.cell(row=row, column=ci).font = Font(size=9)
                row += 1

                # ── Price row ──
                price_row = row
                ws.cell(row=row, column=2, value="参考单价 ->").font = Font(size=9, color="808080")
                for ri in range(n_res):
                    col = 15 + ri
                    if col > MAX_COL: break
                    ws.cell(row=row, column=col).number_format = "#,##0.00"
                for ci in range(1, MAX_COL + 1):
                    ws.cell(row=row, column=ci).border = thin
                    ws.cell(row=row, column=ci).font = Font(size=9, color="808080")
                row += 1

                # ── Data rows ──
                for item in sub_data["items"]:
                    data_row = row
                    ws.cell(row=row, column=1, value=item["code"]).number_format = "@"
                    ws.cell(row=row, column=2, value=item["quota_name"])

                    # Unit with optional normalization
                    raw_unit = item.get("table_unit", "")
                    unit_factor, base_unit = parse_unit_factor(raw_unit)
                    display_unit = base_unit if normalize_unit else raw_unit
                    display_unit = display_unit.replace("m3", "m³").replace("m2", "m²")
                    ws.cell(row=row, column=3, value=display_unit)

                    ws.cell(row=row, column=5, value=item["subsection"])
                    ws.cell(row=row, column=6, value=item.get("work_content", "详见定额原文"))
                    ws.cell(row=row, column=7, value="按设计图示尺寸计算")

                    res_map = {r["label"]: r["amount"] for r in item["resources"]}
                    for ri, res in enumerate(resources):
                        col = 15 + ri
                        if col > MAX_COL: break
                        val = res_map.get(res["label"])
                        if val is not None:
                            if normalize_unit and unit_factor != 1:
                                ws.cell(row=row, column=col, value=f"={val}/{unit_factor}")
                            else:
                                ws.cell(row=row, column=col, value=val)
                            ws.cell(row=row, column=col).number_format = "#,##0.0000"

                    # K (col 11) = L + M + N
                    ws.cell(row=row, column=11).value = f"=L{data_row}+M{data_row}+N{data_row}"
                    ws.cell(row=row, column=11).number_format = "#,##0.00"

                    # L (col 12) = SUMPRODUCT(labour prices, labour qty)
                    if labour_cols:
                        if len(labour_cols) == 1:
                            lc = labour_cols[0]; cl = get_column_letter(lc)
                            ws.cell(row=row, column=12).value = f"={cl}{price_row}*{cl}{data_row}"
                        else:
                            pl = f"{get_column_letter(labour_cols[0])}{price_row}:{get_column_letter(labour_cols[-1])}{price_row}"
                            ql = f"{get_column_letter(labour_cols[0])}{data_row}:{get_column_letter(labour_cols[-1])}{data_row}"
                            ws.cell(row=row, column=12).value = f"=SUMPRODUCT({pl},{ql})"
                    ws.cell(row=row, column=12).number_format = "#,##0.00"

                    # M (col 13) = SUMPRODUCT(material prices, material qty)
                    if material_cols:
                        if len(material_cols) == 1:
                            mc = material_cols[0]; cl = get_column_letter(mc)
                            ws.cell(row=row, column=13).value = f"={cl}{price_row}*{cl}{data_row}"
                        else:
                            pm = f"{get_column_letter(material_cols[0])}{price_row}:{get_column_letter(material_cols[-1])}{price_row}"
                            qm = f"{get_column_letter(material_cols[0])}{data_row}:{get_column_letter(material_cols[-1])}{data_row}"
                            ws.cell(row=row, column=13).value = f"=SUMPRODUCT({pm},{qm})"
                    ws.cell(row=row, column=13).number_format = "#,##0.00"

                    # N (col 14) = SUMPRODUCT(mech prices, mech qty)
                    if mech_cols:
                        if len(mech_cols) == 1:
                            nc = mech_cols[0]; cl = get_column_letter(nc)
                            ws.cell(row=row, column=14).value = f"={cl}{price_row}*{cl}{data_row}"
                        else:
                            pn = f"{get_column_letter(mech_cols[0])}{price_row}:{get_column_letter(mech_cols[-1])}{price_row}"
                            qn = f"{get_column_letter(mech_cols[0])}{data_row}:{get_column_letter(mech_cols[-1])}{data_row}"
                            ws.cell(row=row, column=14).value = f"=SUMPRODUCT({pn},{qn})"
                    ws.cell(row=row, column=14).number_format = "#,##0.00"

                    for ci in range(1, MAX_COL + 1):
                        ws.cell(row=row, column=ci).border = thin
                        ws.cell(row=row, column=ci).font = Font(size=9)
                    row += 1
                row += 1  # blank row between subsections
            if has_l2:
                row += 1  # extra blank row between L2 sections
        row += 1  # extra blank row between chapters

    # Column widths
    for cl, w in {"A": 12, "B": 36, "C": 10, "D": 14, "E": 14, "F": 18, "G": 16, "H": 18}.items():
        ws.column_dimensions[cl].width = w
    for ci in range(9, 15):
        ws.column_dimensions[get_column_letter(ci)].width = 12
    for ci in range(15, MAX_COL + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 13
    ws.freeze_panes = "A4"

    # ── Price table sheet ──
    # Collect all unique resources with total quantities (normalize names and units)
    def _norm_unit(u):
        if not u: return u
        u = u.strip()
        u = u.replace("m3", "m³").replace("m2", "m²")
        return u

    def _norm_name(n):
        if not n: return n
        return " ".join(n.strip().split())

    all_res = {}
    for ch_name, ch_data in chapters.items():
        for l2_name, l2_data in ch_data["l2_sections"].items():
            for sub_name, sub_data in l2_data["subsections"].items():
                for item in sub_data["items"]:
                    for res in item["resources"]:
                        norm_name = _norm_name(res["name"])
                        norm_unit = _norm_unit(res["unit"])
                        key = f"{norm_name}\n{norm_unit}"
                        if key not in all_res:
                            all_res[key] = {"name": norm_name, "unit": norm_unit or res["unit"],
                                            "cls": res["cls"], "total_qty": 0, "chapters": set()}
                        all_res[key]["total_qty"] += (res["amount"] or 0)
                        all_res[key]["chapters"].add(ch_name[:3])

    ps = wb.create_sheet("资源价格表")
    # Headers
    price_headers = ["序号", "类型", "资源名称", "单位", "参考数量", "除税单价", "含税单价", "适用章节", "备注"]
    price_widths = {"A": 8, "B": 10, "C": 32, "D": 10, "E": 16, "F": 14, "G": 14, "H": 28, "I": 16}
    for ci, h in enumerate(price_headers, 1):
        cl = get_column_letter(ci)
        c = ps.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, size=9)
        c.fill = hdr_fill
        c.border = thin
        c.alignment = Alignment(horizontal="center")
        ps.column_dimensions[cl].width = price_widths.get(cl, 12)

    ps.freeze_panes = "A2"

    cls_groups = [
        ("labour", "人工", None),
        ("materials", "材料", None),
        ("mech", "机械", None),
    ]
    # Sub-groups for materials (order matters: more specific first)
    mat_subgroups = [
        ("混凝土/砂浆", ["混凝土", "砂浆"]),
        ("水泥", ["水泥"]),
        ("钢材", ["钢筋", "钢板", "钢管", "型钢", "工字钢", "槽钢", "角钢", "钢轨", "钢丝", "铁件"]),
        ("木材", ["板枋材", "圆木", "原木", "枕木", "板材"]),
        ("砂石料", ["砂", "石", "碎石", "块石", "片石", "砾石", "卵石", "毛石", "料石"]),
        ("土工材料", ["土工", "塑料", "泡沫", "橡胶", "PVC", "尼龙"]),
        ("燃料/油料", ["柴油", "汽油", "煤", "沥青"]),
        ("炸药/爆破", ["炸药", "雷管", "导火线", "导爆"]),
        ("其他材料", []),
    ]

    grouped = {"labour": [], "materials": [], "mech": []}
    ungrouped_materials = list(all_res.items())

    for cls, cls_name, _ in cls_groups:
        if cls == "materials":
            mat_by_sub = {sg[0]: [] for sg in mat_subgroups}
            mat_by_sub["其他材料"] = []
            remaining = []
            for key, data in all_res.items():
                if data["cls"] != "materials":
                    continue
                matched = False
                for sg_name, kws in mat_subgroups:
                    if not kws:
                        continue
                    for kw in kws:
                        if kw in data["name"]:
                            mat_by_sub[sg_name].append((key, data))
                            matched = True
                            break
                    if matched:
                        break
                if not matched:
                    mat_by_sub["其他材料"].append((key, data))
            for sg_name, _ in mat_subgroups:
                sg_items = mat_by_sub.get(sg_name, [])
                if sg_items:
                    sg_items.sort(key=lambda x: x[1]["name"])
                    grouped["materials"].append((sg_name, sg_items))
        else:
            items = [(k, v) for k, v in all_res.items() if v["cls"] == cls]
            items.sort(key=lambda x: x[1]["name"])
            if items:
                grouped[cls].append((None, items))

    price_row = 2
    seq = 1
    cls_names = {"labour": "人工", "materials": "材料", "mech": "机械"}

    for cls, cls_name, _ in cls_groups:
        sections = grouped[cls]
        if not sections:
            continue

        # Class header row
        for ci in range(1, len(price_headers) + 1):
            ps.cell(row=price_row, column=ci).fill = l2_fill
            ps.cell(row=price_row, column=ci).border = thin
            ps.cell(row=price_row, column=ci).font = Font(size=9)
        ps.cell(row=price_row, column=2, value=f"【{cls_names[cls]}】").font = Font(bold=True, size=10)
        price_row += 1

        for sg_name, items in sections:
            if sg_name:
                # Sub-group header
                for ci in range(1, len(price_headers) + 1):
                    ps.cell(row=price_row, column=ci).fill = sub_fill
                    ps.cell(row=price_row, column=ci).border = thin
                    ps.cell(row=price_row, column=ci).font = Font(size=9)
                ps.cell(row=price_row, column=3, value=f"{{{sg_name}}}").font = Font(bold=True, size=9)
                price_row += 1

            for key, data in items:
                ps.cell(row=price_row, column=1, value=seq).number_format = "0"
                ps.cell(row=price_row, column=2, value=cls_names[cls])
                ps.cell(row=price_row, column=3, value=data["name"])
                ps.cell(row=price_row, column=4, value=data["unit"])
                ps.cell(row=price_row, column=5, value=round(data["total_qty"], 4)).number_format = "#,##0.0000"
                ps.cell(row=price_row, column=8, value=", ".join(sorted(data["chapters"])))
                for ci in range(1, len(price_headers) + 1):
                    ps.cell(row=price_row, column=ci).border = thin
                    ps.cell(row=price_row, column=ci).font = Font(size=9)
                price_row += 1
                seq += 1

        price_row += 1  # blank row between classes

    return wb


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Export port quota items to Excel")
    parser.add_argument("input", nargs="?", help="JSON file with quota codes")
    parser.add_argument("-o", "--output", default=None, help="Output xlsx path")
    parser.add_argument("--stdin", action="store_true", help="Read JSON from stdin")
    parser.add_argument("--normalize-unit", action="store_true",
                        help="Normalize quota units to base unit (100m³→m³), divide amounts by factor")
    args = parser.parse_args()

    if args.stdin:
        data = json.load(sys.stdin)
    elif args.input:
        with open(args.input, "r", encoding="utf-8") as f:
            data = json.load(f)
    else:
        print("Usage: python export_quota.py <codes.json> [-o output.xlsx]")
        sys.exit(1)

    out_path = args.output or "港口定额_导出.xlsx"
    wb = build_workbook(data, normalize_unit=args.normalize_unit)
    wb.save(out_path)
    print(f"Exported: {out_path}")
    print(f"Items: {len(data)}")
