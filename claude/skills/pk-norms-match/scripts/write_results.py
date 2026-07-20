"""
Write matching results back to Excel with formulas.

Column layout (appended to existing workbook):
  A (col 1): 定额编号
  B (col 2): 定额名称
  L (col 12): =FORMULATEXT(M) — shows the formula used
  M (col 13): =I{row}/{factor} — BOQ qty / quota unit factor
  N (col 14): 1 (default conversion coefficient, user-adjustable)
  O (col 15): quota unit string (e.g. "100m3")
  P (col 16): =PRODUCT(M,N) — final quota quantity

Usage:
    python write_results.py matching.json source.xlsx [-o output.xlsx]
"""

import json, sys, io, argparse, os, shutil
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

try:
    import openpyxl
except ImportError:
    sys.exit('openpyxl is required: pip install openpyxl')


def load_matching_results(path):
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def write_excel(results, source_path, output_path, sheet_name=None, dry_run=False):
    """Write matching results to Excel with formulas.

    Only writes rows that have matched quota codes (match_type='已匹配').
    Preserves existing content, only modifies A, B, L, M, N, O, P columns.
    """
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
    shutil.copy2(source_path, output_path)
    wb = openpyxl.load_workbook(output_path, keep_links=False)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    _TITLE_MARKER_PAIRS = [('{', '}'), ('【', '】'), ('《', '》')]
    _CONCEPTUAL_UNITS = {'ls', 'l.s.', 'lump sum', 'item', 'lot', 'allow',
                         'allowance', '项', 'sum', 'lump'}

    def _is_title_row(name_val):
        if not name_val:
            return False
        s = str(name_val).strip()
        for open_m, close_m in _TITLE_MARKER_PAIRS:
            if open_m in s and close_m in s:
                return True
        return False

    def _is_conceptual_unit(unit_val):
        if not unit_val:
            return False
        return str(unit_val).strip().lower() in _CONCEPTUAL_UNITS

    # Build name→row mapping from the sheet for robust row lookup
    # (fastexcel and openpyxl may disagree on row numbers due to empty rows)
    name_to_row = {}
    for row in range(1, ws.max_row + 1):
        name_val = ws.cell(row=row, column=5).value
        if name_val:
            name_to_row[str(name_val).strip()] = row

    def _resolve_row(result):
        """Resolve the correct Excel row for a result, using name matching."""
        boq_name = (result.get('boq_name') or '').strip()
        if boq_name and boq_name in name_to_row:
            return name_to_row[boq_name]
        # Fallback to result's row number
        return result.get('row', 0)

    written = 0
    skipped = 0
    cleared = 0

    for r in results:
        row = _resolve_row(r)
        if row == 0:
            skipped += 1
            continue

        # Safety net: skip title rows and conceptual units
        col_e = ws.cell(row=row, column=5).value
        col_f = ws.cell(row=row, column=6).value
        if _is_title_row(col_e) or _is_conceptual_unit(col_f):
            for col in [1, 2]:
                ws.cell(row=row, column=col).value = None
            cleared += 1
            continue

        match_type = r.get('match_type', '')
        matches = r.get('matches', [])

        if match_type != '已匹配' or not matches:
            for col in [1, 2]:
                ws.cell(row=row, column=col).value = None
            skipped += 1
            continue

        best = matches[0]

        ws.cell(row=row, column=1).value = best['quota_code']
        ws.cell(row=row, column=2).value = best['quota_name']

        # 暂不填写换算公式和定额工程量，位置待确认
        # ws.cell(row=row, column=15).value = best['quota_unit_raw'] or best.get('phys_unit', '')
        # factor = best.get('factor', 1.0)
        # if factor and factor != 0:
        #     factor_display = int(factor) if factor == int(factor) else factor
        #     m_formula = f'=I{row}/{factor_display}'
        #     ws.cell(row=row, column=13).value = m_formula
        #     ws.cell(row=row, column=12).value = f'=FORMULATEXT(M{row})'
        #     existing_n = ws.cell(row=row, column=14).value
        #     if existing_n is None or existing_n == 0:
        #         ws.cell(row=row, column=14).value = 1
        #     ws.cell(row=row, column=16).value = f'=PRODUCT(M{row},N{row})'

        written += 1

    if dry_run:
        print(f'[DRY RUN] Would write {written} items, skip {skipped} items')
        return

    wb.save(output_path)
    print(f'写入完成: {written} 项已匹配, {skipped} 项未匹配/跳过, {cleared} 项标题行/概念单位已清除')
    print(f'输出文件: {output_path}')


def main():
    parser = argparse.ArgumentParser(description='匹配结果写回Excel (含公式)')
    parser.add_argument('matching_json', help='match_quota.py 输出的匹配JSON')
    parser.add_argument('source_xlsx', help='原始BOQ清单Excel文件')
    parser.add_argument('-o', '--output', help='输出Excel路径 (默认: 源文件_matched.xlsx)')
    parser.add_argument('--sheet', help='目标sheet名 (默认: 第一个sheet)')
    parser.add_argument('--dry-run', action='store_true', help='仅预览，不实际写入')
    args = parser.parse_args()

    results = load_matching_results(args.matching_json)
    print(f'加载匹配结果: {len(results)} 条')

    if args.output:
        out_path = args.output
    else:
        src = Path(args.source_xlsx)
        out_path = str(src.parent / f'{src.stem}_matched{src.suffix}')

    write_excel(results, args.source_xlsx, out_path,
                sheet_name=args.sheet, dry_run=args.dry_run)


if __name__ == '__main__':
    main()
