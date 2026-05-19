import argparse
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


def split_markdown_row(line):
    content = line.strip()
    if not content.startswith("|") or not content.endswith("|"):
        return []
    content = content[1:-1]
    cells = []
    current = []
    escape = False
    for ch in content:
        if escape:
            current.append(ch)
            escape = False
        elif ch == "\\":
            escape = True
        elif ch == "|":
            cells.append("".join(current).strip())
            current = []
        else:
            current.append(ch)
    cells.append("".join(current).strip())
    return cells


def unescape_md(value):
    text = value.replace(r"\|", "|")
    text = text.replace("<br>", "\n")
    return text.strip()


def parse_excel_column(value):
    text = str(value).strip()
    if text.isdigit():
        return int(text)

    text = text.upper()
    if not text.isalpha():
        raise argparse.ArgumentTypeError(
            f"invalid column value: {value!r}; use a zero-based index like 2 or a column letter like C"
        )

    result = 0
    for ch in text:
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result - 1


def import_review_markdown(md_path, excel_path, tgt_col_idx, output_path=None, sheet_name=None):
    markdown_lines = Path(md_path).read_text(encoding="utf-8").splitlines()
    updates = {}

    for line in markdown_lines:
        if not line.startswith("|"):
            continue
        cells = split_markdown_row(line)
        if len(cells) != 6:
            continue
        if cells[0] in {"row", "---"}:
            continue
        try:
            excel_row = int(cells[0])
        except ValueError:
            continue
        updates[excel_row] = unescape_md(cells[3])

    wb = load_workbook(excel_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    for excel_row, text in updates.items():
        ws.cell(row=excel_row, column=tgt_col_idx + 1).value = text

    save_path = output_path if output_path else excel_path
    wb.save(save_path)
    print(save_path)
    print(f"updated_rows={len(updates)}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("md_path")
    parser.add_argument("excel_path")
    parser.add_argument("--tgt-col", type=parse_excel_column, default=2)
    parser.add_argument("--output", type=str, default=None)
    parser.add_argument("--sheet", type=str, default=None)
    args = parser.parse_args()

    import_review_markdown(
        args.md_path,
        args.excel_path,
        tgt_col_idx=args.tgt_col,
        output_path=args.output,
        sheet_name=args.sheet,
    )


if __name__ == "__main__":
    main()
