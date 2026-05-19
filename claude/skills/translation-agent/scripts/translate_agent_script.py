"""
[NOTE] Excel Helsinki-NLP batch translation script (offline model).
Retained for reference. The standard workflow uses Flash -> Pro instead.
"""
import pandas as pd
from transformers import MarianMTModel, MarianTokenizer
import torch
from openpyxl import load_workbook
import os
import sys
import argparse
from tqdm import tqdm
import re
import json
from modelscope.hub.snapshot_download import snapshot_download

def get_device():
    if torch.cuda.is_available():
        return "cuda"
    return "cpu"

def load_model(model_name="Helsinki-NLP/opus-mt-zh-en"):
    # Set local model directory relative to the script
    # Assume script is in .trae/skills/translation-agent/scripts/
    # Models are in e:\Code\FeynmanLibrary\models
    # We need to find the project root relative to this script
    script_dir = os.path.dirname(os.path.abspath(__file__))
    # script_dir is .../scripts
    # skill_dir is .../translation-agent
    # skills_dir is .../.trae/skills
    # trae_dir is .../.trae
    # project_root is .../FeynmanLibrary
    project_root = os.path.abspath(os.path.join(script_dir, "../../../../"))
    local_model_root = os.path.join(project_root, "models")
    
    print(f"Model will be stored in: {local_model_root}")
    
    try:
        model_dir = snapshot_download(model_name, cache_dir=local_model_root)
    except Exception as e:
        print(f"ModelScope download failed: {e}")
        return None, None

    print(f"Loading model from {model_dir}...")
    tokenizer = MarianTokenizer.from_pretrained(model_dir)
    model = MarianMTModel.from_pretrained(model_dir)
    
    device = get_device()
    print(f"Using device: {device}")
    model.to(device)
    
    return tokenizer, model

def translate_excel(file_path, src_col_idx, tgt_col_idx, sheet_name=None, model_name="Helsinki-NLP/opus-mt-zh-en", output_path=None, glossary_path=None):
    if not os.path.exists(file_path):
        print(f"Error: File not found at {file_path}")
        return

    # Load Glossary
    glossary = {}
    if glossary_path and os.path.exists(glossary_path):
        try:
            with open(glossary_path, 'r', encoding='utf-8') as f:
                glossary = json.load(f)
            print(f"Loaded glossary with {len(glossary)} terms.")
        except Exception as e:
            print(f"Error loading glossary: {e}")

    # Load model
    tokenizer, model = load_model(model_name)
    if not tokenizer or not model:
        return
    
    device = get_device()

    print(f"Reading Excel file: {file_path}")
    try:
        wb = load_workbook(file_path)
        if sheet_name:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                print(f"Error: Sheet '{sheet_name}' not found.")
                return
        else:
            ws = wb.active
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return
    
    df = pd.read_excel(file_path, sheet_name=0 if sheet_name is None else sheet_name, engine='openpyxl')
    
    chinese_pattern = re.compile(r'[\u4e00-\u9fff]')
    is_en_to_zh = "en-zh" in model_name.lower()

    rows_to_translate = []
    
    # Load glossary and pre-compile regex
    glossary_patterns = []
    if glossary:
        print("Pre-compiling glossary patterns...")
        sorted_terms = sorted(glossary.keys(), key=len, reverse=True)
        for term in sorted_terms:
            pattern = re.compile(re.escape(term), re.IGNORECASE)
            trans = glossary[term]
            glossary_patterns.append((pattern, trans))
    
    print("Scanning rows...")
    for index, row in df.iterrows():
        if src_col_idx >= len(row) or tgt_col_idx >= len(row):
            continue

        b_val = row.iloc[src_col_idx]
        h_val = row.iloc[tgt_col_idx]
        
        if pd.isna(b_val) or str(b_val).strip() == "":
            continue
            
        b_str = str(b_val).strip()
        h_str = str(h_val).strip() if pd.notna(h_val) else ""
        
        should_translate = False
        if h_str == "":
            should_translate = True
        elif is_en_to_zh:
            if not chinese_pattern.search(h_str):
                should_translate = True
        else:
            if chinese_pattern.search(h_str):
                should_translate = True
        
        if should_translate:
            # Pre-process with glossary for consistency (Simple replacement)
            # This helps the model by providing the correct translation in-context or enforcing it.
            
            temp_b_str = b_str
            for pattern, trans in glossary_patterns:
                if pattern.search(temp_b_str):
                    temp_b_str = pattern.sub(trans, temp_b_str)
            
            rows_to_translate.append((index, b_str, temp_b_str)) # Store original and modified

    total_rows = len(rows_to_translate)
    print(f"Found {total_rows} rows to translate.")
    
    if total_rows == 0:
        return

    batch_size = 8 # Reduced batch size for faster updates
    indices = [r[0] for r in rows_to_translate]
    texts_to_translate = [r[2] for r in rows_to_translate] # Use modified text
    
    print("Starting translation...")
    
    translated_results = []
    
    for i in tqdm(range(0, total_rows, batch_size)):
        batch_texts = texts_to_translate[i:i + batch_size]
        batch_indices = indices[i:i + batch_size]
        
        inputs = tokenizer(batch_texts, return_tensors="pt", padding=True, truncation=True, max_length=512).to(device)
        
        with torch.no_grad():
            # num_beams=1 for greedy (fastest), num_beams=3 for better quality
            translated = model.generate(**inputs, num_beams=1, max_new_tokens=512)
        
        decoded_texts = [tokenizer.decode(t, skip_special_tokens=True) for t in translated]
        
        # Post-process Verification not strictly needed if we did pre-replacement, 
        # but let's just use the decoded text.
        for j, text in enumerate(decoded_texts):
            final_text = text
            # Clean up potential artifacts if needed
            translated_results.append(final_text)

            excel_row = batch_indices[j] + 2
            ws.cell(row=excel_row, column=tgt_col_idx + 1).value = final_text

    print("Saving file...")
    save_path = output_path if output_path else file_path
    try:
        wb.save(save_path)
        print(f"Saved to {save_path}")
    except Exception as e:
        print(f"Error saving: {e}")
        if output_path is None:
            base, ext = os.path.splitext(file_path)
            alt_path = f"{base}_translated{ext}"
            wb.save(alt_path)
            print(f"Saved to alternate: {alt_path}")

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("file_path")
    parser.add_argument("--src-col", type=int, default=1)
    parser.add_argument("--tgt-col", type=int, default=7)
    parser.add_argument("--sheet", type=str, default=None)
    parser.add_argument("--model", type=str, default="Helsinki-NLP/opus-mt-zh-en")
    parser.add_argument("--output", type=str, default=None)
    parser.add_argument("--glossary", type=str, default=None)

    args = parser.parse_args()
    translate_excel(args.file_path, args.src_col, args.tgt_col, args.sheet, args.model, args.output, args.glossary)

if __name__ == "__main__":
    main()
