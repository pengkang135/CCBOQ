#!/usr/bin/env python3
"""
Direct Excel translation for large BOQ / schedule files.
Usage: python translate_direct.py <excel_path> [--src-col N] [--tgt-col N] [--output path] [--clean]

Progressive save every row; auto-saves to a new file if original is locked.
Retries translation with exponential backoff; skips already-translated rows.
--clean strips BOQ hierarchy markers (《》{}【】) before sending to translator.
"""

import json
import openpyxl
import re
import time
import sys
import os
from pathlib import Path
from deep_translator import GoogleTranslator


def clean_boq_text(text):
    """Strip BOQ hierarchy markers before sending to translator.
    Original text with markers is preserved in the Excel cell."""
    if not text:
        return text
    text = text.replace('《', '').replace('》', '')  # 《》
    text = re.sub(r'^\{([^}]*)\}$', r'\1', text.strip())
    text = re.sub(r'^【([^】]*)】$', r'\1', text.strip())  # 【】
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def translate_with_retry(translator, text, max_retries=5):
    for attempt in range(max_retries):
        try:
            return translator.translate(text)
        except Exception:
            if attempt < max_retries - 1:
                time.sleep(2 ** attempt)  # 1s, 2s, 4s, 8s
    return None


def detect_cols(ws):
    """Auto-detect English source and Chinese target columns."""
    header = [c.value for c in ws[1]]
    src_col = tgt_col = None
    for idx, h in enumerate(header, 1):
        if h and ("description" in str(h).lower() or "desc" in str(h).lower()) and not src_col:
            src_col = idx
        if h and ("中文" in str(h) or "translation_zh" in str(h).lower()
                  or "chinese" in str(h).lower() or "zh" in str(h).lower()):
            tgt_col = idx
    # Fallbacks for BOQ-specific headers
    if not src_col:
        for idx, h in enumerate(header, 1):
            if h and ("item" in str(h).lower() or "boq" in str(h).lower()):
                src_col = idx
                break
    if not tgt_col:
        tgt_col = src_col + 1 if src_col else 3
    if not src_col:
        src_col = 2  # default B column
    return src_col, tgt_col


def main():
    if len(sys.argv) < 2:
        print("Usage: python translate_direct.py <excel_path> [--src-col N] [--tgt-col N] [--output path] [--clean]")
        sys.exit(1)

    excel_path = sys.argv[1]
    src_col = None
    tgt_col = None
    output_path = None
    use_clean = False

    # Parse args
    i = 2
    while i < len(sys.argv):
        if sys.argv[i] == "--src-col" and i + 1 < len(sys.argv):
            src_col = int(sys.argv[i + 1])
            i += 2
        elif sys.argv[i] == "--tgt-col" and i + 1 < len(sys.argv):
            tgt_col = int(sys.argv[i + 1])
            i += 2
        elif sys.argv[i] == "--output" and i + 1 < len(sys.argv):
            output_path = sys.argv[i + 1]
            i += 2
        elif sys.argv[i] == "--clean":
            use_clean = True
            i += 1
        else:
            i += 1

    if not Path(excel_path).exists():
        print(f"[错误] 文件不存在: {excel_path}")
        sys.exit(1)

    print("=" * 60)
    print("Direct Excel Translation (BOQ / Large Schedule)")
    print("=" * 60)

    print("\n[1/4] 打开 Excel...")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    if not src_col or not tgt_col:
        src_col, tgt_col = detect_cols(ws)
    print(f"  源列: {src_col}  目标列: {tgt_col}  清洗: {use_clean}")

    # Determine output path (new file to avoid lock issues)
    if not output_path:
        p = Path(excel_path)
        output_path = p.parent / (p.stem + ".translated" + p.suffix)
    print(f"  输出: {output_path}")

    print("\n[2/4] 初始化翻译器...")
    translator = GoogleTranslator(source="en", target="zh-CN")

    # Collect pending rows
    print("\n[3/4] 扫描待翻译行...")
    todo = []
    for row in range(2, ws.max_row + 1):
        en = ws.cell(row, src_col).value
        zh = ws.cell(row, tgt_col).value
        if en and not zh:
            todo.append(row)

    total = len(todo)
    print(f"  待翻译: {total} 行")
    if total == 0:
        print("  ✓ 全部已翻译")
        wb.close()
        return

    print("\n[4/4] 开始翻译...")
    print("-" * 60)
    success = 0
    fail = 0
    failed_rows = []

    try:
        for i, row in enumerate(todo, 1):
            en_text = ws.cell(row, src_col).value
            raw = str(en_text) if en_text else ""
            display = raw[:55]
            print(f"[{i}/{total}] Row {row}: {display}")

            input_text = clean_boq_text(raw) if use_clean else raw
            result = translate_with_retry(translator, input_text)
            if result:
                ws.cell(row, tgt_col).value = result
                success += 1
                print(f"  ✓ {result[:55]}")
            else:
                fail += 1
                failed_rows.append({"row": row, "text": raw})
                print(f"  ✗ 翻译失败")

            # Progressive save every row
            wb.save(output_path)
            time.sleep(1.0)

    except KeyboardInterrupt:
        print("\n! 用户中断，进度已保存")
    except Exception as e:
        print(f"\n! 错误: {e}")
    finally:
        try:
            wb.save(output_path)
        except Exception:
            pass
        wb.close()

    print("-" * 60)
    print(f"完成: 成功 {success} 行, 失败 {fail} 行")
    print(f"输出文件: {output_path}")

    if failed_rows:
        dump_path = str(Path(output_path).with_suffix('')) + '.failed.json'
        with open(dump_path, 'w', encoding='utf-8') as f:
            json.dump(failed_rows, f, ensure_ascii=False, indent=2)
        print(f"失败条目: {dump_path}")


if __name__ == "__main__":
    main()
