"""Restructure a messy xlsx by grouping rows with overlapping content spans.

Algorithm (domain-agnostic, no keyword matching):
  - For each row, compute span = (first non-empty col, last non-empty col).
  - Consecutive rows with overlapping spans form a "group".
  - When a row's span is disjoint from the current group (e.g. section
    title in col A while table uses B-Q), close the group and start new.
  - Empty rows close the current group.
  - Within each group, drop columns that are entirely empty.
  - Stack all groups vertically; each group starts at column A.

Preserves cell styles, row heights, column widths, and merged cells that
fall entirely within a single group.

Usage:
    python stack_by_span.py <input.xlsx> [output.xlsx]

If output path is omitted, defaults to `<input_basename>_stacked.xlsx` in
the same directory as the input.
"""

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from copy import copy
import argparse
import os
import sys


def non_empty_cols(cells):
    """Return set of column indices with non-empty content."""
    return set(c.column for c in cells
               if c.value is not None and str(c.value).strip())


def row_span(cells):
    """Return (first_col, last_col) of non-empty cells; None for empty row."""
    cols = non_empty_cols(cells)
    if not cols:
        return None
    return (min(cols), max(cols))


def spans_overlap(a, b):
    """True if two closed intervals [a0,a1] and [b0,b1] intersect."""
    return not (a[1] < b[0] or b[1] < a[0])


def copy_style(src, dst):
    if not src.has_style:
        return
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.number_format = src.number_format
    dst.protection = copy(src.protection)


def resolve_merges(ws):
    """For each merged cell range, fill the top-left value into every row of
    the range's LEFTMOST column, then unmerge. Convention: a vertical/
    rectangular merge represents 'same value as above' for downstream cells.
    Horizontal expansion is intentionally NOT filled — the visual multi-column
    span is a rendering detail that would confuse column semantics."""
    ranges = list(ws.merged_cells.ranges)
    resolved = 0
    for mr in ranges:
        val = ws.cell(row=mr.min_row, column=mr.min_col).value
        ws.unmerge_cells(str(mr))
        if val is None:
            continue
        for r in range(mr.min_row + 1, mr.max_row + 1):
            ws.cell(row=r, column=mr.min_col, value=val)
        resolved += 1
    return resolved


def restructure(xlsx_in: str, xlsx_out: str) -> dict:
    """Main entrypoint. Returns a stats dict describing the run."""
    wb_in = load_workbook(xlsx_in)
    ws_in = wb_in.active
    merges_resolved = resolve_merges(ws_in)
    max_row = ws_in.max_row
    max_col = ws_in.max_column

    rows = []
    for row in ws_in.iter_rows(min_row=1, max_row=max_row,
                               min_col=1, max_col=max_col):
        rows.append(list(row))

    # Build groups by span overlap, with singleton-outlier break rule:
    # a single-cell row whose column is NOT in the current group's set of
    # "multi-cell columns" (the union of columns used by any multi-cell row
    # in the group) is treated as a boundary marker, not a data continuation.
    groups = []
    current = None
    for i, cells in enumerate(rows):
        cols = non_empty_cols(cells)
        if not cols:
            if current:
                groups.append(current)
                current = None
            continue
        span = (min(cols), max(cols))
        is_singleton = len(cols) == 1

        # Singleton-outlier break: singleton row in a column outside the
        # group's established data footprint.
        if (is_singleton and current is not None and current["multi_cols"]
                and next(iter(cols)) not in current["multi_cols"]):
            groups.append(current)
            current = {"span": span, "rows": [i], "multi_cols": set()}
            continue

        if current is not None and spans_overlap(current["span"], span):
            current["span"] = (min(current["span"][0], span[0]),
                               max(current["span"][1], span[1]))
            current["rows"].append(i)
            if not is_singleton:
                current["multi_cols"] |= cols
        else:
            if current:
                groups.append(current)
            current = {"span": span, "rows": [i], "multi_cols": set()}
            if not is_singleton:
                current["multi_cols"] = set(cols)
    if current:
        groups.append(current)

    # Emit to new workbook
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = ws_in.title

    out_col_widths = {}
    out_r = 1
    total_merges = 0

    for gi, g in enumerate(groups):
        row_indices = g["rows"]
        start, end = row_indices[0], row_indices[-1] + 1

        used_cols = set()
        for i in row_indices:
            for cell in rows[i]:
                v = cell.value
                if v is not None and str(v).strip():
                    used_cols.add(cell.column)
        if not used_cols:
            continue
        used_cols = sorted(used_cols)
        col_map = {orig: new + 1 for new, orig in enumerate(used_cols)}

        for orig_col in used_cols:
            letter = get_column_letter(orig_col)
            w = ws_in.column_dimensions[letter].width
            if w and w > out_col_widths.get(col_map[orig_col], 0):
                out_col_widths[col_map[orig_col]] = w

        group_first_out_row = out_r
        for i in row_indices:
            src_row_1b = i + 1
            src_h = ws_in.row_dimensions[src_row_1b].height
            if src_h:
                ws_out.row_dimensions[out_r].height = src_h
            for cell in rows[i]:
                if cell.column not in col_map or cell.value is None:
                    continue
                new_col = col_map[cell.column]
                dst = ws_out.cell(row=out_r, column=new_col, value=cell.value)
                copy_style(cell, dst)
            out_r += 1

        for mr in ws_in.merged_cells.ranges:
            if not (mr.min_row - 1 >= start and mr.max_row - 1 < end):
                continue
            if mr.min_col not in col_map or mr.max_col not in col_map:
                continue
            new_min_col = col_map[mr.min_col]
            new_max_col = col_map[mr.max_col]
            new_min_row = group_first_out_row + (mr.min_row - 1 - start)
            new_max_row = group_first_out_row + (mr.max_row - 1 - start)
            try:
                ws_out.merge_cells(
                    start_row=new_min_row, end_row=new_max_row,
                    start_column=new_min_col, end_column=new_max_col
                )
                total_merges += 1
            except Exception:
                pass

    for new_col, width in out_col_widths.items():
        ws_out.column_dimensions[get_column_letter(new_col)].width = width

    os.makedirs(os.path.dirname(os.path.abspath(xlsx_out)) or ".", exist_ok=True)
    wb_out.save(xlsx_out)

    return {
        "input": xlsx_in,
        "output": xlsx_out,
        "source_rows": max_row,
        "source_cols": max_col,
        "output_rows": out_r - 1,
        "groups": len(groups),
        "merges_resolved": merges_resolved,
        "merges_preserved": total_merges,
    }


def default_output_path(xlsx_in: str) -> str:
    base, ext = os.path.splitext(xlsx_in)
    return base + "_stacked" + ext


def main():
    ap = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    ap.add_argument("input", help="Path to the source .xlsx file")
    ap.add_argument("output", nargs="?", default=None,
                    help="Output path (default: <input>_stacked.xlsx)")
    args = ap.parse_args()

    xlsx_in = args.input
    xlsx_out = args.output or default_output_path(xlsx_in)

    if not os.path.isfile(xlsx_in):
        print(f"ERROR: input file not found: {xlsx_in}", file=sys.stderr)
        sys.exit(1)

    print(f"Loading: {xlsx_in}")
    stats = restructure(xlsx_in, xlsx_out)
    print(f"Wrote:   {stats['output']}")
    print(f"  source: {stats['source_rows']} rows x {stats['source_cols']} cols")
    print(f"  merges resolved (filled down): {stats['merges_resolved']}")
    print(f"  output: {stats['output_rows']} rows in {stats['groups']} groups")
    print(f"  merges preserved in output: {stats['merges_preserved']}")


if __name__ == "__main__":
    main()
