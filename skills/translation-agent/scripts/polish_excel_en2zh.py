"""
Polish existing Chinese translations in an Excel file (standalone post-processor).
Applies term fixes, symbol cleanup, and repetition removal.
Does NOT use any local translation model — designed for Flash -> Pro workflow.

Optional: load external pitfalls JSON and glossary JSON for domain-specific corrections.
"""

import argparse
import json
import os
import re

from openpyxl import load_workbook


def normalize_text(text):
    if text is None:
        return ""
    s = str(text)
    if s.lower() == "nan":
        return ""
    s = s.replace(" ", " ").replace("​", " ")
    return s.strip()


def collapse_repetition_zh(text):
    """Remove repeated character patterns common in machine translation artifacts."""
    s = normalize_text(text)
    if not s:
        return s
    for _ in range(6):
        before = s
        s = re.sub(r"([一-鿿]{1,4})\1{2,}", r"\1", s)
        s = re.sub(r"(\b[一-鿿]{1,6}\b)(?:\s+\1){2,}", r"\1", s)
        if s == before:
            break
    s = re.sub(r"\s{2,}", " ", s).strip()
    return s


def cleanup_symbols(text):
    """Fix common symbol artifacts in machine translations."""
    s = normalize_text(text)
    if not s:
        return s
    s = s.replace("—", "-").replace("–", "-")
    s = re.sub(r"\s*-\s*-\s*", " - ", s)
    s = re.sub(r"\s*--\s*", " - ", s)
    s = s.replace(" ,", ",").replace(" .", ".")
    s = s.replace(",和(或)", "，和/或")
    s = s.replace("和(或)", "和/或")
    s = s.replace("（", "(").replace("）", ")")
    s = s.replace("“", '"').replace("”", '"')
    s = s.strip("`")
    s = s.replace("]", "】").replace("[", "【")
    s = re.sub(r"【\s*】", "【】", s)
    s = re.sub(r"【\s+", "【", s)
    s = re.sub(r"\s+】", "】", s)
    return s.strip()


def fix_known_terms(en, zh, pitfalls=None):
    """Apply domain-specific term corrections based on source text.
    pitfalls: optional list of {"en_pattern", "wrong_zh", "correct_zh"} dicts."""
    en_s = normalize_text(en)
    zh_s = normalize_text(zh)

    if not en_s:
        return zh_s

    # Built-in defaults
    if "dredg" in en_s.lower():
        zh_s = zh_s.replace("减少", "疏浚")

    if "gate complex" in en_s.lower():
        zh_s = zh_s.replace("大门", "闸门")
        zh_s = zh_s.replace("门复杂", "闸门综合")
        zh_s = zh_s.replace("复杂", "综合")

    if "quay" in en_s.lower():
        zh_s = zh_s.replace("填报", "码头")
        zh_s = zh_s.replace("水箱", "码头")

    if "pile" in en_s.lower():
        zh_s = zh_s.replace("切片", "桩")
        zh_s = zh_s.replace("管道", "桩基")
        zh_s = zh_s.replace("孔径", "桩")

    if "revetment" in en_s.lower():
        zh_s = zh_s.replace("重审", "护岸")
        zh_s = zh_s.replace("翻新", "护岸")

    zh_s = zh_s.replace("招募", "吹填")
    zh_s = zh_s.replace("配偶体体", "材料")
    zh_s = zh_s.replace("标称口袋", "泊位港池")

    # Apply external pitfalls
    if pitfalls:
        for p in pitfalls:
            if p.get("en_pattern", "").lower() in en_s.lower():
                wrong = p.get("wrong_zh", "")
                correct = p.get("correct_zh", "")
                if wrong and correct and wrong in zh_s:
                    zh_s = zh_s.replace(wrong, correct)

    return zh_s


def apply_glossary(zh, glossary):
    """Apply batch term replacement from a glossary dict {en_term: zh_term}."""
    if not glossary or not zh:
        return zh
    result = zh
    # Sort by term length descending to avoid partial replacements
    for en_term, zh_term in sorted(glossary.items(), key=lambda x: len(x[0]), reverse=True):
        if en_term in result:
            result = result.replace(en_term, zh_term)
    return result


def post_polish(en, zh, pitfalls=None, glossary=None):
    """Run the full post-processing pipeline on a (source, translation) pair."""
    if not normalize_text(en):
        return ""
    if not normalize_text(zh):
        return zh
    s = fix_known_terms(en, zh, pitfalls=pitfalls)
    if glossary:
        s = apply_glossary(s, glossary)
    s = cleanup_symbols(s)
    s = collapse_repetition_zh(s)
    return s


def main():
    parser = argparse.ArgumentParser(
        description="Polish existing Chinese translations in an Excel file"
    )
    parser.add_argument("excel_path", help="Path to Excel file with translations")
    parser.add_argument("--src-col", type=int, default=1, help="Source column index (0-based)")
    parser.add_argument("--tgt-col", type=int, default=2, help="Target column index (0-based)")
    parser.add_argument("--sheet", type=str, default=None, help="Worksheet name")
    parser.add_argument("--output", type=str, default=None, help="Output path")
    parser.add_argument("--pitfalls", type=str, default=None,
                        help="Path to JSON pitfalls file [{en_pattern, wrong_zh, correct_zh}]")
    parser.add_argument("--glossary", type=str, default=None,
                        help="Path to JSON glossary file {en_term: zh_term}")
    args = parser.parse_args()

    pitfalls = None
    glossary = None
    if args.pitfalls and os.path.exists(args.pitfalls):
        with open(args.pitfalls, 'r', encoding='utf-8') as f:
            pitfalls = json.load(f)
        print(f"pitfalls_loaded={len(pitfalls)}")
    if args.glossary and os.path.exists(args.glossary):
        with open(args.glossary, 'r', encoding='utf-8') as f:
            glossary = json.load(f)
        print(f"glossary_loaded={len(glossary)}")

    if not os.path.exists(args.excel_path):
        raise FileNotFoundError(args.excel_path)

    out_path = args.output
    if not out_path:
        base, ext = os.path.splitext(args.excel_path)
        out_path = f"{base}_polished{ext}"

    wb = load_workbook(args.excel_path)
    ws = wb[args.sheet] if args.sheet else wb.active
    max_row = ws.max_row

    fixed_count = 0
    for r in range(2, max_row + 1):
        en = ws.cell(row=r, column=args.src_col + 1).value
        zh = ws.cell(row=r, column=args.tgt_col + 1).value
        if zh:
            polished = post_polish(en, zh, pitfalls=pitfalls, glossary=glossary)
            if polished != normalize_text(zh):
                ws.cell(row=r, column=args.tgt_col + 1).value = polished
                fixed_count += 1

    wb.save(out_path)
    print(out_path)
    print(f"rows_total={max_row - 1}")
    print(f"rows_fixed={fixed_count}")


if __name__ == "__main__":
    main()
